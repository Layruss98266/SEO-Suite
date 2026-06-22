"""Report listing, preview, download, delete, and PDF export routes.

Owns ``/api/reports*``, ``/api/open/<file>``, ``/api/download/<file>``,
``/api/history``, and ``/api/reports/pdf/<file>``.

Path traversal is enforced via :func:`app.state._safe_report_path` for every
filesystem touch. PDF generation forks a Playwright chromium process per call;
concurrency is capped by :data:`app.state._pdf_semaphore` so a burst of
requests cannot OOM the host.
"""

from __future__ import annotations

import csv as _csv
import json
import logging
import re
from datetime import datetime

from flask import Blueprint, jsonify, request, send_file

from app.blueprints import api_error
from app.state import (
    REPORTS_DIR,
    _REPORT_STEM_RE,
    _pdf_semaphore,
    _safe_report_path,
)
from core.auth import login_required
from core.checker import load_history

logger = logging.getLogger(__name__)

bp = Blueprint("reports", __name__)


# ── Internal: stem-based deletion guard ───────────────────────────────────────

def _delete_report_stem(stem: str) -> list[str]:
    """Delete every file in REPORTS_DIR that shares *stem*, regardless of extension.

    Only touches files whose stem matches a known report pattern, so unrelated
    files in the same directory are never affected even if a caller passes a
    crafted name.
    """
    if not _REPORT_STEM_RE.match(stem):
        return []
    removed = []
    for p in REPORTS_DIR.glob(f"{stem}.*"):
        try:
            p.unlink()
            removed.append(p.name)
        except Exception as exc:
            logger.warning("Could not delete %s: %s", p.name, exc)
    return removed


# ── Listing ───────────────────────────────────────────────────────────────────
@bp.route("/api/reports")
@login_required
def api_reports():
    # Indexing reports: csv files (with paired html)
    csv_files = sorted(REPORTS_DIR.glob("indexing_report_*.csv"), reverse=True)[:20]

    # Audit reports: prefer html files, fall back to json-only sidecars.
    # html_bases must cover ALL html files (not just the display cap) so that
    # reports beyond position 20 don't reappear as phantom JSON-only entries.
    all_audit_html = sorted(REPORTS_DIR.glob("seo_audit_*.html"), reverse=True)
    html_bases = {f.stem for f in all_audit_html}
    audit_html_files = all_audit_html[:20]
    audit_json_only = sorted(
        [f for f in REPORTS_DIR.glob("seo_audit_*.json") if f.stem not in html_bases],
        reverse=True,
    )[:20]

    def _size_fmt(n: int) -> str:
        return f"{n // 1024} KB" if n > 1024 else f"{n} B"

    result = []
    for f in csv_files:
        stat = f.stat()
        html_ver = f.with_suffix(".html")
        result.append(
            {
                "name": f.name,
                "html_name": html_ver.name if html_ver.exists() else f.name,
                "date": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "size": _size_fmt(stat.st_size),
                "kind": "indexing",
            }
        )

    def _read_avg_score(jsn_path):
        try:
            return json.loads(jsn_path.read_text(encoding="utf-8")).get("avg_score")
        except Exception:
            return None

    for f in audit_html_files:
        stat = f.stat()
        xlsx = f.with_suffix(".xlsx")
        jsn = f.with_suffix(".json")
        result.append(
            {
                "name": f.name,
                "html_name": f.name,
                "xlsx_name": xlsx.name if xlsx.exists() else None,
                "json_name": jsn.name if jsn.exists() else None,
                "score": _read_avg_score(jsn) if jsn.exists() else None,
                "date": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "size": _size_fmt(stat.st_size),
                "kind": "audit",
            }
        )
    for f in audit_json_only:
        stat = f.stat()
        xlsx = f.with_suffix(".xlsx")
        result.append(
            {
                "name": f.name,
                "html_name": None,
                "xlsx_name": xlsx.name if xlsx.exists() else None,
                "json_name": f.name,
                "score": _read_avg_score(f),
                "date": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "size": _size_fmt(stat.st_size),
                "kind": "audit",
            }
        )
    return jsonify(result)


# ── Open + download ───────────────────────────────────────────────────────────
@bp.route("/api/open/<filename>")
@login_required
def api_open(filename):
    """Serve a report HTML inline (for in-browser viewing)."""
    if not filename.endswith(".html"):
        return "Only HTML reports can be opened inline", 400
    base = filename
    safe = _safe_report_path(base, (".html",))
    if safe is None or not safe.is_file():
        return "Report not found", 404
    from flask import make_response
    response = make_response(safe.read_text(encoding="utf-8"), 200)
    response.headers["Content-Type"] = "text/html"
    # Sandbox the report iframe.  Explicitly omit `allow-same-origin` so
    # the document is treated as an opaque origin — cannot read cookies,
    # localStorage, or call back to the app.  `allow-downloads` keeps the
    # Save-As menu working for embedded report exports.
    response.headers["Content-Security-Policy"] = "sandbox allow-downloads"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@bp.route("/api/download/<filename>")
@login_required
def api_download(filename):
    """Download a report file (XLSX, CSV, HTML, JSON, or PDF)."""
    safe = _safe_report_path(filename, (".xlsx", ".csv", ".html", ".json", ".pdf"))
    if safe is None or not safe.is_file():
        return "Not found", 404
    return send_file(str(safe), as_attachment=True, download_name=safe.name)


# ── Delete ────────────────────────────────────────────────────────────────────
@bp.route("/api/reports/delete/<filename>", methods=["DELETE"])
@login_required
def api_reports_delete(filename):
    if not re.match(r"^[\w\-\.]+$", filename):
        return api_error("Invalid filename", 400)
    stem = filename.rsplit(".", 1)[0]
    deleted = _delete_report_stem(stem)
    if not deleted:
        return api_error("File not found", 404)
    return jsonify({"deleted": deleted})


@bp.route("/api/reports/delete_bulk", methods=["POST"])
@login_required
def api_reports_delete_bulk():
    """Delete multiple reports in one call. Body: ``{names: [filename, ...]}``."""
    data = request.get_json(force=True) or {}
    names = data.get("names", [])
    if not isinstance(names, list) or not names:
        return api_error("names list required", 400)

    deleted, failed = [], []
    for name in names:
        if not isinstance(name, str) or not re.match(r"^[\w\-\.]+$", name):
            failed.append({"name": name, "error": "invalid filename"})
            continue
        removed = _delete_report_stem(name.rsplit(".", 1)[0])
        if removed:
            deleted.append(name)
        else:
            failed.append({"name": name, "error": "not found"})
    return jsonify(
        {
            "ok": not bool(failed),
            "deleted": deleted,
            "failed": failed,
            "deleted_count": len(deleted),
            "failed_count": len(failed),
        }
    )


@bp.route("/api/reports/delete_all", methods=["POST"])
@login_required
def api_reports_delete_all():
    """Nuclear option — wipes all reports. Body must include ``{confirm: 'YES'}``."""
    data = request.get_json(force=True) or {}
    if data.get("confirm") != "YES":
        return api_error("confirm:'YES' required", 400)
    deleted = 0
    # Per-stem deletion enforces the pattern guard for every group.
    seen = set()
    for p in list(REPORTS_DIR.glob("indexing_report_*.*")) + list(
        REPORTS_DIR.glob("seo_audit_*.*")
    ):
        if p.stem not in seen:
            seen.add(p.stem)
            deleted += len(_delete_report_stem(p.stem))
    return jsonify({"deleted_count": deleted})


# ── Summary + preview ─────────────────────────────────────────────────────────
@bp.route("/api/reports/summary")
@login_required
def api_reports_summary():
    """Return avg score + total issues — accepts .json sidecar or .xlsx."""
    name = request.args.get("file", "")
    p = _safe_report_path(name, (".json", ".xlsx"))
    if p is None:
        return api_error("invalid filename", 400)
    if not p.exists():
        return api_error("not found", 404)
    try:
        if name.endswith(".json"):
            data = json.loads(p.read_text(encoding="utf-8"))
            return jsonify(data)
        import openpyxl as _ox

        wb = _ox.load_workbook(p, read_only=True, data_only=True)
        ws = wb["Summary"]
        scores, issues, warns = [], 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[1] is None:
                continue
            if row[2] is not None:
                scores.append(row[2])
            warns += row[4] or 0
            issues += row[5] or 0
        avg = round(sum(scores) / len(scores)) if scores else 0
        return jsonify(
            {
                "urls": len(scores),
                "avg_score": avg,
                "total_issues": issues,
                "total_warnings": warns,
            }
        )
    except Exception as e:
        logger.error("Summary error: %s", e, exc_info=True)
        return api_error("An internal error occurred", 500)


@bp.route("/api/reports/preview/<filename>")
@login_required
def api_reports_preview(filename):
    """Rich preview data for the side drawer — works for both indexing + audit reports."""
    # Route through _safe_report_path for resolve()+relative_to() traversal guard.
    if _safe_report_path(filename, (".csv", ".html", ".json", ".xlsx")) is None:
        return api_error("Invalid filename", 400)

    base = filename.rsplit(".", 1)[0]

    # ── Indexing report (CSV) ─────────────────────────────────────────────
    csv_path = REPORTS_DIR / (base + ".csv")
    if csv_path.exists() or filename.endswith(".csv"):
        p = REPORTS_DIR / filename if filename.endswith(".csv") else csv_path
        if not p.exists():
            return api_error("not found", 404)
        try:
            rows = []
            with open(p, encoding="utf-8") as f:
                reader = _csv.DictReader(f)
                for row in reader:
                    rows.append(row)
            total = len(rows)
            indexed = sum(1 for r in rows if r.get("Status", "").strip() == "Indexed")
            not_indexed = sum(
                1 for r in rows if r.get("Status", "").strip() == "Not Indexed"
            )
            errors = total - indexed - not_indexed
            rate = round(indexed / total * 100, 1) if total else 0
            ni_urls = [
                r.get("URL", "")
                for r in rows
                if r.get("Status", "").strip() == "Not Indexed"
            ][:10]
            return jsonify(
                {
                    "kind": "indexing",
                    "total": total,
                    "indexed": indexed,
                    "not_indexed": not_indexed,
                    "errors": errors,
                    "rate": rate,
                    "not_indexed_sample": ni_urls,
                }
            )
        except Exception as e:
            logger.error("Preview indexing error: %s", e, exc_info=True)
            return api_error("An internal error occurred", 500)

    # ── Audit report (JSON sidecar preferred, xlsx fallback) ─────────────
    json_path = REPORTS_DIR / (base + ".json")
    xlsx_path = REPORTS_DIR / (base + ".xlsx")
    if json_path.exists():
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
            urls_data = data.get("urls", [])

            if isinstance(urls_data, list) and urls_data:
                scores = [u.get("score", 0) for u in urls_data if u.get("score") is not None]
                avg = round(sum(scores) / len(scores)) if scores else data.get("avg_score", 0)
                excellent = sum(1 for s in scores if s >= 80)
                good = sum(1 for s in scores if 50 <= s < 80)
                poor = sum(1 for s in scores if s < 50)
                issue_counts: dict = {}
                for u in urls_data:
                    for chk in u.get("results", []):
                        if chk.get("status") in ("fail", "warning"):
                            key = chk.get("label") or chk.get("tool", "unknown")
                            issue_counts[key] = issue_counts.get(key, 0) + 1
                top_issues = sorted(issue_counts.items(), key=lambda x: -x[1])[:8]
            else:
                # Summary-only sidecar: {avg_score, total_issues, total_warnings, urls(int)}
                n_urls = urls_data if isinstance(urls_data, int) else data.get("urls", 0)
                scores = []
                avg = data.get("avg_score", 0)
                excellent, good, poor = 0, 0, n_urls
                top_issues = []

            return jsonify(
                {
                    "kind": "audit",
                    "urls": len(scores)
                    if scores
                    else (
                        data.get("urls", 0)
                        if not isinstance(data.get("urls"), list)
                        else len(urls_data)
                    ),
                    "avg_score": avg,
                    "total_issues": data.get("total_issues", 0),
                    "total_warnings": data.get("total_warnings", 0),
                    "score_dist": {"excellent": excellent, "good": good, "poor": poor},
                    "top_issues": [{"label": k, "count": v} for k, v in top_issues],
                }
            )
        except Exception as e:
            logger.error("Preview audit JSON error: %s", e, exc_info=True)
            return api_error("An internal error occurred", 500)

    if xlsx_path.exists():
        try:
            import openpyxl as _ox

            wb = _ox.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb["Summary"]
            scores, issues, warns = [], 0, 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[1] is None:
                    continue
                if row[2] is not None:
                    scores.append(row[2])
                warns += row[4] or 0
                issues += row[5] or 0
            avg = round(sum(scores) / len(scores)) if scores else 0
            excellent = sum(1 for s in scores if s >= 80)
            good = sum(1 for s in scores if 50 <= s < 80)
            poor = sum(1 for s in scores if s < 50)
            return jsonify(
                {
                    "kind": "audit",
                    "urls": len(scores),
                    "avg_score": avg,
                    "total_issues": issues,
                    "total_warnings": warns,
                    "score_dist": {"excellent": excellent, "good": good, "poor": poor},
                    "top_issues": [],
                }
            )
        except Exception as e:
            logger.error("Preview audit XLSX error: %s", e, exc_info=True)
            return api_error("An internal error occurred", 500)

    return api_error("Report data not found", 404)


# ── History ───────────────────────────────────────────────────────────────────
@bp.route("/api/history")
@login_required
def api_history():
    return jsonify(load_history())


# ── PDF export ────────────────────────────────────────────────────────────────
@bp.route("/api/reports/pdf/<filename>")
@login_required
def api_reports_pdf(filename):
    """Render a report HTML to PDF via Playwright Chromium.

    Concurrency cap (``_pdf_semaphore``) ensures a burst of requests doesn't
    fork unlimited chromium processes and OOM the host. Non-blocking acquire:
    callers see a fast 503 instead of piling up worker threads.
    """
    base = filename.rsplit(".", 1)[0] + ".html"
    html_path = _safe_report_path(base, (".html",))
    if html_path is None or not html_path.is_file():
        return "Report not found", 404
    stem = html_path.stem
    if not _pdf_semaphore.acquire(blocking=False):
        return api_error("PDF generation busy — retry shortly", 503)
    try:
        from playwright.sync_api import sync_playwright

        abs_uri = html_path.resolve().as_uri()
        with sync_playwright() as pw:
            # SECURITY (S15): Do NOT add --no-sandbox here. Reports contain
            # user-influenced HTML (audit results), so the Chromium renderer
            # sandbox is the last line of defense against a renderer 0-day.
            # The container running this service MUST run as a non-root user
            # (e.g. `pwuser` in the official Playwright image); root + sandbox
            # is what forces people to reach for --no-sandbox.
            browser = pw.chromium.launch(args=["--disable-javascript"])
            page = browser.new_page()
            page.goto(abs_uri, wait_until="networkidle", timeout=30000)
            pdf_bytes = page.pdf(format="A4", print_background=True)
            browser.close()
        return pdf_bytes, 200, {
            "Content-Type": "application/pdf",
            "Content-Disposition": f"attachment; filename={stem}.pdf",
        }
    except Exception as e:
        logger.error("PDF generation failed: %s", e, exc_info=True)
        return api_error("An internal error occurred", 500)
    finally:
        _pdf_semaphore.release()


# ── Public exports for other blueprints that need the stem deleter ────────────
__all__ = ["bp", "register", "_delete_report_stem"]


def register(app) -> None:
    app.register_blueprint(bp)
