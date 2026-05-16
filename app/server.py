"""
SEO Suite — unified route file (monolith, ~2 200 lines).

MIGRATION PLAN: routes are being moved to app/routes/ blueprints incrementally.
See app/routes/__init__.py for the blueprint registry and migration status.
Shared mutable state lives in app/state.py — import from there, not here.

Current shared-state objects re-exported from app/state for backwards compat:
    CFG, index_status, audit_status, index_queue, audit_queue,
    index_subscribers, audit_subscribers, state_lock, index_cancel,
    audit_cancel, index_paused, audit_paused, last_index_run,
    audit_partial, audit_full_results, REPORTS_DIR, DATA_DIR, CONFIG_PATH
"""

import json, csv, threading, queue, time, logging, os, re
from pathlib import Path
from datetime import datetime
from urllib.parse import urlparse
from flask import Flask, Response, request, jsonify, send_from_directory, session
from flask_cors import CORS
from werkzeug.utils import secure_filename
from core.version import VERSION

from core.checker import (
    fetch_sitemap_urls, fetch_from_domain, load_from_csv_excel,
    filter_urls, find_latest_report, compare_runs, load_history,
    execute_and_save, load_config, build_gsc_service, crawl_site
)
from core.seo_audit import audit_single_url, generate_html_report, generate_excel_report
from core.security import is_safe_url, validate_public_url
from core.auth import (
    init_auth, auth_enabled, verify_credentials, login_required, is_authed, LOGIN_PAGE,
)

# Cap on per-run audit result lists so a huge sitemap can't blow up memory.
# Past this point new results are dropped from the live progress feed; the
# completed report still gets every URL written to disk.
MAX_AUDIT_RESULTS = 5000


def _require_public_url(value: str, field_name: str = "url"):
    """Validate `value` as a public URL. Returns (url, None) on success or
    (None, (response, status)) on failure — ready to be `return`ed from a route.
    """
    try:
        return validate_public_url(value), None
    except ValueError as exc:
        return None, (jsonify({"error": f"Invalid {field_name}: {exc}"}), 400)


def _safe_public_url_list(raw: str) -> list[str]:
    """Split a comma-separated URL list and drop any URL that fails SSRF check."""
    urls = []
    for part in raw.split(","):
        candidate = part.strip()
        if not candidate:
            continue
        try:
            urls.append(validate_public_url(candidate))
        except ValueError as exc:
            logger.warning("Blocked URL in list %s: %s", candidate, exc)
    return urls


def _norm_url(url: str) -> str:
    """Prepend https:// if the user omitted a scheme (e.g. 'example.com')."""
    if url and not url.startswith(("http://", "https://")):
        return f"https://{url}"
    return url


def _reject_unsafe(url: str):
    """Return a 400 JSON response if *url* fails the SSRF safety check, else None.

    The tool endpoints all accept a URL from the request body and fetch it
    server-side. Without this guard, anyone reaching the server can pivot
    requests to internal services (cloud metadata, the dashboard's own
    delete_all endpoint, etc.). See core.security.is_safe_url for the policy.
    """
    ok, reason = is_safe_url(url)
    if ok:
        return None
    return jsonify({"ok": False, "error": f"URL refused: {reason}"}), 400


def _int(data: dict, key: str, default: int, lo: int, hi: int) -> int:
    """Safely parse an integer from a request-data dict, clamped to [lo, hi].

    Returns *default* (clamped) when the key is absent, None, or not
    convertible to int — preventing bare ``int()`` calls from raising
    ValueError / TypeError and producing HTTP 500 responses.
    """
    try:
        return max(lo, min(hi, int(data.get(key, default))))
    except (TypeError, ValueError):
        return max(lo, min(hi, default))


TEMPLATE_DIR = Path(__file__).parent / "templates"
STATIC_DIR   = Path(__file__).parent / "static"

# ── Logging ───────────────────────────────────────────────────────────────────
Path("data").mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("data/app.log", encoding="utf-8"),
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__, template_folder=str(TEMPLATE_DIR), static_folder=str(STATIC_DIR))
init_auth(app)
# CORS origins are env-configurable so production deployments don't need a code
# change. Default is 8080 (the canonical port) — set CORS_ALLOWED_ORIGINS if
# you front the app on a different port.
_cors_origins = [
    o.strip()
    for o in os.getenv(
        "CORS_ALLOWED_ORIGINS",
        "http://localhost:8080,http://127.0.0.1:8080",
    ).split(",")
    if o.strip()
]
CORS(app, origins=_cors_origins, supports_credentials=True)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB upload cap

# ── Sentry error tracking (Stage 1-D) ────────────────────────────────────────
# Opt-in: set SENTRY_DSN env var to activate. No-ops silently when absent so
# local / air-gapped installs are unaffected.
_sentry_dsn = os.getenv("SENTRY_DSN", "")
if _sentry_dsn:
    import sentry_sdk
    from sentry_sdk.integrations.flask import FlaskIntegration
    sentry_sdk.init(
        dsn=_sentry_dsn,
        integrations=[FlaskIntegration()],
        traces_sample_rate=0.05,           # capture 5 % of requests for perf
        environment=os.getenv("SEO_SUITE_ENV", "development"),
    )
    logger.info("Sentry error tracking enabled (env=%s)", os.getenv("SEO_SUITE_ENV", "development"))

# ── Rate limiting (Stage 1-B) ─────────────────────────────────────────────────
# In-memory store — no Redis dependency needed for single-server deployments.
# Limits apply per remote IP address.
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[],           # no global default — limits are route-specific
    storage_uri="memory://",
)

# Anchor every filesystem path to the project root so the server behaves the
# same regardless of cwd. Previously `Path("config.json")` could write to the
# wrong directory when launched outside the repo root.
PROJECT_ROOT = Path(__file__).parent.parent.resolve()
CONFIG_PATH  = PROJECT_ROOT / "config.json"
CFG = load_config()

DATA_DIR    = PROJECT_ROOT / "data";    DATA_DIR.mkdir(exist_ok=True)
REPORTS_DIR = DATA_DIR / "reports";     REPORTS_DIR.mkdir(exist_ok=True)

# PDF generation spawns Playwright per request — cap concurrency so a burst of
# /api/reports/pdf calls can't OOM the host with parallel chromium processes.
_PDF_CONCURRENCY = max(1, int(os.environ.get("SEO_SUITE_PDF_WORKERS", "2")))
_pdf_semaphore   = threading.Semaphore(_PDF_CONCURRENCY)


def _safe_report_path(filename: str, allowed_exts: tuple[str, ...]) -> Path | None:
    """Validate that *filename* resolves inside REPORTS_DIR and has an allowed extension.
    Returns the resolved Path or None. Used to block path traversal in /api/open,
    /api/download, /api/reports/pdf."""
    if not isinstance(filename, str) or not filename:
        return None
    # Reject anything that looks like a separator or traversal segment up front.
    if "/" in filename or "\\" in filename or ".." in filename or filename.startswith("."):
        return None
    if not re.match(r"^[\w\-]+(?:\.[\w\-]+)*$", filename):
        return None
    if allowed_exts and not filename.lower().endswith(allowed_exts):
        return None
    try:
        candidate = (REPORTS_DIR / filename).resolve()
        candidate.relative_to(REPORTS_DIR.resolve())
    except (ValueError, OSError):
        return None
    return candidate


def _safe_upload_path(raw: str) -> Path | None:
    """Validate that *raw* points to a file inside data/uploads/.
    Returns the resolved Path if safe, else None. Used to prevent a client from
    making the audit/index handlers read arbitrary filesystem paths via the
    csv/xlsx input mode."""
    try:
        uploads = (DATA_DIR / "uploads").resolve()
        target  = Path(raw).resolve()
        target.relative_to(uploads)
        return target if target.is_file() else None
    except (ValueError, OSError):
        return None

@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "File too large — 10 MB maximum"}), 413

# ── Shared state ──────────────────────────────────────────────────────────────
# SSE subscribers — each connected client gets its own bounded queue so multiple
# browser tabs (or a tab that reloads mid-run) all see every progress event.
# The previous single-Queue design let the first subscriber drain the queue,
# leaving later/reconnected clients stuck on the last seen state.
_index_subscribers: list[queue.Queue] = []
_audit_subscribers: list[queue.Queue] = []
_sub_lock = threading.Lock()

# Legacy aliases — internal cancel/error paths still call `_index_queue.put(...)`.
# We keep them as thin adapters that broadcast to all subscribers instead.
class _BroadcastQueue:
    def __init__(self, subs: list[queue.Queue]):
        self._subs = subs
    def put(self, msg):
        with _sub_lock:
            dead = []
            for q in self._subs:
                try:
                    q.put_nowait(msg)
                except queue.Full:
                    dead.append(q)
            for q in dead:
                try: self._subs.remove(q)
                except ValueError: pass

_index_queue      = _BroadcastQueue(_index_subscribers)
_audit_queue      = _BroadcastQueue(_audit_subscribers)


def _broadcast_index(msg): _index_queue.put(msg)
def _broadcast_audit(msg): _audit_queue.put(msg)


def _subscribe(subs: list[queue.Queue]) -> queue.Queue:
    q = queue.Queue(maxsize=1000)
    with _sub_lock:
        subs.append(q)
    return q

def _unsubscribe(subs: list[queue.Queue], q: queue.Queue) -> None:
    with _sub_lock:
        try: subs.remove(q)
        except ValueError: pass


def _cleanup_subscribers() -> None:
    """Background thread: periodically drop fully-queued (stale) SSE queues.

    A subscriber that has disconnected without calling /api/index/stream
    cleanup leaves a full queue behind.  Without this sweep those queues
    accumulate indefinitely, slowly leaking memory.  We remove any queue
    whose buffer is already at capacity — a live consumer would have drained
    it.  Runs every 5 minutes; daemon so it doesn't block process exit.
    """
    while True:
        time.sleep(300)
        with _sub_lock:
            for subs in (_index_subscribers, _audit_subscribers):
                subs[:] = [q for q in subs if not q.full()]


threading.Thread(target=_cleanup_subscribers, daemon=True, name="sse-cleanup").start()

_index_status     = {"running": False, "total": 0, "done": 0}
_audit_status     = {"running": False, "total": 0, "done": 0}
_audit_cancel     = threading.Event()
_index_cancel     = threading.Event()
_lock             = threading.Lock()
_last_index_run   = {}   # {url: status_string} — updated live during a run
_index_paused     = threading.Event()   # set = running, clear = paused
_index_paused.set()                     # start in "running" (not paused) state
_audit_paused     = threading.Event()   # mirrors _index_paused for audit
_audit_paused.set()
_audit_partial    = []                  # slim per-URL dicts for mid-run CSV export
_audit_full_results = []               # full audit dicts for cancel partial report

ERROR_STATUSES = {"Error", "Timeout", "Other"}
ERROR_PREFIXES = ("GSC Error", "Error:", "Error ", "Browser Error", "Playwright Error")


def _snapshot_last_index_run() -> dict[str, str]:
    with _lock:
        return dict(_last_index_run)


def _reset_last_index_run() -> None:
    with _lock:
        _last_index_run.clear()


def _update_last_index_run(url: str, status: str) -> None:
    with _lock:
        _last_index_run[url] = status


def _replace_last_index_run(results: dict[str, str]) -> None:
    with _lock:
        _last_index_run.clear()
        _last_index_run.update(results)






# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Indexing
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
@login_required
def index():
    return (TEMPLATE_DIR / "dashboard.html").read_text(encoding="utf-8"), 200, {"Content-Type": "text/html"}


# ── Auth routes ──────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
@limiter.limit("20 per minute")
def login():
    if not auth_enabled():
        # Auth disabled (no SEO_SUITE_PASSWORD_HASH env). Send users straight in.
        return ("", 302, {"Location": "/"})
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        if verify_credentials(username, password):
            session.clear()
            session["authed"] = True
            session.permanent = True
            return ("", 302, {"Location": "/"})
        page = LOGIN_PAGE.replace("__ERROR__", "<div class='err'>Invalid credentials</div>")
        return page, 401, {"Content-Type": "text/html"}
    page = LOGIN_PAGE.replace("__ERROR__", "")
    return page, 200, {"Content-Type": "text/html"}


@app.route("/logout", methods=["POST", "GET"])
def logout():
    session.clear()
    return ("", 302, {"Location": "/login"})


@app.route("/api/index/run", methods=["POST"])
@login_required
@limiter.limit("10 per hour")
def api_index_run():
    global _index_status
    # Quick reject without holding the lock — final atomic check happens below.
    if _index_status["running"]:
        return jsonify({"error": "Already running"}), 400

    try:
        data = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"error": "Invalid JSON in request body"}), 400

    input_type = data.get("input_type", "sitemap")
    raw        = data.get("input", "").strip()
    if not raw:
        return jsonify({"error": "No URL or sitemap provided"}), 400
    # SSRF guard for URL-bearing input modes. CSV/list inputs are filesystem or
    # multi-URL — those URLs are checked individually downstream by Playwright,
    # which targets google.com (not the user-supplied host).
    if input_type in ("sitemap", "domain"):
        raw = _norm_url(raw)
        ok, reason = is_safe_url(raw)
        if not ok:
            return jsonify({"error": f"URL refused: {reason}"}), 400
    pattern    = data.get("pattern", "")
    # Clamp limit to [1, 500] so a client can't request a 999999-URL run that
    # ties up Playwright workers and disk.
    limit      = _int(data, "limit", 20, 1, 500)
    quiet      = data.get("quiet", False)
    headless   = data.get("headless", False)
    do_compare = data.get("compare", False)

    estimated_total = limit
    # Atomic check+set so a concurrent POST that snuck past the early check
    # can't also flip running=True and spawn a second worker thread.
    with _lock:
        if _index_status["running"]:
            return jsonify({"error": "Already running"}), 400
        _index_status = {"running": True, "total": estimated_total, "done": 0}

    def run():
        global _last_index_run
        try:
            # Fetch URLs inside the thread — keeps the HTTP response fast
            if input_type == "sitemap":
                urls = fetch_sitemap_urls(raw)
            elif input_type == "domain":
                urls = fetch_from_domain(raw)
            elif input_type == "csv":
                safe = _safe_upload_path(raw)
                if safe is None:
                    _index_queue.put({"type": "error",
                                      "message": "CSV path must point to an uploaded file in data/uploads/"})
                    return
                urls = load_from_csv_excel(str(safe))
            elif input_type in ("paste", "list"):
                # Comma-separated list of URLs pasted by the user.
                # _safe_public_url_list drops anything that fails SSRF check.
                urls = _safe_public_url_list(raw)
            else:  # 'multi' — treat as list of sitemaps OR URLs
                parts = [s.strip() for s in raw.split(",") if s.strip()]
                urls = []
                for p in parts:
                    if p.lower().endswith(".xml") or "sitemap" in p.lower():
                        urls += fetch_sitemap_urls(p)
                    elif p.startswith("http"):
                        urls.append(p)

            urls = filter_urls(urls, pattern)[:limit]
            if not urls:
                _index_queue.put({"type": "error", "message": "No URLs found — check your sitemap URL or filter pattern"})
                return

            with _lock:
                _index_status["total"] = len(urls)
            _reset_last_index_run()

            _index_cancel.clear()
            _index_paused.set()   # ensure not paused at start of run
            prev = find_latest_report() if do_compare else None
            _run_results: dict[str, str] = {}

            def cb(num, total, url, status):
                _index_paused.wait()   # block here while paused
                if _index_cancel.is_set():
                    return
                _run_results[url] = status
                _update_last_index_run(url, status)   # live update for cancel/partial export
                from core.checker import get_priority_score, get_crawl_depth, get_url_type
                _index_queue.put({"type": "progress", "num": num, "total": total, "url": url,
                                  "status": status, "priority": get_priority_score(url),
                                  "depth": get_crawl_depth(url), "url_type": get_url_type(url)})
                with _lock:
                    _index_status["done"] = num

            html = execute_and_save(urls, headless=headless, quiet=quiet,
                                    do_compare=do_compare, prev_report=prev, progress_cb=cb)
            _replace_last_index_run(_run_results)
            error_count = sum(1 for s in _run_results.values() if _is_error_status(s))
            _index_queue.put({"type": "done", "report": str(html), "error_count": error_count})
        except Exception as e:
            logger.error("Indexing thread error: %s", e, exc_info=True)
            _index_queue.put({"type": "error", "message": str(e)})
        finally:
            with _lock:
                _index_status["running"] = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"total": estimated_total, "started": True})


def _is_error_status(status: str) -> bool:
    """True if a per-URL status string indicates a failure to verify indexing.
    Covers the exact tokens 'Error', 'Timeout', 'Other' and also 'GSC Error: …'
    and 'Error: …' prefixes returned by the checker on exceptions.
    """
    if not isinstance(status, str):
        return False
    if status in ERROR_STATUSES:
        return True
    return any(status.startswith(prefix) for prefix in ERROR_PREFIXES)


@app.route("/api/index/stream")
@login_required
def api_index_stream():
    sub = _subscribe(_index_subscribers)
    def gen():
        try:
            while True:
                try:
                    msg = sub.get(timeout=30)
                    yield f"data: {json.dumps(msg)}\n\n"
                    if msg.get("type") in ("done", "error", "cancelled"): break
                except queue.Empty:
                    yield 'data: {"type":"ping"}\n\n'
        finally:
            _unsubscribe(_index_subscribers, sub)
    return Response(gen(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — SEO Audit
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/audit/run", methods=["POST"])
@login_required
@limiter.limit("10 per hour")
def api_audit_run():
    global _audit_status
    if _audit_status["running"]:
        return jsonify({"error":"Already running"}), 400

    try:
        data = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"error": "Invalid JSON in request body"}), 400

    input_type = data.get("input_type","sitemap")
    raw        = data.get("input","").strip()
    if not raw:
        return jsonify({"error": "No URL or sitemap provided"}), 400
    if input_type in ("sitemap", "domain", "crawl"):
        raw = _norm_url(raw)
        ok, reason = is_safe_url(raw)
        if not ok:
            return jsonify({"error": f"URL refused: {reason}"}), 400
    pattern    = data.get("pattern","")
    limit      = _int(data, "limit", 10, 1, 500)
    keywords   = data.get("keywords",[])
    use_cases  = data.get("use_cases", None)
    tasks      = data.get("tasks", None)
    workers    = _int(data, "workers", 3, 1, 8)  # parallel URL audits
    _audit_cancel.clear()

    # Estimate total immediately so the frontend can show a progress bar.
    # URL fetching happens inside the thread so this route returns in <50 ms.
    estimated_total = limit
    # Clear partial buffers BEFORE flipping running=True. Previously the order
    # was reversed: a cancel POST arriving between the running flip and the
    # clears would snapshot an empty list and report an empty partial report.
    with _lock:
        if _audit_status["running"]:
            return jsonify({"error": "Already running"}), 400
        _audit_partial.clear()
        _audit_full_results.clear()
        _audit_status = {"running": True, "total": estimated_total, "done": 0}
    _audit_cancel.clear()
    _audit_paused.set()   # ensure not paused at start of new run

    current_cfg = load_config()
    cfg_with_kw = {**current_cfg, "track_keywords": keywords}

    def run():
        try:
            # Fetch URLs inside the thread — keeps the HTTP response fast
            if input_type == "sitemap":
                urls = fetch_sitemap_urls(raw)
            elif input_type == "domain":
                urls = fetch_from_domain(raw)
            elif input_type == "crawl":
                urls = crawl_site(raw, max_pages=limit, max_depth=int(data.get("crawl_depth", 2)))
            elif input_type in ("paste", "list"):
                urls = _safe_public_url_list(raw)
            else:  # csv / xlsx — raw is the uploaded file path
                safe = _safe_upload_path(raw)
                if safe is None:
                    _audit_queue.put({"type": "error",
                                      "message": "CSV path must point to an uploaded file in data/uploads/"})
                    return
                urls = load_from_csv_excel(str(safe))

            urls = filter_urls(urls, pattern)[:limit]
            if not urls:
                _audit_queue.put({"type": "error", "message": "No URLs found — check your sitemap URL or filter pattern"})
                return

            with _lock:
                _audit_status["total"] = len(urls)

            from core.seo_audit import generate_html_report, generate_excel_report
            from tools.phase3 import audit_site as p3_site

            gsc_service  = build_gsc_service() if current_cfg.get("gsc", {}).get("enabled") else None
            audits       = []
            p3_site_data = []

            if gsc_service and urls:
                _p           = urlparse(urls[0])
                site_url     = f"{_p.scheme}://{_p.netloc}/" if _p.netloc else ""
                p3_site_data = p3_site(gsc_service, site_url, urls[:5])

            from concurrent.futures import ThreadPoolExecutor, as_completed as _ac
            i_counter = {"n": 0}
            def _audit_one(u):
                if _audit_cancel.is_set():
                    return None
                _audit_paused.wait()   # block while paused
                if _audit_cancel.is_set():
                    return None
                return audit_single_url(u, cfg_with_kw, gsc_service,
                                        use_cases=use_cases, tasks=tasks)
            with ThreadPoolExecutor(max_workers=workers) as _ex:
                fut_to_url = {_ex.submit(_audit_one, u): u for u in urls}
                for fut in _ac(fut_to_url):
                    if _audit_cancel.is_set():
                        for f in fut_to_url: f.cancel()
                        break
                    url = fut_to_url[fut]
                    try:
                        audit = fut.result()
                    except Exception as ex:
                        logger.error("audit error %s: %s", url, ex)
                        audit = None
                    if audit is None:
                        continue
                    audits.append(audit)
                    with _lock:
                        # Bounded live state — past MAX_AUDIT_RESULTS new entries
                        # stop accumulating to keep memory flat on huge sitemaps.
                        if len(_audit_partial) < MAX_AUDIT_RESULTS:
                            _audit_partial.append({
                                "url":      url,
                                "score":    audit.get("score", 0),
                                "issues":   len(audit.get("issues", [])),
                                "warnings": len(audit.get("warnings", [])),
                            })
                        if len(_audit_full_results) < MAX_AUDIT_RESULTS:
                            _audit_full_results.append(audit)
                    i_counter["n"] += 1
                    i = i_counter["n"]
                    # Slim per-result payload — drawer needs tool/status/message/value
                    slim_results = [{"tool": r.get("tool"),
                                     "status": r.get("status"),
                                     "message": r.get("message",""),
                                     "value": r.get("value"),
                                     "details": r.get("details") or {}}
                                    for r in audit.get("results", [])]
                    _audit_queue.put({
                        "type": "progress", "num": i, "total": len(urls), "url": url,
                        "score":    audit.get("score", 0),
                        "issues":   len(audit.get("issues", [])),
                        "warnings": len(audit.get("warnings", [])),
                        "results":  slim_results,
                        "counts":   audit.get("counts", {}),
                    })
                    with _lock:
                        _audit_status["done"] = i

            timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
            html_path  = REPORTS_DIR / f"seo_audit_{timestamp}.html"
            excel_path = REPORTS_DIR / f"seo_audit_{timestamp}.xlsx"
            _json_path = REPORTS_DIR / f"seo_audit_{timestamp}.json"

            # Write HTML first — it is the primary file.  JSON and Excel are
            # companions; they must not be written before HTML so a crash between
            # writes cannot leave orphaned sidecars without a matching HTML.
            html_content = generate_html_report(audits, p3_site_data, timestamp)
            html_path.write_text(html_content, encoding="utf-8")

            xlsx_ok = False
            try:
                generate_excel_report(audits, excel_path)
                xlsx_ok = True
            except Exception as _xe:
                logger.error("Excel report failed (non-fatal): %s", _xe)

            try:
                _sidecar = {
                    "avg_score":      round(sum(a["score"] for a in audits) / len(audits)) if audits else 0,
                    "total_issues":   sum(len(a["issues"])   for a in audits),
                    "total_warnings": sum(len(a["warnings"]) for a in audits),
                    "urls":           len(audits),
                }
                _json_path.write_text(json.dumps(_sidecar), encoding="utf-8")
            except Exception as _je:
                logger.warning("JSON sidecar write failed: %s", _je)

            _audit_queue.put({
                "type":  "done",
                "report": str(html_path),
                "xlsx":   str(excel_path) if xlsx_ok else "",
            })
        except Exception as e:
            logger.error("Audit thread error: %s", e, exc_info=True)
            _audit_queue.put({"type": "error", "message": str(e)})
        finally:
            with _lock:
                _audit_status["running"] = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"total": estimated_total, "started": True, "workers": workers})

@app.route("/api/audit/stream")
@login_required
def api_audit_stream():
    sub = _subscribe(_audit_subscribers)
    def gen():
        try:
            while True:
                try:
                    msg = sub.get(timeout=30)
                    yield f"data: {json.dumps(msg)}\n\n"
                    if msg.get("type") in ("done", "error", "cancelled"): break
                except queue.Empty:
                    yield 'data: {"type":"ping"}\n\n'
        finally:
            _unsubscribe(_audit_subscribers, sub)
    return Response(gen(), mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Shared
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/reports")
@login_required
def api_reports():
    # Indexing reports: csv files (with paired html)
    csv_files = sorted(REPORTS_DIR.glob("indexing_report_*.csv"), reverse=True)[:20]

    # Audit reports: prefer html files, fall back to json-only sidecars.
    # html_bases must cover ALL html files (not just the display cap) so that
    # reports beyond position 20 don't reappear as phantom JSON-only entries.
    all_audit_html   = sorted(REPORTS_DIR.glob("seo_audit_*.html"), reverse=True)
    html_bases       = {f.stem for f in all_audit_html}
    audit_html_files = all_audit_html[:20]
    audit_json_only  = sorted(
        [f for f in REPORTS_DIR.glob("seo_audit_*.json") if f.stem not in html_bases],
        reverse=True
    )[:20]

    result = []
    for f in csv_files:
        stat     = f.stat()
        html_ver = f.with_suffix(".html")
        result.append({
            "name":      f.name,
            "html_name": html_ver.name if html_ver.exists() else f.name,
            "date":      datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "size":      f"{stat.st_size//1024} KB" if stat.st_size > 1024 else f"{stat.st_size} B",
            "kind":      "indexing",
        })
    def _read_avg_score(jsn_path):
        try:
            return json.loads(jsn_path.read_text(encoding="utf-8")).get("avg_score")
        except Exception:
            return None

    for f in audit_html_files:
        stat = f.stat()
        xlsx = f.with_suffix(".xlsx")
        jsn  = f.with_suffix(".json")
        result.append({
            "name":      f.name,
            "html_name": f.name,
            "xlsx_name": xlsx.name if xlsx.exists() else None,
            "json_name": jsn.name  if jsn.exists()  else None,
            "score":     _read_avg_score(jsn) if jsn.exists() else None,
            "date":      datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "size":      f"{stat.st_size//1024} KB" if stat.st_size > 1024 else f"{stat.st_size} B",
            "kind":      "audit",
        })
    for f in audit_json_only:
        stat = f.stat()
        xlsx = f.with_suffix(".xlsx")
        result.append({
            "name":      f.name,
            "html_name": None,
            "xlsx_name": xlsx.name if xlsx.exists() else None,
            "json_name": f.name,
            "score":     _read_avg_score(f),
            "date":      datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "size":      f"{stat.st_size//1024} KB" if stat.st_size > 1024 else f"{stat.st_size} B",
            "kind":      "audit",
        })
    return jsonify(result)

@app.route("/api/open/<filename>")
@login_required
def api_open(filename):
    # Path traversal hardening — only serve .html that resolves inside REPORTS_DIR.
    base = filename.rsplit(".", 1)[0] + ".html"
    safe = _safe_report_path(base, (".html",))
    if safe is None or not safe.is_file():
        return "Report not found", 404
    return safe.read_text(encoding="utf-8"), 200, {"Content-Type": "text/html"}

@app.route("/api/download/<filename>")
@login_required
def api_download(filename):
    # Path traversal hardening — restrict to known report extensions inside REPORTS_DIR.
    safe = _safe_report_path(filename, (".xlsx", ".csv", ".html", ".json", ".pdf"))
    if safe is None or not safe.is_file():
        return "Not found", 404
    if filename.endswith(".xlsx"):
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif filename.endswith(".pdf"):
        mime = "application/pdf"
    elif filename.endswith(".json"):
        mime = "application/json"
    elif filename.endswith(".html"):
        mime = "text/html"
    else:
        mime = "text/csv"
    return safe.read_bytes(), 200, {
        "Content-Type": mime,
        "Content-Disposition": f"attachment; filename={safe.name}",
    }

import re as _re

_REPORT_STEM_RE = _re.compile(r'^(indexing_report|seo_audit)_[\w\-]+$')

def _delete_report_stem(stem: str) -> list[str]:
    """Delete every file in REPORTS_DIR that shares *stem*, regardless of extension.
    Only touches files whose stem matches a known report pattern, so unrelated files
    in the same directory are never affected even if a caller passes a crafted name."""
    if not _REPORT_STEM_RE.match(stem):
        return []
    removed = []
    for p in REPORTS_DIR.glob(f"{stem}.*"):
        try:
            p.unlink()
            removed.append(p.name)
        except Exception as exc:
            logger.warning("Could not delete %s: %s", p.name, exc)
    return removed

@app.route("/api/reports/delete/<filename>", methods=["DELETE"])
@login_required
def api_reports_delete(filename):
    if not _re.match(r'^[\w\-\.]+$', filename):
        return jsonify({"error": "Invalid filename"}), 400
    stem = filename.rsplit(".", 1)[0]
    deleted = _delete_report_stem(stem)
    if not deleted:
        return jsonify({"error": "File not found"}), 404
    return jsonify({"deleted": deleted})

@app.route("/api/reports/summary")
@login_required
def api_reports_summary():
    """Return avg score + total issues — accepts .json sidecar or .xlsx."""
    import re
    name = request.args.get("file", "")
    if not re.match(r'^[\w\.\-]+\.(json|xlsx)$', name):
        return jsonify({"error": "invalid filename"}), 400
    p = REPORTS_DIR / name
    if not p.exists():
        return jsonify({"error": "not found"}), 404
    try:
        if name.endswith(".json"):
            data = json.loads(p.read_text(encoding="utf-8"))
            return jsonify(data)
        import openpyxl as _ox
        wb = _ox.load_workbook(p, read_only=True, data_only=True)
        ws = wb["Summary"]
        scores, issues, warns = [], 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[1] is None: continue
            if row[2] is not None: scores.append(row[2])
            warns  += row[4] or 0
            issues += row[5] or 0
        avg = round(sum(scores)/len(scores)) if scores else 0
        return jsonify({"urls": len(scores), "avg_score": avg,
                        "total_issues": issues, "total_warnings": warns})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/reports/preview/<filename>")
@login_required
def api_reports_preview(filename):
    """Rich preview data for the side drawer — works for both indexing + audit reports."""
    import re, csv as _csv
    if not re.match(r'^[\w\-\.]+$', filename):
        return jsonify({"error": "Invalid filename"}), 400

    base = filename.rsplit(".", 1)[0]

    # ── Indexing report (CSV) ─────────────────────────────────────────────────
    csv_path = REPORTS_DIR / (base + ".csv")
    if csv_path.exists() or filename.endswith(".csv"):
        p = REPORTS_DIR / filename if filename.endswith(".csv") else csv_path
        if not p.exists():
            return jsonify({"error": "not found"}), 404
        try:
            rows = []
            with open(p, encoding="utf-8") as f:
                reader = _csv.DictReader(f)
                for row in reader:
                    rows.append(row)
            total = len(rows)
            indexed     = sum(1 for r in rows if r.get("Status","").strip() == "Indexed")
            not_indexed = sum(1 for r in rows if r.get("Status","").strip() == "Not Indexed")
            errors      = total - indexed - not_indexed
            rate        = round(indexed / total * 100, 1) if total else 0
            # Sample of not-indexed URLs as "key issues"
            ni_urls = [r.get("URL","") for r in rows if r.get("Status","").strip() == "Not Indexed"][:10]
            return jsonify({
                "kind": "indexing",
                "total": total, "indexed": indexed, "not_indexed": not_indexed,
                "errors": errors, "rate": rate,
                "not_indexed_sample": ni_urls,
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    # ── Audit report (JSON sidecar preferred, xlsx fallback) ──────────────────
    json_path = REPORTS_DIR / (base + ".json")
    xlsx_path = REPORTS_DIR / (base + ".xlsx")
    if json_path.exists():
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
            urls_data = data.get("urls", [])

            # Full JSON format (list of per-URL dicts) vs summary-only sidecar (int urls count)
            if isinstance(urls_data, list) and urls_data:
                scores = [u.get("score", 0) for u in urls_data if u.get("score") is not None]
                avg = round(sum(scores) / len(scores)) if scores else data.get("avg_score", 0)
                excellent = sum(1 for s in scores if s >= 80)
                good      = sum(1 for s in scores if 50 <= s < 80)
                poor      = sum(1 for s in scores if s < 50)
                issue_counts: dict = {}
                for u in urls_data:
                    for chk in u.get("results", []):
                        if chk.get("status") in ("fail", "warning"):
                            key = chk.get("label") or chk.get("tool", "unknown")
                            issue_counts[key] = issue_counts.get(key, 0) + 1
                top_issues = sorted(issue_counts.items(), key=lambda x: -x[1])[:8]
            else:
                # Summary-only sidecar: {avg_score, total_issues, total_warnings, urls(int)}
                n_urls = urls_data if isinstance(urls_data, int) else data.get("urls", 0)
                scores = []
                avg    = data.get("avg_score", 0)
                excellent, good, poor = 0, 0, n_urls  # can't know distribution
                top_issues = []

            return jsonify({
                "kind": "audit",
                "urls": len(scores) if scores else (data.get("urls", 0) if not isinstance(data.get("urls"), list) else len(urls_data)),
                "avg_score": avg,
                "total_issues":   data.get("total_issues", 0),
                "total_warnings": data.get("total_warnings", 0),
                "score_dist": {"excellent": excellent, "good": good, "poor": poor},
                "top_issues": [{"label": k, "count": v} for k, v in top_issues],
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    if xlsx_path.exists():
        try:
            import openpyxl as _ox
            wb = _ox.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb["Summary"]
            scores, issues, warns = [], 0, 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[1] is None: continue
                if row[2] is not None: scores.append(row[2])
                warns  += row[4] or 0
                issues += row[5] or 0
            avg = round(sum(scores)/len(scores)) if scores else 0
            excellent = sum(1 for s in scores if s >= 80)
            good      = sum(1 for s in scores if 50 <= s < 80)
            poor      = sum(1 for s in scores if s < 50)
            return jsonify({
                "kind": "audit",
                "urls": len(scores), "avg_score": avg,
                "total_issues": issues, "total_warnings": warns,
                "score_dist": {"excellent": excellent, "good": good, "poor": poor},
                "top_issues": [],
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    return jsonify({"error": "Report data not found"}), 404

@app.route("/api/reports/delete_bulk", methods=["POST"])
@login_required
def api_reports_delete_bulk():
    """Delete multiple reports in one call. Body: {names: [filename, ...]}"""
    data = request.get_json(force=True) or {}
    names = data.get("names", [])
    if not isinstance(names, list) or not names:
        return jsonify({"error": "names list required"}), 400

    deleted, failed = [], []
    for name in names:
        if not isinstance(name, str) or not _re.match(r'^[\w\-\.]+$', name):
            failed.append({"name": name, "error": "invalid filename"}); continue
        removed = _delete_report_stem(name.rsplit(".", 1)[0])
        if removed:
            deleted.append(name)
        else:
            failed.append({"name": name, "error": "not found"})
    return jsonify({"deleted": deleted, "failed": failed,
                    "deleted_count": len(deleted), "failed_count": len(failed)})

@app.route("/api/reports/delete_all", methods=["POST"])
@login_required
def api_reports_delete_all():
    """Nuclear option — wipes all reports. Body must include {confirm: 'YES'}."""
    data = request.get_json(force=True) or {}
    if data.get("confirm") != "YES":
        return jsonify({"error": "confirm:'YES' required"}), 400
    deleted = 0
    # Use _delete_report_stem per stem so the pattern guard is always enforced
    seen = set()
    for p in list(REPORTS_DIR.glob("indexing_report_*.*")) + list(REPORTS_DIR.glob("seo_audit_*.*")):
        if p.stem not in seen:
            seen.add(p.stem)
            deleted += len(_delete_report_stem(p.stem))
    return jsonify({"deleted_count": deleted})

@app.route("/api/history")
@login_required
def api_history():
    from core.checker import load_history
    return jsonify(load_history())

# Keys allowed in POST /api/settings — anything not in this allowlist is dropped
# silently so a malicious client can't inject arbitrary keys into config.json.
_SETTINGS_ALLOWED_KEYS = {
    "parallel_tabs", "gsc", "email", "slack", "teams", "schedule",
    "priority", "track_keywords", "timings", "thresholds", "proxies",
    # API keys — kept in config so they persist across restarts without re-entry
    "pagespeed_api_key",              # PageSpeed Insights (was wrongly named pagespeed_key)
    "serpapi_key",                    # SerpAPI rank tracking
    "dataforseo_login",               # DataForSEO login email
    "dataforseo_password",            # DataForSEO API password
    "moz_access_id",                  # Moz domain authority
    "moz_secret_key",                 # Moz secret
    "indexnow_key", "indexnow_host",  # Stage 2-B: IndexNow
    "bing_api_key",                   # Stage 3-A: Bing Webmaster
    "groq_api_key",                   # Stage 3-D: Groq AI
}

@app.route("/api/settings", methods=["GET","POST"])
@login_required
def api_settings():
    import core.checker as _checker_mod
    import core.seo_audit as _audit_mod
    cfg_path = CONFIG_PATH
    if request.method == "POST":
        new_cfg  = request.get_json() or {}
        if not isinstance(new_cfg, dict):
            return jsonify({"error": "JSON object required"}), 400
        # Drop unknown keys so attackers can't write arbitrary paths/URLs into config.
        filtered = {k: v for k, v in new_cfg.items() if k in _SETTINGS_ALLOWED_KEYS}
        rejected = sorted(set(new_cfg) - _SETTINGS_ALLOWED_KEYS)
        existing = json.loads(cfg_path.read_text()) if cfg_path.exists() else {}
        existing.update(filtered)
        cfg_path.write_text(json.dumps(existing, indent=2))
        # Refresh in-process globals so the next run uses the saved values
        refreshed = load_config()
        global CFG
        CFG = refreshed
        _checker_mod.CFG  = refreshed
        _audit_mod.CFG    = refreshed
        resp = {"message":"Settings saved ✓"}
        if rejected:
            resp["rejected_keys"] = rejected
        return jsonify(resp)
    return jsonify(json.loads(cfg_path.read_text()) if cfg_path.exists() else {})

UPLOAD_DIR    = DATA_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
PROFILES_PATH = DATA_DIR / "profiles.json"

def _load_profiles():
    if PROFILES_PATH.exists():
        try: return json.loads(PROFILES_PATH.read_text())
        except Exception: pass
    return {}

def _save_profiles(p):
    PROFILES_PATH.write_text(json.dumps(p, indent=2))

@app.route("/api/profiles", methods=["GET", "POST", "DELETE"])
@login_required
def api_profiles():
    profiles = _load_profiles()
    if request.method == "GET":
        return jsonify(profiles)
    if request.method == "DELETE":
        name = request.args.get("name", "").strip()
        if name in profiles:
            del profiles[name]; _save_profiles(profiles)
        return jsonify({"ok": True, "profiles": profiles})
    # POST: save
    data = request.get_json(force=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name required"}), 400
    profiles[name] = {
        "use_cases": data.get("use_cases", []),
        "tasks":     data.get("tasks", []),
        "keywords":  data.get("keywords", ""),
        "limit":     data.get("limit", 10),
        "saved_at":  datetime.now().isoformat(),
    }
    _save_profiles(profiles)
    return jsonify({"ok": True, "profiles": profiles})

@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    """Upload a CSV/Excel file. Returns server path for use as 'input' with input_type='csv'."""
    if "file" not in request.files:
        return jsonify({"error": "No file in request"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400
    name = f.filename
    if not name.lower().endswith((".csv", ".xlsx", ".xls", ".tsv", ".txt")):
        return jsonify({"error": "Only .csv .xlsx .xls .tsv .txt allowed"}), 400
    # sanitize filename
    # Use werkzeug's hardened sanitizer instead of a hand-rolled allowlist —
    # it handles unicode normalisation and reserved Windows names.
    safe = secure_filename(name) or "upload"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = UPLOAD_DIR / f"{ts}_{safe}"
    try:
        f.save(str(dest))
    except Exception as e:
        return jsonify({"error": f"Save failed: {e}"}), 500

    # Try to count URLs
    warning = ""
    try:
        urls = load_from_csv_excel(str(dest))
        count = len(urls)
    except Exception as exc:
        count = 0
        warning = str(exc)
    body = {"path": str(dest), "filename": safe, "url_count": count}
    if warning:
        body["warning"] = warning
    return jsonify(body)


@app.route("/api/compare")
@login_required
def api_compare():
    """Compare two audit XLSX reports — returns score diffs per URL."""
    a_name = request.args.get("a", "")
    b_name = request.args.get("b", "")
    import re as _re
    if not (_re.match(r"^[\w\.\-]+\.xlsx$", a_name) and _re.match(r"^[\w\.\-]+\.xlsx$", b_name)):
        return jsonify({"error": "invalid filenames"}), 400
    pa = REPORTS_DIR / a_name; pb = REPORTS_DIR / b_name
    if not (pa.exists() and pb.exists()):
        return jsonify({"error": "report(s) not found"}), 404
    try:
        import openpyxl as _ox
        def _scores(p):
            wb = _ox.load_workbook(p, read_only=True, data_only=True); ws = wb["Summary"]
            out = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[1] is None: continue
                out[str(row[1])] = {"score": row[2] or 0,
                                    "pass": row[3] or 0, "warn": row[4] or 0, "fail": row[5] or 0,
                                    "top_issue": row[6] or ""}
            return out
        a = _scores(pa); b = _scores(pb)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    all_urls = sorted(set(a) | set(b))
    rows = []
    for u in all_urls:
        sa = a.get(u, {}).get("score"); sb = b.get(u, {}).get("score")
        rows.append({"url": u,
                     "a_score": sa, "b_score": sb,
                     "delta": (sb - sa) if sa is not None and sb is not None else None,
                     "added": sa is None, "removed": sb is None})
    improved = sum(1 for r in rows if r["delta"] is not None and r["delta"] > 0)
    declined = sum(1 for r in rows if r["delta"] is not None and r["delta"] < 0)
    return jsonify({"a": a_name, "b": b_name, "rows": rows,
                    "improved": improved, "declined": declined,
                    "added": sum(1 for r in rows if r["added"]),
                    "removed": sum(1 for r in rows if r["removed"])})

# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Single Use-Case Runner
# ══════════════════════════════════════════════════════════════════════════════

def _run_usecase_for_url(url: str, use_case: str, cfg: dict, keywords: str = "") -> dict:
    """Shared logic: run a use case against a single resolved URL."""
    from core.seo_audit import audit_single_url, calc_seo_score
    from core.checker   import build_gsc_service

    gsc_service = None
    if cfg.get("gsc", {}).get("enabled"):
        try: gsc_service = build_gsc_service()
        except Exception: pass

    extra = {"keywords": keywords} if keywords else {}
    audit  = audit_single_url(url, cfg, gsc_service=gsc_service, use_cases=[use_case], **extra)
    checks = audit.get("results", []) if isinstance(audit, dict) else list(audit)
    score  = audit.get("score", calc_seo_score(checks)) if isinstance(audit, dict) else calc_seo_score(checks)
    return {
        "ok": True, "url": url, "use_case": use_case,
        "score": score,
        "passes":   sum(1 for r in checks if r.get("status") == "pass"),
        "warnings": sum(1 for r in checks if r.get("status") == "warning"),
        "fails":    sum(1 for r in checks if r.get("status") == "fail"),
        "results":  checks,
    }


@app.route("/api/usecase/run", methods=["POST"])
@login_required
def api_usecase_run():
    """Run a single use-case. Supports input_format: url | domain | sitemap."""
    data         = request.get_json(force=True) or {}
    raw_url      = (data.get("url") or "").strip()
    use_case     = (data.get("use_case") or "").strip()
    input_format = (data.get("input_format") or "url").strip()
    keywords     = (data.get("keywords") or "").strip()

    if not raw_url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if not use_case:
        return jsonify({"ok": False, "error": "use_case required"}), 400
    # Normalize scheme so bare domains like "example.com" work end-to-end
    if raw_url and not raw_url.startswith("http"):
        raw_url = f"https://{raw_url}"
    if (rej := _reject_unsafe(raw_url)):
        return rej

    from core.seo_audit import USE_CASES
    from core.checker   import fetch_sitemap_urls, fetch_from_domain

    if use_case not in USE_CASES:
        return jsonify({"ok": False, "error": f"Unknown use_case: {use_case}"}), 400

    try:
        if input_format == "sitemap":
            urls = fetch_sitemap_urls(raw_url)[:1]
            if not urls:
                return jsonify({
                    "ok": False,
                    "error": "Sitemap returned no valid URLs — check the URL and try again",
                }), 400
            target = urls[0]
        elif input_format == "domain":
            urls = fetch_from_domain(raw_url)[:1]
            if not urls:
                return jsonify({
                    "ok": False,
                    "error": "No crawlable URLs found for this domain — check the URL and try again",
                }), 400
            target = urls[0]
        else:
            target = raw_url

        return jsonify(_run_usecase_for_url(target, use_case, CFG, keywords))
    except Exception as e:
        logger.error("usecase/run error: %s", e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/usecase/run_bulk", methods=["POST"])
@login_required
def api_usecase_run_bulk():
    """Run a use-case against all URLs from an uploaded CSV/XLSX (max 20 URLs)."""
    import re as _re
    use_case = (request.form.get("use_case") or "").strip()
    keywords = (request.form.get("keywords") or "").strip()
    f        = request.files.get("file")

    if not f or not f.filename:
        return jsonify({"ok": False, "error": "file required"}), 400
    if not _re.search(r'\.(csv|xlsx)$', f.filename, _re.IGNORECASE):
        return jsonify({"ok": False, "error": "Only .csv or .xlsx accepted"}), 400

    from core.seo_audit import USE_CASES
    if use_case not in USE_CASES:
        return jsonify({"ok": False, "error": f"Unknown use_case: {use_case}"}), 400

    try:
        import tempfile, csv as _csv
        suffix = ".csv" if f.filename.lower().endswith(".csv") else ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = Path(tmp.name)

        from core.checker import load_from_csv_excel
        urls = load_from_csv_excel(tmp_path)[:20]
        tmp_path.unlink(missing_ok=True)

        if not urls:
            return jsonify({"ok": False, "error": "No valid URLs found in file"}), 400

        from core.seo_audit import calc_seo_score
        all_results = []
        for u in urls:
            try:
                all_results.append(_run_usecase_for_url(u, use_case, CFG, keywords))
            except Exception as ue:
                all_results.append({"ok": False, "url": u, "error": str(ue)})

        # Aggregate: return summary + per-url breakdown
        ok_results = [r for r in all_results if r.get("ok")]
        avg_score  = round(sum(r["score"] for r in ok_results) / len(ok_results)) if ok_results else 0
        # Merge checks (first URL's checks for display, summary stats across all)
        first = ok_results[0] if ok_results else {}
        return jsonify({
            "ok": True, "url": f"{len(urls)} URLs", "use_case": use_case,
            "score":    avg_score,
            "passes":   sum(r.get("passes",0)   for r in ok_results),
            "warnings": sum(r.get("warnings",0) for r in ok_results),
            "fails":    sum(r.get("fails",0)     for r in ok_results),
            "results":  first.get("results", []),
            "bulk":     all_results,
        })
    except Exception as e:
        logger.error("usecase/run_bulk error: %s", e, exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Quick Tools (Phase A)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/tools/serp_preview", methods=["POST"])
@login_required
def api_serp_preview():
    data = request.get_json(force=True) or {}
    url  = _norm_url((data.get("url") or "").strip())
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    from tools.quick_tools import serp_snippet_preview
    return jsonify(serp_snippet_preview(url))

@app.route("/api/tools/redirect_chain", methods=["POST"])
@login_required
def api_redirect_chain():
    data = request.get_json(force=True) or {}
    url  = _norm_url((data.get("url") or "").strip())
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    from tools.quick_tools import redirect_chain
    return jsonify(redirect_chain(url))

@app.route("/api/tools/http_headers", methods=["POST"])
@login_required
def api_http_headers():
    data = request.get_json(force=True) or {}
    url  = _norm_url((data.get("url") or "").strip())
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    from tools.quick_tools import http_headers
    return jsonify(http_headers(url))

@app.route("/api/tools/keyword_density", methods=["POST"])
@login_required
def api_keyword_density():
    data  = request.get_json(force=True) or {}
    url   = _norm_url((data.get("url") or "").strip())
    top_n = int(data.get("top_n", 20))
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    from tools.quick_tools import keyword_density
    return jsonify(keyword_density(url, top_n=top_n))

@app.route("/api/tools/code_text_ratio", methods=["POST"])
@login_required
def api_code_text_ratio():
    data = request.get_json(force=True) or {}
    url  = _norm_url((data.get("url") or "").strip())
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    from tools.quick_tools import code_to_text_ratio
    return jsonify(code_to_text_ratio(url))

@app.route("/api/tools/compression", methods=["POST"])
@login_required
def api_compression():
    data = request.get_json(force=True) or {}
    url  = _norm_url((data.get("url") or "").strip())
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    from tools.quick_tools import compression_headers
    return jsonify(compression_headers(url))


@app.route("/api/tools/indexnow_submit", methods=["POST"])
@login_required
def api_indexnow_submit():
    """Submit one or more URLs to IndexNow (Bing / Yandex).

    Body fields:
      urls   – list of URL strings OR a single URL string  (required)
      key    – IndexNow API key  (falls back to config "indexnow_key")
      host   – hostname that owns the key  (falls back to config "indexnow_host")
    """
    data = request.get_json(force=True) or {}

    # Accept both a list and a newline/comma separated string
    raw_urls = data.get("urls") or data.get("url") or []
    if isinstance(raw_urls, str):
        raw_urls = [u.strip() for u in raw_urls.replace(",", "\n").splitlines() if u.strip()]
    if not raw_urls:
        return jsonify({"ok": False, "error": "urls required"}), 400

    key  = (data.get("key")  or CFG.get("indexnow_key",  "")).strip()
    host = (data.get("host") or CFG.get("indexnow_host", "")).strip()
    if not key:
        return jsonify({"ok": False, "error": "IndexNow API key required (set in Settings → indexnow_key)"}), 400
    if not host:
        return jsonify({"ok": False, "error": "Host required (set in Settings → indexnow_host)"}), 400

    from tools.indexnow import submit_url, submit_bulk
    if len(raw_urls) == 1:
        return jsonify(submit_url(raw_urls[0], key, host))
    return jsonify(submit_bulk(raw_urls, key, host))


@app.route("/api/tools/indexnow_generate_key", methods=["POST"])
@login_required
def api_indexnow_generate_key():
    """Generate a fresh IndexNow API key."""
    from tools.indexnow import generate_key
    return jsonify({"ok": True, "key": generate_key()})


@app.route("/api/tools/sitemap_audit", methods=["POST"])
@login_required
def api_sitemap_audit():
    """Fetch and audit a sitemap — checks structure, duplicates, size, HTTP links."""
    data = request.get_json(force=True) or {}
    url  = _norm_url((data.get("url") or "").strip())
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    from tools.sitemap_audit import audit_sitemap
    return jsonify(audit_sitemap(url))


@app.route("/api/tools/schema_validate", methods=["POST"])
@login_required
def api_schema_validate():
    """Fetch a URL, extract JSON-LD blocks, validate against Schema.org."""
    data = request.get_json(force=True) or {}
    url  = _norm_url((data.get("url") or "").strip())
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    # schema_validator does its own validate_public_url; we still pre-check so
    # the error shape matches the other tool endpoints.
    if (rej := _reject_unsafe(url)): return rej
    from tools.schema_validator import validate_url
    return jsonify(validate_url(url))


@app.route("/api/tools/keyword_research", methods=["POST"])
@login_required
def api_keyword_research():
    """DataForSEO Labs keyword research — related / suggestions / ideas."""
    data = request.get_json(force=True) or {}
    keywords = data.get("keywords") or []
    if isinstance(keywords, str):
        keywords = [k.strip() for k in keywords.split(",") if k.strip()]
    if not keywords:
        return jsonify({"ok": False, "error": "keywords required"}), 400
    mode = (data.get("mode") or "auto").strip()
    location_code = _int(data, "location_code", 2840, 1, 99999)
    language_code = (data.get("language_code") or "en").strip()
    limit         = _int(data, "limit", 150, 1, 1000)

    # Credentials come from config (env-overlaid) so callers never need to ship
    # secrets in the request body.
    login    = CFG.get("dataforseo_login", "")
    password = CFG.get("dataforseo_password", "")

    from tools.keyword_research import research_keywords
    return jsonify(research_keywords(
        keywords, login, password,
        location_code=location_code, language_code=language_code,
        limit=limit, mode=mode,
    ))


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Stage 3-A: Bing Webmaster
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/tools/bing/overview", methods=["POST"])
@login_required
def api_bing_overview():
    """Full Bing Webmaster overview: traffic + crawl stats + sitemap status."""
    data     = request.get_json(force=True) or {}
    site_url = _norm_url((data.get("site_url") or "").strip())
    api_key  = (data.get("api_key")  or CFG.get("bing_api_key", "")).strip()
    period   = _int(data, "period", 2, 1, 3)
    if not site_url:
        return jsonify({"ok": False, "error": "site_url required"}), 400
    if (rej := _require_public_url(site_url, "site_url"))[1]:
        return rej[1]
    if not api_key:
        return jsonify({"ok": False, "error": "Bing API key required (set in Settings → bing_api_key)"}), 400
    from tools.bing_webmaster import site_overview
    return jsonify(site_overview(site_url, api_key, period=period))


@app.route("/api/tools/bing/inspect", methods=["POST"])
@login_required
def api_bing_inspect():
    """Inspect a specific URL via Bing Webmaster API."""
    data     = request.get_json(force=True) or {}
    page_url = _norm_url((data.get("url")      or "").strip())
    site_url = _norm_url((data.get("site_url") or "").strip())
    api_key  = (data.get("api_key")  or CFG.get("bing_api_key", "")).strip()
    if not page_url or not site_url:
        return jsonify({"ok": False, "error": "url and site_url required"}), 400
    if (rej := _reject_unsafe(page_url)):  return rej
    if (rej := _require_public_url(site_url, "site_url"))[1]: return rej[1]
    if not api_key:
        return jsonify({"ok": False, "error": "Bing API key required"}), 400
    from tools.bing_webmaster import inspect_url
    return jsonify(inspect_url(page_url, site_url, api_key))


@app.route("/api/tools/bing/submit", methods=["POST"])
@login_required
def api_bing_submit():
    """Submit a URL to Bing for crawling."""
    data     = request.get_json(force=True) or {}
    page_url = _norm_url((data.get("url")      or "").strip())
    site_url = _norm_url((data.get("site_url") or "").strip())
    api_key  = (data.get("api_key")  or CFG.get("bing_api_key", "")).strip()
    if not page_url or not site_url:
        return jsonify({"ok": False, "error": "url and site_url required"}), 400
    if (rej := _reject_unsafe(page_url)):  return rej
    if (rej := _require_public_url(site_url, "site_url"))[1]: return rej[1]
    if not api_key:
        return jsonify({"ok": False, "error": "Bing API key required"}), 400
    from tools.bing_webmaster import submit_url
    return jsonify(submit_url(page_url, site_url, api_key))


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Stage 3-B: GSC Opportunity Layer
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/tools/gsc_opportunities", methods=["POST"])
@login_required
def api_gsc_opportunities():
    """Run GSC opportunity analysis: low-CTR pages, position decay, cannibalization."""
    data     = request.get_json(force=True) or {}
    site_url = _norm_url((data.get("site_url") or "").strip())
    if not site_url:
        return jsonify({"ok": False, "error": "site_url required"}), 400
    if (rej := _require_public_url(site_url, "site_url"))[1]: return rej[1]

    gsc_cfg = CFG.get("gsc", {})
    if not gsc_cfg.get("enabled"):
        return jsonify({"ok": False,
                        "error": "Google Search Console not enabled — go to Settings → GSC"}), 400
    try:
        service = build_gsc_service()
    except Exception as exc:
        return jsonify({"ok": False, "error": f"GSC auth error: {exc}"}), 400

    from tools.phase3 import gsc_opportunities
    try:
        return jsonify(gsc_opportunities(service, site_url))
    except Exception as exc:
        logger.error("gsc_opportunities error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — GSC Analytics Tools (position tracker, CTR analyzer, coverage, sitemaps)
# ══════════════════════════════════════════════════════════════════════════════

def _gsc_service_or_error():
    """Return (service, None) or (None, error_response)."""
    gsc_cfg = CFG.get("gsc", {})
    if not gsc_cfg.get("enabled"):
        return None, (jsonify({"ok": False,
                               "error": "Google Search Console not enabled — go to Settings → GSC"}), 400)
    try:
        return build_gsc_service(), None
    except Exception as exc:
        return None, (jsonify({"ok": False, "error": f"GSC auth error: {exc}"}), 400)


@app.route("/api/tools/gsc_position_tracker", methods=["POST"])
@login_required
def api_gsc_position_tracker():
    """Average position trend for a URL over the last 90 days."""
    data = request.get_json(force=True) or {}
    url      = _norm_url((data.get("url")      or "").strip())
    site_url = _norm_url((data.get("site_url") or "").strip())
    if not url or not site_url:
        return jsonify({"ok": False, "error": "url and site_url required"}), 400
    if (rej := _require_public_url(url, "url"))[1]:      return rej[1]
    if (rej := _require_public_url(site_url, "site_url"))[1]: return rej[1]
    svc, err = _gsc_service_or_error()
    if err: return err
    from tools.phase3 import position_tracker
    try:
        return jsonify(position_tracker(url, svc, site_url))
    except Exception as exc:
        logger.error("gsc_position_tracker error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/tools/gsc_ctr_analyzer", methods=["POST"])
@login_required
def api_gsc_ctr_analyzer():
    """Find pages with high impressions but low CTR across the whole site."""
    data = request.get_json(force=True) or {}
    site_url        = _norm_url((data.get("site_url") or "").strip())
    min_impressions = int(data.get("min_impressions") or 100)
    if not site_url:
        return jsonify({"ok": False, "error": "site_url required"}), 400
    if (rej := _require_public_url(site_url, "site_url"))[1]: return rej[1]
    svc, err = _gsc_service_or_error()
    if err: return err
    from tools.phase3 import ctr_analyzer
    try:
        return jsonify(ctr_analyzer(svc, site_url, min_impressions=min_impressions))
    except Exception as exc:
        logger.error("gsc_ctr_analyzer error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/tools/gsc_coverage_errors", methods=["POST"])
@login_required
def api_gsc_coverage_errors():
    """Fetch coverage errors and indexing gaps from GSC sitemaps."""
    data = request.get_json(force=True) or {}
    site_url = _norm_url((data.get("site_url") or "").strip())
    if not site_url:
        return jsonify({"ok": False, "error": "site_url required"}), 400
    if (rej := _require_public_url(site_url, "site_url"))[1]: return rej[1]
    svc, err = _gsc_service_or_error()
    if err: return err
    from tools.phase3 import coverage_errors
    try:
        return jsonify(coverage_errors(svc, site_url))
    except Exception as exc:
        logger.error("gsc_coverage_errors error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/tools/gsc_sitemaps_status", methods=["POST"])
@login_required
def api_gsc_sitemaps_status():
    """List all sitemaps submitted to GSC with submission and indexing stats."""
    data = request.get_json(force=True) or {}
    site_url = _norm_url((data.get("site_url") or "").strip())
    if not site_url:
        return jsonify({"ok": False, "error": "site_url required"}), 400
    if (rej := _require_public_url(site_url, "site_url"))[1]: return rej[1]
    svc, err = _gsc_service_or_error()
    if err: return err
    from tools.phase3 import sitemaps_status
    try:
        return jsonify(sitemaps_status(svc, site_url))
    except Exception as exc:
        logger.error("gsc_sitemaps_status error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Notification test endpoints
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/tools/notify_test", methods=["POST"])
@login_required
def api_notify_test():
    """Send a test notification via email, slack, or teams."""
    data    = request.get_json(force=True) or {}
    channel = (data.get("channel") or "").strip().lower()
    if channel not in ("email", "slack", "teams"):
        return jsonify({"ok": False, "error": "channel must be email, slack, or teams"}), 400
    from core.notifier import NotificationService
    svc = NotificationService(CFG)
    try:
        if channel == "email":
            svc.send_email("SEO Suite — test notification",
                           "<h2>Test OK</h2><p>Your email notification is configured correctly.</p>")
        elif channel == "slack":
            svc.send_slack("SEO Suite test notification: Slack is configured correctly.")
        else:
            svc.send_teams("SEO Suite test notification: Teams is configured correctly.")
        return jsonify({"ok": True, "message": f"Test {channel} sent successfully"})
    except Exception as exc:
        logger.error("notify_test error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Individual phase execution
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/audit/phase/<int:phase_num>", methods=["POST"])
@login_required
def api_audit_single_phase(phase_num: int):
    """Run a single audit phase (1–4) against a URL and return results immediately."""
    if phase_num not in (1, 2, 3, 4):
        return jsonify({"ok": False, "error": "phase must be 1, 2, 3, or 4"}), 400
    data = request.get_json(force=True) or {}
    url  = _norm_url((data.get("url") or "").strip())
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej

    results = []
    try:
        if phase_num == 1:
            from tools.phase1 import (
                robots_check, http_status_check, redirect_check, canonical_check,
                title_check, meta_description_check, heading_check, image_alt_check,
                word_count_check, broken_link_check, internal_links_check,
                sitemap_validate, schema_check,
            )
            fns = [robots_check, http_status_check, redirect_check, canonical_check,
                   title_check, meta_description_check, heading_check, image_alt_check,
                   word_count_check, broken_link_check, internal_links_check,
                   sitemap_validate, schema_check]
            from concurrent.futures import ThreadPoolExecutor
            with ThreadPoolExecutor(max_workers=6) as ex:
                futs = [ex.submit(fn, url) for fn in fns]
                results = [f.result() for f in futs]

        elif phase_num == 2:
            api_key = CFG.get("pagespeed_api_key", "")
            if not api_key:
                return jsonify({"ok": False,
                                "error": "PageSpeed API key not set — go to Settings → Performance"}), 400
            from tools.phase2 import audit_url as p2_audit
            results = p2_audit(url, api_key=api_key)

        elif phase_num == 3:
            gsc_cfg = CFG.get("gsc", {})
            if not gsc_cfg.get("enabled"):
                return jsonify({"ok": False,
                                "error": "GSC not enabled — go to Settings → Indexing & Crawling"}), 400
            svc = build_gsc_service()
            from urllib.parse import urlparse
            parsed   = urlparse(url)
            site_url = f"{parsed.scheme}://{parsed.netloc}/"
            from tools.phase3 import clicks_impressions, top_queries, position_tracker, ctr_analyzer, coverage_errors, sitemaps_status
            from concurrent.futures import ThreadPoolExecutor
            fns = [
                lambda: clicks_impressions(url, svc, site_url),
                lambda: top_queries(url, svc, site_url),
                lambda: position_tracker(url, svc, site_url),
                lambda: ctr_analyzer(svc, site_url),
                lambda: coverage_errors(svc, site_url),
                lambda: sitemaps_status(svc, site_url),
            ]
            with ThreadPoolExecutor(max_workers=3) as ex:
                results = [f.result() for f in [ex.submit(fn) for fn in fns]]

        elif phase_num == 4:
            from tools.phase4 import backlink_check, domain_authority, keyword_rank_tracker
            from concurrent.futures import ThreadPoolExecutor
            fns = []
            dfs_login = CFG.get("dataforseo_login", "")
            dfs_pass  = CFG.get("dataforseo_password", "")
            moz_id    = CFG.get("moz_access_id", "")
            moz_sec   = CFG.get("moz_secret_key", "")
            keywords  = data.get("keywords") or CFG.get("track_keywords", [])
            if dfs_login or moz_id:
                fns.append(lambda: backlink_check(url, dataforseo_login=dfs_login, dataforseo_password=dfs_pass))
            if moz_id:
                fns.append(lambda: domain_authority(url, moz_id, moz_sec))
            serpapi_key = CFG.get("serpapi_key", "")
            if keywords and (serpapi_key or dfs_login):
                fns.append(lambda: keyword_rank_tracker(url, keywords,
                                                         serpapi_key=serpapi_key,
                                                         dataforseo_login=dfs_login,
                                                         dataforseo_password=dfs_pass))
            if not fns:
                return jsonify({"ok": False,
                                "error": "Phase 4 requires at least one of: DataForSEO, Moz, or SerpAPI credentials"}), 400
            with ThreadPoolExecutor(max_workers=len(fns)) as ex:
                results = [f.result() for f in [ex.submit(fn) for fn in fns]]

        from core.seo_audit import calc_seo_score
        score    = calc_seo_score(results)
        issues   = [r for r in results if r.get("status") in ("fail", "error")]
        warnings = [r for r in results if r.get("status") == "warning"]
        passed   = [r for r in results if r.get("status") == "pass"]
        return jsonify({
            "ok": True, "url": url, "phase": phase_num,
            "score": score, "results": results,
            "counts": {"pass": len(passed), "warning": len(warnings), "fail": len(issues)},
        })

    except Exception as exc:
        logger.error("audit_phase_%d error: %s", phase_num, exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Stage 3-C: Performance Opportunity Layer
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/tools/perf_opportunities", methods=["POST"])
@login_required
def api_perf_opportunities():
    """Run PageSpeed analysis and return prioritised Core Web Vitals fix plan."""
    data    = request.get_json(force=True) or {}
    url     = _norm_url((data.get("url")     or "").strip())
    api_key = (data.get("api_key") or CFG.get("pagespeed_api_key", "")).strip()
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    if not api_key:
        return jsonify({"ok": False,
                        "error": "PageSpeed API key required (set in Settings → PageSpeed API Key)"}), 400
    from tools.phase2 import performance_opportunities
    try:
        return jsonify(performance_opportunities(url, api_key))
    except Exception as exc:
        logger.error("perf_opportunities error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Stage 3-D: Groq AI Assistance
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/tools/ai_explain", methods=["POST"])
@login_required
def api_ai_explain():
    """Explain audit results in plain English using Groq LLM."""
    data    = request.get_json(force=True) or {}
    results = data.get("results") or data.get("audit_results") or []
    url     = _norm_url((data.get("url") or "").strip())
    api_key = (data.get("api_key") or CFG.get("groq_api_key", "")
               or os.getenv("GROQ_API_KEY", "")).strip()
    if not results:
        return jsonify({"ok": False, "error": "results required"}), 400
    if not api_key:
        return jsonify({"ok": False,
                        "error": "Groq API key required (Settings → groq_api_key or GROQ_API_KEY env)"}), 400
    from tools.ai_assist import explain_audit
    try:
        return jsonify(explain_audit(results, api_key, url=url))
    except Exception as exc:
        logger.error("ai_explain error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/tools/ai_draft_meta", methods=["POST"])
@login_required
def api_ai_draft_meta():
    """Draft improved title + meta description variants using Groq LLM."""
    data         = request.get_json(force=True) or {}
    url          = _norm_url((data.get("url")          or "").strip())
    current_title = (data.get("title")       or "").strip()
    current_desc  = (data.get("description") or "").strip()
    top_queries   = data.get("top_queries")  or []
    api_key       = (data.get("api_key") or CFG.get("groq_api_key", "")
                     or os.getenv("GROQ_API_KEY", "")).strip()
    if not url:
        return jsonify({"ok": False, "error": "url required"}), 400
    if (rej := _reject_unsafe(url)): return rej
    if not api_key:
        return jsonify({"ok": False,
                        "error": "Groq API key required (Settings → groq_api_key)"}), 400
    from tools.ai_assist import draft_meta
    try:
        return jsonify(draft_meta(url, current_title, current_desc, top_queries, api_key))
    except Exception as exc:
        logger.error("ai_draft_meta error: %s", exc, exc_info=True)
        return jsonify({"ok": False, "error": str(exc)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES — Generators (Phase B)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/tools/schema_fields/<schema_type>")
@login_required
def api_schema_fields(schema_type):
    from tools.generators import get_schema_fields
    return jsonify(get_schema_fields(schema_type))

@app.route("/api/tools/schema_generate", methods=["POST"])
@login_required
def api_schema_generate():
    data = request.get_json(force=True) or {}
    schema_type = (data.get("schema_type") or "").strip()
    form_data   = data.get("data", {})
    if not schema_type:
        return jsonify({"ok": False, "error": "schema_type required"}), 400
    from tools.generators import generate_schema
    return jsonify(generate_schema(schema_type, form_data))

@app.route("/api/tools/robots_txt", methods=["POST"])
@login_required
def api_robots_txt():
    data = request.get_json(force=True) or {}
    from tools.generators import generate_robots_txt
    return jsonify(generate_robots_txt(data))

@app.route("/api/tools/sitemap_generate", methods=["POST"])
@login_required
def api_sitemap_generate():
    data = request.get_json(force=True) or {}
    from tools.generators import generate_sitemap
    return jsonify(generate_sitemap(data))

@app.route("/api/tools/hreflang_generate", methods=["POST"])
@login_required
def api_hreflang_generate():
    data = request.get_json(force=True) or {}
    from tools.generators import generate_hreflang
    return jsonify(generate_hreflang(data))

@app.route("/api/tools/meta_tags_generate", methods=["POST"])
@login_required
def api_meta_tags_generate():
    data = request.get_json(force=True) or {}
    from tools.generators import generate_meta_tags
    return jsonify(generate_meta_tags(data))


# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/auth_status")
@login_required
def api_auth_status():
    """Return current auth configuration state (for Settings → Security card)."""
    return jsonify({
        "auth_enabled": auth_enabled(),
        "username":     os.getenv("SEO_SUITE_USERNAME", "admin"),
        "secret_set":   bool(os.getenv("SEO_SUITE_SECRET")),
    })


@app.route("/api/auth/change_credentials", methods=["POST"])
@login_required
def api_auth_change_credentials():
    """
    Generate a new password hash for the supplied username + password and
    return it so the user can paste it into their .env file.

    We intentionally do NOT write to .env automatically — env changes must be
    explicit and deliberate. The hash is returned in the response so the user
    can update SEO_SUITE_USERNAME / SEO_SUITE_PASSWORD_HASH themselves.
    """
    data     = request.get_json(force=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username:
        return jsonify({"ok": False, "error": "username required"}), 400
    if len(password) < 12:
        return jsonify({"ok": False, "error": "password must be at least 12 characters"}), 400
    from werkzeug.security import generate_password_hash
    pw_hash = generate_password_hash(password)
    logger.info("Credential change requested for username: %s", username)
    return jsonify({
        "ok":      True,
        "message": "Hash generated — update your .env file with these values and restart the server",
        "env_snippet": (
            f"SEO_SUITE_USERNAME={username}\n"
            f"SEO_SUITE_PASSWORD_HASH={pw_hash}"
        ),
    })


@app.route("/api/use_cases")
@login_required
def api_use_cases():
    from core.seo_audit import USE_CASES
    return jsonify(USE_CASES)

@app.route("/api/tasks")
@login_required
def api_tasks():
    from core.seo_audit import TASKS
    return jsonify(TASKS)

@app.route("/health")
def health():
    # Public, intentionally minimal — no version, no run-state. Use authed
    # endpoints (e.g. /api/history) for in-depth status. Probes only need 200.
    return jsonify({"status": "ok"})

def _save_partial_index_report() -> tuple[str, str]:
    """Save a partial indexing CSV+HTML from whatever has been collected so far.
    Returns (html_filename, csv_filename) — empty strings on failure."""
    from core.report_generator import ReportGenerator
    snapshot = _snapshot_last_index_run()
    if not snapshot:
        return "", ""
    try:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = f"indexing_report_{ts}_partial"
        csv_path  = REPORTS_DIR / f"{stem}.csv"
        html_path = REPORTS_DIR / f"{stem}.html"
        counts    = {}
        rows      = []
        from core.checker import get_priority_score, get_crawl_depth, get_url_type
        ts_now = datetime.now().isoformat()
        for i, (url, status) in enumerate(snapshot.items(), 1):
            rows.append({"num": i, "url": url, "status": status,
                         "priority": get_priority_score(url),
                         "depth": get_crawl_depth(url),
                         "url_type": get_url_type(url),
                         "checked_at": ts_now})
            counts[status] = counts.get(status, 0) + 1
        import csv as _csv
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = _csv.writer(f)
            w.writerow(["#", "URL", "Google Indexed", "Priority", "Depth", "URL Type", "Checked At"])
            for r in rows:
                w.writerow([r["num"], r["url"], r["status"], r["priority"],
                             r["depth"], r["url_type"], r["checked_at"]])
        ReportGenerator().html(rows, html_path, counts, {}, None, [])
        return html_path.name, csv_path.name
    except Exception as exc:
        logger.warning("Partial index report save failed: %s", exc)
        return "", ""

@app.route("/api/index/cancel", methods=["POST"])
@login_required
def api_index_cancel():
    # Signal cancellation only — let the worker's finally clear `running`.
    # If we flipped `running` here while the worker thread was still alive, a
    # second /api/index/run posted immediately after this would be accepted
    # and race against the still-draining prior thread on _last_index_run.
    _index_cancel.set()
    _index_paused.set()   # unblock thread so it can exit
    html_name, csv_name = _save_partial_index_report()
    with _lock:
        done = _index_status.get("done", 0)
    _broadcast_index({
        "type":    "cancelled",
        "report":  html_name,
        "csv":     csv_name,
        "done":    done,
    })
    return jsonify({"cancelled": True})

@app.route("/api/index/pause", methods=["POST"])
@login_required
def api_index_pause():
    if not _index_status["running"]:
        return jsonify({"error": "Not running"}), 400
    _index_paused.clear()   # block the worker thread
    _index_queue.put({"type": "paused"})
    return jsonify({"paused": True})

@app.route("/api/index/resume", methods=["POST"])
@login_required
def api_index_resume():
    _index_paused.set()     # unblock the worker thread
    _index_queue.put({"type": "resumed"})
    return jsonify({"resumed": True})

@app.route("/api/index/errors")
@login_required
def api_index_errors():
    data = _snapshot_last_index_run()
    error_urls = [url for url, status in data.items() if _is_error_status(status)]
    return jsonify({"error_urls": error_urls, "count": len(error_urls)})

@app.route("/api/index/retry", methods=["POST"])
@login_required
def api_index_retry():
    global _index_status
    if _index_status["running"]:
        return jsonify({"error": "Already running"}), 400

    error_urls = [
        url for url, status in _snapshot_last_index_run().items()
        if _is_error_status(status)
    ]

    if not error_urls:
        return jsonify({"error": "No failed URLs to retry"}), 400

    data     = request.get_json() or {}
    headless = data.get("headless", False)
    quiet    = data.get("quiet", False)

    with _lock:
        if _index_status["running"]:
            return jsonify({"error": "Already running"}), 400
        _index_status = {"running": True, "total": len(error_urls), "done": 0}

    def run():
        try:
            _run_results: dict[str, str] = {}

            def cb(num, total, url, status):
                _run_results[url] = status
                _update_last_index_run(url, status)
                _index_queue.put({"type": "progress", "num": num, "total": total, "url": url, "status": status})
                with _lock:
                    _index_status["done"] = num

            html = execute_and_save(error_urls, headless=headless, quiet=quiet,
                                    do_compare=False, prev_report=None, progress_cb=cb)
            with _lock:
                _last_index_run.update(_run_results)
            error_count = sum(1 for s in _run_results.values() if _is_error_status(s))
            _index_queue.put({"type": "done", "report": str(html), "error_count": error_count})
        except Exception as e:
            logger.error("Retry thread error: %s", e, exc_info=True)
            _index_queue.put({"type": "error", "message": str(e)})
        finally:
            with _lock:
                _index_status["running"] = False

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"total": len(error_urls), "started": True})

@app.route("/api/reports/pdf/<filename>")
@login_required
def api_reports_pdf(filename):
    # Path traversal hardening — resolve under REPORTS_DIR, require .html sibling.
    base = filename.rsplit(".", 1)[0] + ".html"
    html_path = _safe_report_path(base, (".html",))
    if html_path is None or not html_path.is_file():
        return "Report not found", 404
    stem = html_path.stem
    # Concurrency cap — sync Playwright forks a chromium process per call, so
    # without this a burst of PDF requests can exhaust memory. Non-blocking
    # acquire: callers see 503 instead of piling up request threads.
    if not _pdf_semaphore.acquire(blocking=False):
        return jsonify({"error": "PDF generation busy — retry shortly"}), 503
    try:
        from playwright.sync_api import sync_playwright
        abs_uri = html_path.resolve().as_uri()
        with sync_playwright() as pw:
            browser = pw.chromium.launch()
            page = browser.new_page()
            page.goto(abs_uri, wait_until="networkidle", timeout=30000)
            pdf_bytes = page.pdf(format="A4", print_background=True)
            browser.close()
        return pdf_bytes, 200, {
            "Content-Type": "application/pdf",
            "Content-Disposition": f"attachment; filename={stem}.pdf",
        }
    except Exception as e:
        logger.error("PDF generation failed: %s", e, exc_info=True)
        return jsonify({"error": str(e)}), 500
    finally:
        _pdf_semaphore.release()

def _save_partial_audit_report() -> tuple[str, str]:
    """Save a partial audit HTML+Excel from full results collected so far.
    Returns (html_filename, xlsx_filename) — empty strings on failure/no data."""
    with _lock:
        snapshot = list(_audit_full_results)
    if not snapshot:
        return "", ""
    try:
        ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem       = f"seo_audit_{ts}_partial"
        html_path  = REPORTS_DIR / f"{stem}.html"
        excel_path = REPORTS_DIR / f"{stem}.xlsx"
        json_path  = REPORTS_DIR / f"{stem}.json"
        html_path.write_text(generate_html_report(snapshot, [], ts), encoding="utf-8")
        xlsx_ok = False
        try:
            generate_excel_report(snapshot, excel_path)
            xlsx_ok = True
        except Exception as _xe:
            logger.warning("Partial Excel report failed (non-fatal): %s", _xe)
        try:
            sidecar = {
                "avg_score":      round(sum(a["score"] for a in snapshot) / len(snapshot)),
                "total_issues":   sum(len(a.get("issues", [])) for a in snapshot),
                "total_warnings": sum(len(a.get("warnings", [])) for a in snapshot),
                "urls":           len(snapshot),
            }
            json_path.write_text(json.dumps(sidecar), encoding="utf-8")
        except Exception:
            pass
        return html_path.name, (excel_path.name if xlsx_ok else "")
    except Exception as exc:
        logger.warning("Partial audit report save failed: %s", exc)
        return "", ""

@app.route("/api/audit/cancel", methods=["POST"])
@login_required
def api_audit_cancel():
    # Signal cancellation only — let the worker's finally clear `running`.
    _audit_cancel.set()
    _audit_paused.set()   # unblock thread so it can see the cancel flag
    html_name, xlsx_name = _save_partial_audit_report()
    with _lock:
        done = _audit_status.get("done", 0)
    _broadcast_audit({
        "type":   "cancelled",
        "report": html_name,
        "xlsx":   xlsx_name,
        "done":   done,
    })
    return jsonify({"cancelled": True})

@app.route("/api/audit/pause", methods=["POST"])
@login_required
def api_audit_pause():
    if not _audit_status["running"]:
        return jsonify({"error": "Not running"}), 400
    _audit_paused.clear()
    _audit_queue.put({"type": "paused"})
    return jsonify({"paused": True})

@app.route("/api/audit/resume", methods=["POST"])
@login_required
def api_audit_resume():
    _audit_paused.set()
    _audit_queue.put({"type": "resumed"})
    return jsonify({"resumed": True})

@app.route("/api/audit/partial")
@login_required
def api_audit_partial():
    """Return completed-so-far audit results as CSV for mid-run export."""
    import csv as _csv, io as _io
    with _lock:
        rows = list(_audit_partial)
    if not rows:
        return jsonify({"error": "No results yet"}), 404
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["URL", "Score", "Issues", "Warnings"])
    for r in rows:
        w.writerow([r["url"], r["score"], r["issues"], r["warnings"]])
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return buf.getvalue(), 200, {
        "Content-Type": "text/csv",
        "Content-Disposition": f"attachment; filename=audit_partial_{ts}.csv",
    }

@app.route("/api/index/partial")
@login_required
def api_index_partial():
    """Return completed-so-far indexing results as CSV for mid-run export."""
    import csv as _csv, io as _io
    data = _snapshot_last_index_run()
    if not data:
        return jsonify({"error": "No results yet"}), 404
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["URL", "Status"])
    for url, status in data.items():
        w.writerow([url, status])
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return buf.getvalue(), 200, {
        "Content-Type": "text/csv",
        "Content-Disposition": f"attachment; filename=indexing_partial_{ts}.csv",
    }


# NOTE: the canonical entry point is main.py at the project root. This module
# is intentionally NOT runnable directly — drift between two __main__ blocks
# (different ports, different env defaults) is too easy a footgun.
