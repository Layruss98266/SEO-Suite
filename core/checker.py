"""
Google Indexing Status Checker
Features: Parallel tabs · GSC API · Proxy rotation · Email/Slack/Teams
          Resume · Retry · Excel · HTML · Compare · Filter · Schedule
          Crawl depth · Priority score · Trend chart · Multiple inputs
"""

import csv
import json
import logging
import os
import random
import re
import sys
import threading
import time
import xml.etree.ElementTree as ET
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

logger = logging.getLogger(__name__)

# Re-export focused classes so callers can import from core.checker or core.<module>
from core.notifier import NotificationService  # noqa: E402
from core.report_generator import ReportGenerator  # noqa: E402
from core.security import esc as _esc  # noqa: E402
from core.security import filter_public_urls as _filter_urls
from core.security import validate_public_url as _validate_url

# ── Colors ────────────────────────────────────────────────────────────────────
try:
    from colorama import Fore, Style, init
    init(autoreset=True)
    GREEN = Fore.GREEN; RED = Fore.RED; YELLOW = Fore.YELLOW
    CYAN = Fore.CYAN;   BOLD = Style.BRIGHT; RESET = Style.RESET_ALL
except ImportError:
    GREEN = RED = YELLOW = CYAN = BOLD = RESET = ""

# ── Optional deps ─────────────────────────────────────────────────────────────
try:
    from playwright.sync_api import TimeoutError as PWTimeout
    from playwright.sync_api import sync_playwright
    _PLAYWRIGHT_IMPORT_ERROR = None
except ImportError as _e:
    # Defer the hard failure until indexing actually runs. This lets the module
    # import in audit-only environments and CI (where Playwright might not be
    # installed) without bringing down every other route in app/server.py.
    sync_playwright = None
    PWTimeout = Exception  # type: ignore[misc,assignment]
    _PLAYWRIGHT_IMPORT_ERROR = _e


def _require_playwright() -> None:
    """Raise a runtime error only when browser-based indexing is actually used."""
    if sync_playwright is None:
        raise RuntimeError(
            "Playwright is required for indexing checks. "
            "Run: pip install playwright && playwright install chromium"
        ) from _PLAYWRIGHT_IMPORT_ERROR

try:
    from tqdm import tqdm; HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    import schedule as sched; HAS_SCHEDULE = True
except ImportError:
    HAS_SCHEDULE = False

try:
    from playwright_stealth import stealth_sync; HAS_STEALTH = True
except ImportError:
    HAS_STEALTH = False
    def stealth_sync(page): pass  # no-op fallback

try:
    import requests as req_lib; HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

CHROME_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"

# ── Dirs ──────────────────────────────────────────────────────────────────────
# Anchor all runtime paths to the project root so the app behaves consistently
# regardless of the working directory from which it is launched. Previously
# Path("data") resolved relative to cwd, which could split config, reports,
# uploads, and profiles across different trees when launched outside repo root.
_PROJECT_ROOT = Path(__file__).parent.parent.resolve()
DATA_DIR     = Path(os.environ.get("SEO_SUITE_DATA_DIR", str(_PROJECT_ROOT / "data")))
DATA_DIR.mkdir(parents=True, exist_ok=True)
PROGRESS_DIR = DATA_DIR / "progress";            PROGRESS_DIR.mkdir(exist_ok=True)
REPORTS_DIR  = DATA_DIR / "reports";             REPORTS_DIR.mkdir(exist_ok=True)
HISTORY_FILE = DATA_DIR / "history.json"

# ── Config ────────────────────────────────────────────────────────────────────
def load_config() -> dict:
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass

    cfg_file = _PROJECT_ROOT / "config.json"
    cfg = json.loads(cfg_file.read_text()) if cfg_file.exists() else {}

    # Overlay secrets from environment variables so they never live in config.json
    import os
    _env_keys = {
        "pagespeed_api_key": "PAGESPEED_API_KEY",
        "serpapi_key":        "SERPAPI_KEY",
        "moz_access_id":      "MOZ_ACCESS_ID",
        "moz_secret_key":     "MOZ_SECRET_KEY",
        "dataforseo_login":   "DATAFORSEO_LOGIN",
        "dataforseo_password":"DATAFORSEO_PASSWORD",
    }
    for cfg_key, env_key in _env_keys.items():
        val = os.getenv(env_key, "")
        if val:
            cfg[cfg_key] = val

    return cfg

def _get_cfg() -> dict:
    """Return the live config dict.

    Prefers ``app.state.CFG`` (kept up-to-date by the settings save handler)
    so that config changes made via the web UI take effect immediately without
    restarting the process.  Falls back to ``load_config()`` when running as a
    standalone CLI (no Flask app present).
    """
    try:
        from app.state import CFG as _app_cfg
        return _app_cfg
    except ImportError:
        return load_config()


# Module-level alias kept for backward-compatibility with any import that does
# ``from core.checker import CFG``.  Callers inside this module use _get_cfg().
CFG = load_config()

def _t(key: str, default: Any) -> Any:
    """Read a value from the live config's 'timings', falling back to default."""
    return _get_cfg().get("timings", {}).get(key, default)

# ── Beep ──────────────────────────────────────────────────────────────────────
def beep() -> None:
    # Sound is opt-in: set SEO_SUITE_SOUND=1 to enable. Avoids audible side-effects
    # under Flask/gunicorn and keeps the import path cross-platform clean.
    if os.environ.get("SEO_SUITE_SOUND") != "1":
        return
    if sys.platform == "win32":
        try:
            import winsound
            for _ in range(3):
                winsound.Beep(1000, 400)
                time.sleep(0.2)
        except Exception:
            sys.stdout.write("\a\a\a")
            sys.stdout.flush()
    else:
        sys.stdout.write("\a\a\a")
        sys.stdout.flush()


# ══════════════════════════════════════════════════════════════════════════════
# INPUT — Sitemap / CSV / Domain
# ══════════════════════════════════════════════════════════════════════════════

def fetch_sitemap_urls(sitemap_url: str, _depth: int = 0, _seen: set | None = None) -> list[str]:
    # Defend against sitemap-index loops (an index pointing to itself or to a
    # cycle) by capping recursion depth and tracking visited URLs.
    if _depth > 5:
        logger.warning("Sitemap recursion depth exceeded at %s — stopping", sitemap_url)
        return []
    if _seen is None:
        _seen = set()
    if sitemap_url in _seen:
        return []
    try:
        sitemap_url = _validate_url(sitemap_url)
    except Exception as e:
        logger.error("SSRF guard blocked sitemap URL %s: %s", sitemap_url, e)
        return []
    _seen.add(sitemap_url)
    parsed = urlparse(sitemap_url)
    if parsed.scheme not in ("http", "https"):
        logger.error("Invalid sitemap URL (must start with http/https): %s", sitemap_url)
        return []
    logger.info("Fetching sitemap: %s", sitemap_url)
    # Use safe_requests_get so every redirect hop is re-validated against the
    # SSRF allowlist. The previous urllib.request.urlopen call followed redirects
    # with no validation, letting a public sitemap URL bounce the server into
    # private or metadata addresses.
    from core.security import safe_requests_get as _safe_get
    try:
        r = _safe_get(
            sitemap_url,
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=_t("sitemap_fetch_timeout_secs", 15),
        )
        content = r.content
    except Exception as e:
        logger.error("Failed to fetch sitemap %s: %s", sitemap_url, e)
        return []

    ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
    root = None
    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        try:
            from lxml import etree as _lxml_et
            _parser = _lxml_et.XMLParser(recover=True, resolve_entities=False, no_network=True, huge_tree=False)
            root = _lxml_et.fromstring(content, _parser)
            ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
            children_lxml = root.findall("{http://www.sitemaps.org/schemas/sitemap/0.9}sitemap/{http://www.sitemaps.org/schemas/sitemap/0.9}loc")
            if children_lxml:
                urls: list[str] = []
                for node in children_lxml:
                    if node.text:
                        urls.extend(fetch_sitemap_urls(node.text.strip(), _depth + 1, _seen))
                return urls
            return [loc.text.strip() for loc in root.findall(".//{http://www.sitemaps.org/schemas/sitemap/0.9}loc") if loc.text]
        except Exception:
            import re as _re
            return _re.findall(r"<loc>\s*(https?://[^\s<]+)\s*</loc>", content.decode("utf-8", errors="replace"))

    children = root.findall("sm:sitemap/sm:loc", ns)
    if children:
        urls: list[str] = []
        for node in children:
            urls.extend(fetch_sitemap_urls(node.text.strip(), _depth + 1, _seen))
        return urls
    return [loc.text.strip() for loc in root.findall("sm:url/sm:loc", ns)]

def fetch_from_domain(domain: str) -> list[str]:
    domain = domain.strip().rstrip("/")
    if not domain.startswith("http"):
        domain = "https://" + domain
    try:
        domain = _validate_url(domain)
    except ValueError as e:
        logger.error("Refusing to scan non-public domain %s: %s", domain, e)
        return []
    for path in ["/sitemap.xml", "/sitemap_index.xml", "/sitemap/"]:
        urls = fetch_sitemap_urls(domain + path)
        if urls:
            return urls
    logger.warning("No sitemap found for %s — using homepage only", domain)
    return [domain]

def crawl_site(start_url: str, max_pages: int = 50, max_depth: int = 2) -> list[str]:
    """BFS crawl from start_url, following internal <a href> links up to max_depth."""
    from collections import deque
    from urllib.parse import urljoin as _uj
    from urllib.parse import urlparse as _up

    from bs4 import BeautifulSoup as _BS

    # Use safe_requests_get so every redirect hop is re-validated against the
    # SSRF allowlist. The previous raw `requests.get` followed redirects with
    # no validation, letting a public URL bounce the server into private space.
    from core.security import safe_requests_get as _safe_get

    start_url = start_url.strip()
    if not start_url.startswith("http"):
        start_url = "https://" + start_url
    try:
        start_url = _validate_url(start_url)
    except ValueError as e:
        logger.error("Refusing to crawl non-public URL %s: %s", start_url, e)
        return []
    base_host = _up(start_url).netloc

    seen, found = set(), []
    q = deque([(start_url, 0)])
    headers = {"User-Agent": "Mozilla/5.0 (compatible; SEOCrawler/1.0)"}
    while q and len(found) < max_pages:
        url, depth = q.popleft()
        if url in seen: continue
        seen.add(url)
        try:
            r = _safe_get(url, headers=headers, timeout=_t("crawl_request_timeout_secs", 8))
            if r.status_code != 200 or "text/html" not in r.headers.get("Content-Type", ""):
                continue
        except Exception:
            continue
        found.append(url)
        if depth >= max_depth: continue
        try:
            soup = _BS(r.text, "html.parser")
            for a in soup.find_all("a", href=True):
                href = _uj(url, a["href"]).split("#")[0].rstrip("/")
                if not href.startswith("http"): continue
                if _up(href).netloc != base_host: continue
                # Validate the link too — a crafted host header on the base
                # page could otherwise pivot crawls into private space.
                try:
                    href = _validate_url(href)
                except ValueError:
                    continue
                if href not in seen and len(found) + len(q) < max_pages * 3:
                    q.append((href, depth + 1))
        except Exception:
            continue
    return found

def load_from_csv_excel(file_path: str) -> list[str]:
    def _looks_like_header(row: list[str]) -> bool:
        normalized = [cell.strip().lower() for cell in row if cell and cell.strip()]
        if not normalized:
            return False
        return any(cell in {"url", "urls", "link", "links"} for cell in normalized)

    def _find_url_col(row: list[str]) -> int | None:
        for idx, cell in enumerate(row):
            if cell.strip().lower() in {"url", "urls", "link", "links"}:
                return idx
        return None

    def _extract_cells(rows: list[list[str]]) -> list[str]:
        if not rows:
            return []
        header = rows[0]
        url_col = _find_url_col(header) if _looks_like_header(header) else None
        if url_col is not None:
            return [
                row[url_col].strip()
                for row in rows[1:]
                if len(row) > url_col and row[url_col] and row[url_col].strip().startswith("http")
            ]
        urls = [
            cell.strip()
            for row in rows
            for cell in row
            if cell and cell.strip().startswith("http")
        ]
        if not urls and any(cell.strip() for cell in header):
            cols = ", ".join(cell.strip() for cell in header if cell and cell.strip()) or "<empty>"
            raise ValueError(
                f"File must contain a URL column or at least one http(s) URL. Found columns: {cols}"
            )
        return urls

    path = Path(file_path)
    if not path.exists():
        logger.error("File not found: %s", file_path)
        return []
    urls: list[str] = []
    if path.suffix.lower() in (".xlsx", ".xls"):
        if not HAS_XLSX:
            logger.error("openpyxl not installed — cannot read Excel file")
            return []
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        rows = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        urls = _extract_cells(rows)
    else:
        sample = path.read_text(encoding="utf-8", errors="replace")
        try:
            dialect = csv.Sniffer().sniff(sample[:2048], delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel
        with open(path, encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f, dialect)
            rows = [[cell.strip() for cell in row] for row in reader]
        urls = _extract_cells(rows)
    # Filter every URL through the SSRF allowlist — a malicious CSV could
    # otherwise smuggle internal hosts past the upload form.
    safe_urls = _filter_urls(urls)
    if len(safe_urls) != len(urls):
        logger.warning("Dropped %d non-public URLs from %s",
                       len(urls) - len(safe_urls), file_path)
    return safe_urls

def get_urls_from_user() -> list[str]:
    if not sys.stdin.isatty():
        raise RuntimeError("CLI input not available in web context")
    print(f"\n{BOLD}How do you want to provide URLs?{RESET}")
    print("  1. Sitemap URL")
    print("  2. Multiple sitemap URLs (comma-separated)")
    print("  3. Domain only (auto-detect sitemap)")
    print("  4. CSV / Excel file")
    choice = input("\nChoice (1-4): ").strip()

    if choice == "1":
        url = input("Sitemap URL: ").strip()
        if not url:
            print(f"{RED}No URL entered.{RESET}"); return []
        return fetch_sitemap_urls(url)

    elif choice == "2":
        raw   = input("Sitemap URLs (comma-separated): ").strip()
        urls: list[str] = []
        for u in raw.split(","):
            urls.extend(fetch_sitemap_urls(u.strip()))
        return urls

    elif choice == "3":
        domain = input("Domain (e.g. edstellar.com): ").strip()
        return fetch_from_domain(domain)

    elif choice == "4":
        path = input("File path (CSV or Excel): ").strip()
        return load_from_csv_excel(path)

    else:
        print(f"{RED}Invalid choice.{RESET}"); return []


# ══════════════════════════════════════════════════════════════════════════════
# URL HELPERS — filter, depth, priority, type
# ══════════════════════════════════════════════════════════════════════════════

def filter_urls(urls: list[str], pattern: str) -> list[str]:
    return [u for u in urls if pattern.lower() in u.lower()] if pattern else urls

def get_crawl_depth(url: str) -> int:
    return len([p for p in urlparse(url).path.strip("/").split("/") if p])

def get_url_type(url: str) -> str:
    path = urlparse(url).path.strip("/")
    if not path: return "Homepage"
    seg = path.split("/")[0]
    return seg.replace("-", " ").title() if seg else "Other"

def get_priority_score(url: str) -> str:
    high = _get_cfg().get("priority", {}).get("high_value_patterns", ["/category/", "/course/"])
    low  = _get_cfg().get("priority", {}).get("low_value_patterns",  ["/tag/", "/author/"])
    depth = get_crawl_depth(url)
    if any(p in url for p in high):        return "High"
    if any(p in url for p in low):         return "Low"
    if depth <= 1:                         return "High"
    if depth <= 3:                         return "Medium"
    return "Low"


# ══════════════════════════════════════════════════════════════════════════════
# PROGRESS / RESUME
# ══════════════════════════════════════════════════════════════════════════════

def progress_file_for(key: str) -> Path:
    # Use a stable hash of the full key so different URL orderings between
    # runs still resolve to the same progress file (callers join multiple URLs
    # together — without sorting/hashing, ['a','b'] and ['b','a'] would split).
    import hashlib
    digest = hashlib.sha1(key.encode("utf-8", errors="replace")).hexdigest()[:16]
    return PROGRESS_DIR / f"progress_{digest}.json"

def load_progress(pf: Path) -> dict:
    return json.loads(pf.read_text()) if pf.exists() else {}

def save_progress(pf: Path, data: dict):
    # Atomic write — concurrent retry + main run threads otherwise race on the
    # same file and can truncate it mid-write, losing progress.
    tmp = pf.with_suffix(pf.suffix + ".tmp")
    tmp.write_text(json.dumps(data, indent=2))
    os.replace(tmp, pf)


# ══════════════════════════════════════════════════════════════════════════════
# COMPARE RUNS
# ══════════════════════════════════════════════════════════════════════════════

def find_latest_report() -> Path | None:
    reports = sorted(REPORTS_DIR.glob("indexing_report_*.csv"), reverse=True)
    return reports[0] if reports else None

def compare_runs(current: dict[str, str], prev_file: Path) -> dict:
    prev: dict[str, str] = {}
    with open(prev_file, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            prev[row["URL"]] = row["Google Indexed"]
    ni, nd = [], []
    for url, status in current.items():
        old = prev.get(url)
        if not old: continue
        if old != "Indexed" and status == "Indexed":   ni.append(url)
        elif old == "Indexed" and status != "Indexed": nd.append(url)
    return {"newly_indexed": ni, "newly_deindexed": nd}


# ══════════════════════════════════════════════════════════════════════════════
# HISTORY / TREND
# ══════════════════════════════════════════════════════════════════════════════

def load_history() -> list[dict]:
    return json.loads(HISTORY_FILE.read_text()) if HISTORY_FILE.exists() else []

def save_history(counts: dict, total: int, timestamp: str):
    history = load_history()
    history.append({
        "date":        timestamp,
        "total":       total,
        "indexed":     counts.get("Indexed", 0),
        "not_indexed": counts.get("Not Indexed", 0),
    })
    HISTORY_FILE.write_text(json.dumps(history[-30:], indent=2))  # keep last 30 runs


# ══════════════════════════════════════════════════════════════════════════════
# GSC API CHECK
# ══════════════════════════════════════════════════════════════════════════════

# Token-bucket rate limiter sized to GSC URL Inspection quota.
# GSC default per-property quota is ~600 QPM (~10 QPS) and 2000/day. We stay
# comfortably under that with 6 QPS sustained and a burst of 30 to absorb
# short spikes, so a 100-URL batch no longer slams the API and gets 429-ed.
_GSC_QPS = 6.0
_GSC_BURST = 30


class _RateLimiter:
    """Simple thread-safe token bucket. Acquire() blocks until a token is free."""

    def __init__(self, rate_per_sec: float, burst: int) -> None:
        self._rate = float(rate_per_sec)
        self._capacity = float(burst)
        self._tokens = float(burst)
        self._last = time.monotonic()
        self._lock = threading.Lock()

    def acquire(self) -> None:
        while True:
            with self._lock:
                now = time.monotonic()
                # Refill tokens based on elapsed wall-clock time.
                self._tokens = min(
                    self._capacity,
                    self._tokens + (now - self._last) * self._rate,
                )
                self._last = now
                if self._tokens >= 1.0:
                    self._tokens -= 1.0
                    return
                # Not enough — compute sleep outside the lock.
                wait = (1.0 - self._tokens) / self._rate
            time.sleep(wait)


_gsc_limiter = _RateLimiter(_GSC_QPS, _GSC_BURST)


def gsc_check_url(url: str, service, site_url: str = None) -> str:
    # Lazy import — googleapiclient is optional and not present in audit-only envs.
    try:
        from googleapiclient.errors import HttpError as _HttpError
    except ImportError:
        _HttpError = ()  # type: ignore[assignment]

    if not site_url:
        parsed = urlparse(url)
        site_url = f"{parsed.scheme}://{parsed.netloc}/"

    # 3 retries with exponential backoff + jitter on 429 / 5xx: 1s, 2s, 4s base.
    max_attempts = 4
    for attempt in range(max_attempts):
        # Throttle every attempt — retry traffic counts against quota too.
        _gsc_limiter.acquire()
        try:
            resp = service.urlInspectionResult().inspect(
                body={"inspectionUrl": url, "siteUrl": site_url}
            ).execute()
            verdict = resp.get("urlInspectionResult", {}).get("indexStatusResult", {}).get("verdict", "")
            return "Indexed" if verdict == "PASS" else "Not Indexed"
        except Exception as e:  # noqa: BLE001 — re-classified below
            status = getattr(getattr(e, "resp", None), "status", None)
            is_http = _HttpError and isinstance(e, _HttpError)
            retryable = bool(is_http and status is not None and (status == 429 or 500 <= status < 600))
            if retryable and attempt < max_attempts - 1:
                backoff = (2 ** attempt) + random.uniform(0.0, 1.0)
                # Do not log the exception message — it can include credential
                # paths or quota project ids. Log status code only (C3 fix).
                logger.warning(
                    "GSC transient error (status=%s) for %s — retrying in %.2fs (attempt %d/%d)",
                    status, url, backoff, attempt + 1, max_attempts - 1,
                )
                time.sleep(backoff)
                continue
            # Final failure (or non-retryable). Preserve C3: do not leak exception
            # text (it may contain credential paths). Log via logger.exception so
            # the traceback lands in server logs but not in the returned verdict.
            logger.exception("GSC check failed for %s", url)
            return "GSC check failed"
    return "GSC check failed"

def build_gsc_service() -> Any | None:
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        creds_file = _get_cfg().get("gsc", {}).get("credentials_file", "gsc_credentials.json")
        creds = service_account.Credentials.from_service_account_file(
            creds_file,
            scopes=["https://www.googleapis.com/auth/webmasters.readonly"]
        )
        return build("searchconsole", "v1", credentials=creds)
    except Exception as e:
        logger.error("GSC setup failed: %s", e)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# BROWSER CHECK (Playwright)
# ══════════════════════════════════════════════════════════════════════════════

_print_lock = threading.Lock()


def _normalize_url(url: str) -> str:
    """Lowercase, strip scheme, strip www, strip trailing slash and fragment."""
    url = url.lower().split("#")[0].split("?")[0].rstrip("/")
    url = re.sub(r'^https?://', '', url)
    if url.startswith("www."):
        url = url[4:]
    return url


def is_page_alive(page) -> bool:
    """Return True if the Playwright page/browser is still usable."""
    try:
        page.title()
        return True
    except Exception:
        return False


def google_check(page, url: str, delay_state: dict) -> str:
    clean      = url.replace("https://", "").replace("http://", "").rstrip("/")
    # filter=0 disables Google's duplicate-result filtering; num=10 gives more to match against
    search_url = f"https://www.google.com/search?q=site:{clean}&hl=en&num=10&filter=0"
    try:
        if not is_page_alive(page):
            return "Error: Browser context closed"

        _page_timeout = _t("browser_page_timeout_ms", 20000)
        t0 = time.time()
        page.goto(search_url, wait_until="domcontentloaded", timeout=_page_timeout)
        elapsed = time.time() - t0

        # Adaptive delay
        _delay_max = _t("delay_max_secs", 15.0)
        _delay_min = _t("delay_base_secs", 3.0)
        base = delay_state["base"]
        delay_state["base"] = min(base + 1, _delay_max) if elapsed > 5 else max(base - 0.5, _delay_min) if elapsed < 2 else base
        time.sleep(random.uniform(delay_state["base"], delay_state["base"] + 3))

        html = page.content()

        if "detected unusual traffic" in html.lower() or "captcha" in html.lower():
            beep()
            with _print_lock:
                logger.warning("CAPTCHA detected — solve it in the browser window")
            delay_state["base"] = min(delay_state["base"] + 3, _delay_max)
            time.sleep(_t("captcha_wait_secs", 60))
            page.goto(search_url, wait_until="domcontentloaded", timeout=_page_timeout)
            time.sleep(4)
            html = page.content()

        delay_state["count"] = delay_state.get("count", 0) + 1
        _break_every = _t("rate_limit_break_every_n", 10)
        if delay_state["count"] % _break_every == 0:
            wait = random.uniform(
                _t("rate_limit_break_min_secs", 15),
                _t("rate_limit_break_max_secs", 25),
            )
            with _print_lock:
                logger.info("Break %.0fs to avoid rate limiting", wait)
            time.sleep(wait)

        html_lower = html.lower()
        no_results_patterns = [
            "did not match any documents",
            "no results found",
            "your search did not match",
            "did not match any web pages",
            "no webpages were found",
        ]
        if any(t in html_lower for t in no_results_patterns):
            return "Not Indexed"

        target_norm   = _normalize_url(url)
        target_domain = _normalize_url(url).split("/")[0]  # e.g. "example.com"
        target_path   = "/" + "/".join(_normalize_url(url).split("/")[1:]) if "/" in _normalize_url(url) else "/"

        # Collect result URLs — multiple selectors cover old and new Google SERP layouts
        result_urls: list[str] = []
        for selector in (
            "div.g a[href]",
            "#rso a[href]",
            ".yuRUbf a[href]",
            "div[data-sokoban-container] a[href]",
            "div[jscontroller] h3 ~ a[href]",
        ):
            try:
                for el in page.locator(selector).all()[:15]:
                    href = el.get_attribute("href") or ""
                    if href.startswith("http"):
                        result_urls.append(href.split("?")[0].split("#")[0].rstrip("/").lower())
            except Exception:
                pass

        # Fallback regex — restrict to target domain to avoid capturing nav/ad links
        if not result_urls:
            result_urls = [
                u.rstrip("/").lower()
                for u in re.findall(r'href="(https?://[^"?#]+)"', html)
                if target_domain in u.lower()
            ]

        # 1) Exact match (www/scheme-normalized)
        for r in result_urls:
            if _normalize_url(r) == target_norm:
                return "Indexed"

        # 2) Path match — covers www↔non-www and http↔https variants
        for r in result_urls:
            r_norm = _normalize_url(r)
            r_domain = r_norm.split("/")[0]
            r_path   = "/" + "/".join(r_norm.split("/")[1:]) if "/" in r_norm else "/"
            # Same domain (ignoring www) and same path
            if r_domain == target_domain and r_path == target_path:
                return "Indexed"

        # 3) result-stats count > 0 means something from this site is indexed at this path
        try:
            stats_el = page.locator("#result-stats").first
            if stats_el.count() > 0:
                nums = re.findall(r'\d+', (stats_el.text_content() or "").replace(",", ""))
                if nums and int(nums[0]) > 0:
                    return "Indexed"
        except Exception:
            pass

        return "Not Indexed"

    except PWTimeout: return "Timeout"
    except Exception as e: return f"Error: {e}"


def check_parallel(urls: list[str], n: int, proxy_list: list, headless: bool, delay_state: dict, callback) -> dict[str, str]:
    """Check URLs using n parallel browser tabs."""
    _require_playwright()
    results: dict[str, str] = {}
    batches = [urls[i::n] for i in range(n)]

    def worker(batch: list[str], proxy: dict | None):
        with sync_playwright() as p:
            launch_args = {
                "headless": headless,
                "args": ["--disable-blink-features=AutomationControlled", "--no-sandbox"],
            }
            if proxy:
                launch_args["proxy"] = proxy

            browser = p.chromium.launch(**launch_args)
            ctx = browser.new_context(
                user_agent=CHROME_UA,
                viewport={"width": 1280, "height": 800},
                locale="en-US",
            )
            page = ctx.new_page()
            stealth_sync(page)
            page.goto("https://www.google.com", wait_until="domcontentloaded")
            time.sleep(2)
            try:
                btn = page.locator("button:has-text('Accept all'), button:has-text('I agree')")
                if btn.count() > 0:
                    btn.first.click(); time.sleep(1)
            except Exception:
                pass

            local_delay = dict(delay_state)
            for url in batch:
                # Re-launch context if browser crashed mid-batch (check before each URL)
                if not is_page_alive(page):
                    try:
                        ctx2  = browser.new_context(
                            user_agent=CHROME_UA,
                            viewport={"width": 1280, "height": 800}, locale="en-US"
                        )
                        page = ctx2.new_page()
                        stealth_sync(page)
                        page.goto("https://www.google.com", wait_until="domcontentloaded")
                        time.sleep(2)
                    except Exception as e:
                        results[url] = f"Error: Browser restart failed — {e}"
                        callback(url, results[url])
                        continue
                try:
                    status = google_check(page, url, local_delay)
                except Exception as e:
                    status = f"Error: {e}"
                    # Force alive-check on next iteration by invalidating page
                    try:
                        page.close()
                    except Exception:
                        pass
                    if not is_page_alive(page):
                        page = ctx.new_page()
                        stealth_sync(page)
                results[url] = status
                callback(url, status)

            try:
                browser.close()
            except Exception:
                pass

    threads = []
    for i, batch in enumerate(batches):
        if not batch: continue
        proxy = {"server": proxy_list[i % len(proxy_list)]} if proxy_list else None
        t = threading.Thread(target=worker, args=(batch, proxy))
        t.start(); threads.append(t)

    for t in threads:
        t.join()

    return results


# ══════════════════════════════════════════════════════════════════════════════
# REPORTS — Excel + HTML
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel(rows: list[dict], path: Path, type_summary: dict):
    if not HAS_XLSX:
        logger.warning("openpyxl not installed — skipping Excel"); return
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Results"
    HDR  = PatternFill("solid", fgColor="1F4E79")
    GF   = PatternFill("solid", fgColor="C6EFCE")
    RF   = PatternFill("solid", fgColor="FFC7CE")
    YF   = PatternFill("solid", fgColor="FFEB9C")

    headers = ["#", "URL", "Status", "Priority", "Depth", "URL Type", "Checked At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HDR; cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row["num"], row["url"], row["status"], row["priority"], row["depth"], row["url_type"], row["checked_at"]])
        c = ws.cell(ws.max_row, 3)
        c.fill = GF if row["status"] == "Indexed" else RF if row["status"] == "Not Indexed" else YF

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22

    ws2 = wb.create_sheet("By URL Type")
    ws2.append(["URL Type", "Indexed", "Not Indexed", "Other", "Total"])
    for cell in ws2[1]: cell.fill = HDR; cell.font = Font(bold=True, color="FFFFFF")
    for t, c in sorted(type_summary.items()):
        ws2.append([t, c.get("Indexed",0), c.get("Not Indexed",0), c.get("Other",0), sum(c.values())])
    for col in "ABCDE": ws2.column_dimensions[col].width = 22

    wb.save(path)
    logger.info("Excel report written: %s", path)


_HTML_REPORT_ENV = None


def _get_html_report_env():
    """Lazily build the Jinja2 environment for the HTML indexing report.

    Autoescape is enabled so any user-supplied URL, status, or category text
    is HTML-escaped automatically — eliminating the manual _esc() weave that
    the previous f-string version relied on.
    """
    global _HTML_REPORT_ENV
    if _HTML_REPORT_ENV is None:
        from jinja2 import Environment, FileSystemLoader, select_autoescape
        tpl_dir = Path(__file__).parent / "templates"
        _HTML_REPORT_ENV = Environment(
            loader=FileSystemLoader(str(tpl_dir)),
            autoescape=select_autoescape(["html", "htm", "xml"]),
        )
    return _HTML_REPORT_ENV


def generate_html(rows: list[dict], path: Path, counts: dict, type_summary: dict, compare: dict | None, history: list[dict]) -> None:
    indexed     = counts.get("Indexed", 0)
    not_indexed = counts.get("Not Indexed", 0)
    other       = sum(v for k, v in counts.items() if k not in ("Indexed","Not Indexed"))
    total       = indexed + not_indexed + other
    pct         = round(indexed / total * 100, 1) if total else 0

    # Pre-serialize chart data as JSON so Jinja can drop it into <script> blocks
    # with the |safe filter (these are numbers/strings we generated ourselves,
    # not user-supplied — json.dumps handles all quoting/escaping).
    t_labels     = json.dumps(list(type_summary.keys()))
    t_indexed    = json.dumps([v.get("Indexed", 0) for v in type_summary.values()])
    t_not        = json.dumps([v.get("Not Indexed", 0) for v in type_summary.values()])
    hist_dates   = json.dumps([h["date"][:10] for h in history])
    hist_indexed = json.dumps([h["indexed"] for h in history])
    hist_not     = json.dumps([h["not_indexed"] for h in history])

    now = datetime.now()
    template = _get_html_report_env().get_template("indexing_report.html")
    html = template.render(
        rows=rows,
        counts=counts,
        compare=compare,
        history=history,
        indexed=indexed,
        not_indexed=not_indexed,
        other=other,
        total=total,
        pct=pct,
        t_labels=t_labels,
        t_indexed=t_indexed,
        t_not=t_not,
        hist_dates=hist_dates,
        hist_indexed=hist_indexed,
        hist_not=hist_not,
        today=now.strftime("%Y-%m-%d"),
        generated_at=now.strftime("%B %d, %Y at %I:%M %p"),
    )
    path.write_text(html, encoding="utf-8")
    logger.info("HTML report written: %s", path)



# ══════════════════════════════════════════════════════════════════════════════
# CORE RUN LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def _run_browser_pass(urls: list[str], headless: bool, proxy_list: list,
                      delay_state: dict, on_result: Callable) -> None:
    """Check a batch of URLs via a single browser tab (SERP `site:` scrape).

    Used as the fallback when GSC can't answer for a URL. Opens one Chromium
    context, dismisses the consent dialog, and runs google_check per URL,
    reporting each via on_result.
    """
    _require_playwright()
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
            proxy={"server": proxy_list[0]} if proxy_list else None,
        )
        ctx = browser.new_context(
            user_agent=CHROME_UA,
            viewport={"width": 1280, "height": 800}, locale="en-US",
        )
        page = ctx.new_page()
        stealth_sync(page)
        page.goto("https://www.google.com", wait_until="domcontentloaded"); time.sleep(2)
        try:
            btn = page.locator("button:has-text('Accept all'), button:has-text('I agree')")
            if btn.count() > 0: btn.first.click(); time.sleep(1)
        except Exception:
            pass
        iterator = tqdm(urls, unit="url") if HAS_TQDM else urls
        for url in iterator:
            on_result(url, google_check(page, url, delay_state))
        try:
            browser.close()
        except Exception:
            pass


def run_check(urls: list[str], use_gsc: bool | None = None, headless: bool = False,
              quiet: bool = False, progress_cb: Callable | None = None) -> tuple[list[dict], dict, dict, dict]:
    # GSC URL Inspection is the authoritative, ToS-compliant indexation source,
    # so prefer it whenever it's configured. SERP scraping is the fallback — used
    # for URLs GSC can't answer (properties the service account doesn't own) or
    # when GSC isn't set up at all. `use_gsc=None` means "auto-detect from config".
    if use_gsc is None:
        use_gsc = bool(_get_cfg().get("gsc", {}).get("enabled"))
    gsc_service = build_gsc_service() if use_gsc else None
    if use_gsc and gsc_service is None:
        use_gsc = False  # GSC requested but unavailable — fall back to browser
    if not use_gsc:
        _require_playwright()
    # Sort before joining so the resume key is stable across runs that happen
    # to receive the same set of URLs in a different order.
    pf   = progress_file_for(",".join(sorted(urls)[:3]))
    full_saved = load_progress(pf)
    # Scope saved results to only the URLs in this run so the limit is honoured
    urls_set = set(urls)
    saved    = {u: s for u, s in full_saved.items() if u in urls_set}
    # Only skip URLs that definitively resolved — retry Error/Timeout/Other
    already_done   = {u for u, s in saved.items() if s in ("Indexed", "Not Indexed")}
    remaining      = [u for u in urls if u not in already_done]

    counts: dict       = {}
    type_summary: dict = {}
    all_rows: list     = []
    current_results    = dict(saved)
    failed_urls: list  = []
    _save_counter      = [0]  # mutable counter for batching progress writes

    # Seed resumed results (only URLs in this run)
    for url in urls:
        if url not in saved:
            continue
        status = saved[url]
        t = get_url_type(url)
        type_summary.setdefault(t, {})
        type_summary[t][status] = type_summary[t].get(status, 0) + 1
        counts[status] = counts.get(status, 0) + 1
        all_rows.append({"num": len(all_rows)+1, "url": url, "status": status,
                         "priority": get_priority_score(url), "depth": get_crawl_depth(url),
                         "url_type": get_url_type(url), "checked_at": "resumed"})

    # gsc_service was already resolved at the top of run_check (GSC-first logic).
    proxy_list     = _get_cfg().get("proxies", [])
    n_parallel     = min(_get_cfg().get("parallel_tabs", 1), len(remaining)) if remaining else 1
    delay_state    = {"base": _t("delay_base_secs", 3.0), "count": 0}

    _result_lock = threading.Lock()

    def on_result(url: str, status: str):
        ts       = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        priority = get_priority_score(url)
        depth    = get_crawl_depth(url)
        url_type = get_url_type(url)

        with _result_lock:
            if "Error" in status or status == "Timeout":
                failed_urls.append(url)

            counts[status] = counts.get(status, 0) + 1
            current_results[url] = status
            type_summary.setdefault(url_type, {})
            type_summary[url_type][status] = type_summary[url_type].get(status, 0) + 1
            num = len(all_rows) + 1
            all_rows.append({"num": num, "url": url, "status": status, "priority": priority,
                             "depth": depth, "url_type": url_type, "checked_at": ts})
            # Batch disk writes: persist every 10 URLs to avoid N² I/O on large sitemaps
            _save_counter[0] += 1
            if _save_counter[0] % 10 == 0:
                save_progress(pf, {**full_saved, **current_results})

        if progress_cb:
            progress_cb(num, len(urls), url, status)

        if not quiet or status != "Indexed":
            icon = f"{GREEN}✅{RESET}" if status=="Indexed" else (f"{RED}❌{RESET}" if status=="Not Indexed" else f"{YELLOW}⚠{RESET}")
            with _print_lock:
                logger.info("[%4d] %s -> %s", num, url, status)

    if use_gsc and gsc_service:
        logger.info("Using Google Search Console API (primary)")
        iterator = tqdm(remaining, unit="url") if HAS_TQDM else remaining
        # URLs GSC can't resolve (not an owned property, quota exhausted, etc.)
        # are deferred to a browser SERP-scrape fallback so they still get a verdict.
        gsc_deferred: list[tuple[str, str]] = []
        for url in iterator:
            status = gsc_check_url(url, gsc_service)
            if status == "GSC check failed":
                gsc_deferred.append((url, status))
            else:
                on_result(url, status)

        if gsc_deferred:
            if sync_playwright is None:
                # No browser available — surface the GSC error rather than guessing.
                for url, err in gsc_deferred:
                    on_result(url, err)
            else:
                logger.warning(
                    "GSC could not resolve %d URL(s) (not owned / quota) — falling back to browser",
                    len(gsc_deferred),
                )
                _run_browser_pass([u for u, _ in gsc_deferred], headless,
                                  proxy_list, delay_state, on_result)
    elif n_parallel > 1:
        logger.info("Using %d parallel browser tabs", n_parallel)
        check_parallel(remaining, n_parallel, proxy_list, headless, delay_state, on_result)
    else:
        logger.info("Checking %d URLs", len(remaining))
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=headless,
                args=["--disable-blink-features=AutomationControlled","--no-sandbox"],
                proxy={"server": proxy_list[0]} if proxy_list else None
            )
            ctx  = browser.new_context(
                user_agent=CHROME_UA,
                viewport={"width": 1280, "height": 800}, locale="en-US"
            )
            page = ctx.new_page()
            stealth_sync(page)
            page.goto("https://www.google.com", wait_until="domcontentloaded"); time.sleep(2)
            try:
                btn = page.locator("button:has-text('Accept all'), button:has-text('I agree')")
                if btn.count() > 0: btn.first.click(); time.sleep(1)
            except Exception: pass

            iterator = tqdm(remaining, unit="url") if HAS_TQDM else remaining
            for url in iterator:
                on_result(url, google_check(page, url, delay_state))

            # Retry failed URLs with a fresh browser so a crashed context doesn't block retries
            if failed_urls:
                logger.info("Auto-retrying %d failed URLs", len(failed_urls))
                try:
                    with sync_playwright() as retry_pw:
                        retry_browser = retry_pw.chromium.launch(
                            headless=headless,
                            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
                            proxy={"server": proxy_list[0]} if proxy_list else None
                        )
                        retry_ctx  = retry_browser.new_context(
                            user_agent=CHROME_UA,
                            viewport={"width": 1280, "height": 800}, locale="en-US"
                        )
                        retry_page = retry_ctx.new_page()
                        stealth_sync(retry_page)
                        retry_page.goto("https://www.google.com", wait_until="domcontentloaded")
                        time.sleep(2)
                        for url in failed_urls:
                            status = google_check(retry_page, url, delay_state)
                            for row in all_rows:
                                if row["url"] == url:
                                    old = row["status"]
                                    counts[old] = max(0, counts.get(old, 0) - 1)
                                    counts[status] = counts.get(status, 0) + 1
                                    row["status"] = status
                                    current_results[url] = status
                                    _save_counter[0] += 1
                                    if _save_counter[0] % 10 == 0:
                                        save_progress(pf, {**full_saved, **current_results})
                                    break
                            with _print_lock:
                                logger.info("[retry] %s -> %s", url, status)
                        try:
                            retry_browser.close()
                        except Exception:
                            pass
                except Exception as e:
                    with _print_lock:
                        logger.warning("Retry pass failed: %s", e)

            try:
                browser.close()
            except Exception:
                pass

    if pf.exists(): pf.unlink()
    return all_rows, counts, type_summary, current_results


# ══════════════════════════════════════════════════════════════════════════════
# SCHEDULE
# ══════════════════════════════════════════════════════════════════════════════

def scheduled_run():
    cfg  = _get_cfg().get("schedule", {})
    surl = cfg.get("sitemap_url", "")
    if not surl:
        print(f"{RED}[SCHEDULED] No sitemap_url in config — skipping run{RESET}")
        return
    lim  = cfg.get("limit", 100)
    urls = fetch_sitemap_urls(surl)[:lim]
    print(f"\n{CYAN}[SCHEDULED] Running at {datetime.now().strftime('%H:%M:%S')}…{RESET}")
    execute_and_save(urls)

def start_scheduler():
    if not HAS_SCHEDULE:
        print(f"{RED}schedule not installed{RESET}"); return
    cfg      = _get_cfg().get("schedule", {})
    interval = cfg.get("interval", "daily")
    t        = cfg.get("time", "08:00")
    if interval == "daily":
        sched.every().day.at(t).do(scheduled_run)
    elif interval == "weekly":
        sched.every().monday.at(t).do(scheduled_run)
    elif interval == "hourly":
        sched.every().hour.do(scheduled_run)
    print(f"{GREEN}Scheduler started — {interval} at {t}{RESET}")
    while True:
        sched.run_pending(); time.sleep(30)


def execute_and_save(urls: list[str], headless: bool = False, quiet: bool = False,
                     use_gsc: bool | None = None, do_compare: bool = False,
                     prev_report: Path | None = None,
                     progress_cb: Callable | None = None) -> Path:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path    = REPORTS_DIR / f"indexing_report_{timestamp}.csv"
    excel_path  = REPORTS_DIR / f"indexing_report_{timestamp}.xlsx"
    html_path   = REPORTS_DIR / f"indexing_report_{timestamp}.html"

    all_rows, counts, type_summary, current_results = run_check(
        urls, use_gsc=use_gsc, headless=headless, quiet=quiet, progress_cb=progress_cb
    )

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["#","URL","Google Indexed","Priority","Depth","URL Type","Checked At"])
        for r in all_rows:
            w.writerow([r["num"],r["url"],r["status"],r["priority"],r["depth"],r["url_type"],r["checked_at"]])

    compare_data = compare_runs(current_results, prev_report) if do_compare and prev_report else None

    save_history(counts, len(all_rows), timestamp)
    history = load_history()

    logger.info("Generating reports")
    _gen = ReportGenerator()
    try:
        _gen.excel(all_rows, excel_path, type_summary)
        _gen.html(all_rows, html_path, counts, type_summary, compare_data, history)
    except Exception:
        # Clean up any partially-written companion files so the CSV (primary
        # discovery file) is never left with orphaned siblings on disk.
        for _p in (excel_path, html_path):
            if _p.exists():
                try: _p.unlink()
                except OSError: pass
        raise
    logger.info("CSV report written: %s", csv_path)

    NotificationService(CFG).notify_all(counts, len(all_rows), html_path)

    total = len(all_rows)
    errs = sum(v for k, v in counts.items() if k not in ("Indexed", "Not Indexed"))
    logger.info(
        "SUMMARY total=%d indexed=%d not_indexed=%d errors=%d",
        total,
        counts.get("Indexed", 0),
        counts.get("Not Indexed", 0),
        errs,
    )
    for t, c in sorted(type_summary.items()):
        logger.info(
            "  by_type %s indexed=%d not_indexed=%d",
            t, c.get("Indexed", 0), c.get("Not Indexed", 0),
        )
    if compare_data:
        logger.info(
            "Changes: newly_indexed=%d newly_deindexed=%d",
            len(compare_data["newly_indexed"]),
            len(compare_data["newly_deindexed"]),
        )
    logger.info("Report ready: %s", html_path)
    return html_path


# ══════════════════════════════════════════════════════════════════════════════
# MAIN CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    from core.version import CLI_BANNER
    print(f"\n{BOLD}{CYAN}{'='*56}")
    print(f"  {CLI_BANNER} — Indexing Checker")
    print("   Parallel · GSC · Proxy · Schedule · Notify")
    print(f"{'='*56}{RESET}\n")

    urls = get_urls_from_user()
    if not urls: print(f"{RED}No URLs.{RESET}"); sys.exit(1)
    print(f"\n{GREEN}Found {len(urls)} URLs.{RESET}")

    pattern = input("Filter by URL pattern (e.g. /category/) or Enter to skip: ").strip()
    urls    = filter_urls(urls, pattern)
    if pattern: print(f"{CYAN}Filtered to {len(urls)} URLs{RESET}")

    # Resume
    pf      = progress_file_for(",".join(sorted(urls)[:3]))
    saved   = load_progress(pf)
    if saved:
        print(f"\n{YELLOW}Previous progress: {len(saved)} URLs checked.{RESET}")
        if input("Resume? (y/n): ").strip().lower() != "y":
            saved = {}; pf.unlink(missing_ok=True)

    remaining = [u for u in urls if u not in saved]
    print(f"{GREEN}{len(remaining)} URLs to check.{RESET}")

    raw = input("How many to check? (number or 'all'): ").strip()
    if raw.lower() != "all":
        try: remaining = remaining[:int(raw)]
        except ValueError: remaining = remaining[:10]

    # Options
    use_gsc  = _get_cfg().get("gsc", {}).get("enabled") and input("Use GSC API? (y/n): ").strip().lower() == "y"
    quiet    = input("Show only Not Indexed? (y/n): ").strip().lower() == "y"
    headless = input("Run browser in background? (y/n): ").strip().lower() == "y"

    # Schedule
    if _get_cfg().get("schedule", {}).get("enabled"):
        if input("Run on schedule instead? (y/n): ").strip().lower() == "y":
            start_scheduler(); return

    # Compare
    prev = find_latest_report()
    do_compare = False
    if prev:
        print(f"\n{CYAN}Previous report: {prev.name}{RESET}")
        do_compare = input("Compare with previous run? (y/n): ").strip().lower() == "y"

    execute_and_save(remaining, headless=headless, quiet=quiet,
                     use_gsc=use_gsc, do_compare=do_compare, prev_report=prev)


if __name__ == "__main__":
    main()
