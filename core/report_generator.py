"""
ReportGenerator — build Excel + HTML indexing reports.
Extracted from core/checker.py for single-responsibility.
"""

from __future__ import annotations

import json
import logging
import random
import re
from datetime import datetime
from pathlib import Path

from core.security import esc

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate Excel and HTML reports from indexing run results.

    Usage::

        gen = ReportGenerator()
        gen.excel(rows, path, type_summary)
        gen.html(rows, path, counts, type_summary, compare_data, history)
    """

    # ── Excel ─────────────────────────────────────────────────────────────────

    def excel(self, rows: list[dict], path: Path, type_summary: dict) -> None:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            logger.warning("openpyxl not installed — skipping Excel report")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        HDR = PatternFill("solid", fgColor="1F4E79")
        GF  = PatternFill("solid", fgColor="C6EFCE")
        RF  = PatternFill("solid", fgColor="FFC7CE")
        YF  = PatternFill("solid", fgColor="FFEB9C")

        headers = ["#", "URL", "Status", "Priority", "Depth", "URL Type", "Checked At"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HDR
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([row["num"], row["url"], row["status"], row["priority"],
                       row["depth"], row["url_type"], row["checked_at"]])
            c = ws.cell(ws.max_row, 3)
            c.fill = GF if row["status"] == "Indexed" else RF if row["status"] == "Not Indexed" else YF

        for col, w in zip("ABCDEFG", [5, 65, 16, 10, 8, 22, 22]):
            ws.column_dimensions[col].width = w

        ws2 = wb.create_sheet("By URL Type")
        ws2.append(["URL Type", "Indexed", "Not Indexed", "Other", "Total"])
        for cell in ws2[1]:
            cell.fill = HDR
            cell.font = Font(bold=True, color="FFFFFF")
        for t, c in sorted(type_summary.items()):
            ws2.append([t, c.get("Indexed", 0), c.get("Not Indexed", 0),
                        c.get("Other", 0), sum(c.values())])
        for col in "ABCDE":
            ws2.column_dimensions[col].width = 22

        wb.save(path)
        logger.info("Excel report saved: %s", path)

    # ── HTML ──────────────────────────────────────────────────────────────────

    def html(self, rows: list[dict], path: Path, counts: dict,
             type_summary: dict, compare: dict | None, history: list[dict]) -> None:
        indexed     = counts.get("Indexed", 0)
        not_indexed = counts.get("Not Indexed", 0)
        other       = sum(v for k, v in counts.items() if k not in ("Indexed", "Not Indexed"))
        total       = indexed + not_indexed + other
        pct         = round(indexed / total * 100, 1) if total else 0

        t_labels  = json.dumps(list(type_summary.keys()))
        t_indexed = json.dumps([v.get("Indexed", 0) for v in type_summary.values()])
        t_not     = json.dumps([v.get("Not Indexed", 0) for v in type_summary.values()])

        hist_dates   = json.dumps([h["date"][:10] for h in history])
        hist_indexed = json.dumps([h["indexed"] for h in history])
        hist_not     = json.dumps([h["not_indexed"] for h in history])

        compare_html = self._compare_section(compare)
        trend_section = self._trend_section(history)
        rows_html = self._rows_html(rows)

        html = self._build_html(
            total=total, indexed=indexed, not_indexed=not_indexed, other=other, pct=pct,
            t_labels=t_labels, t_indexed=t_indexed, t_not=t_not,
            hist_dates=hist_dates, hist_indexed=hist_indexed, hist_not=hist_not,
            compare_html=compare_html, trend_section=trend_section, rows_html=rows_html,
            history=history,
        )
        path.write_text(html, encoding="utf-8")
        logger.info("HTML report saved: %s", path)

    # ── Private helpers ───────────────────────────────────────────────────────

    @staticmethod
    def _compare_section(compare: dict | None) -> str:
        if not compare:
            return ""
        ni, nd = compare["newly_indexed"], compare["newly_deindexed"]
        ni_li = "".join(f"<li><a href='{esc(u)}' target='_blank'>{esc(u)}</a></li>" for u in ni)
        nd_li = "".join(f"<li><a href='{esc(u)}' target='_blank'>{esc(u)}</a></li>" for u in nd)
        return f"""
        <div class="section">
          <h2>Changes Since Last Run</h2>
          <div class="summary-cards" style="margin-bottom:16px">
            <div class="card green"><span class="num">{len(ni)}</span><span class="lbl">Newly Indexed</span></div>
            <div class="card red"><span class="num">{len(nd)}</span><span class="lbl">Newly De-indexed</span></div>
          </div>
          {"<h3 class='change-head green-txt'>Newly Indexed</h3><ul class='change-list'>" + ni_li + "</ul>" if ni else ""}
          {"<h3 class='change-head red-txt'>Newly De-indexed</h3><ul class='change-list red-list'>" + nd_li + "</ul>" if nd else ""}
        </div>"""

    @staticmethod
    def _trend_section(history: list[dict]) -> str:
        if len(history) <= 1:
            return ""
        return f"""
        <div class="section">
          <h2>Trend — Last {len(history)} Runs</h2>
          <canvas id="trend" height="80"></canvas>
        </div>"""

    @staticmethod
    def _rows_html(rows: list[dict]) -> str:
        parts = []
        for r in rows:
            cls = "indexed" if r["status"] == "Indexed" else "not-indexed" if r["status"] == "Not Indexed" else "other"
            icon = "✅" if r["status"] == "Indexed" else "❌" if r["status"] == "Not Indexed" else "⚠️"
            pri = r["priority"]
            badge = f"badge-{'high' if pri == 'High' else 'med' if pri == 'Medium' else 'low'}"
            parts.append(
                f"<tr class='{cls}'>"
                f"<td>{esc(r['num'])}</td>"
                f"<td><a href='{esc(r['url'])}' target='_blank'>{esc(r['url'])}</a></td>"
                f"<td>{icon} {esc(r['status'])}</td>"
                f"<td><span class='badge {badge}'>{esc(pri)}</span></td>"
                f"<td>{esc(r['depth'])}</td>"
                f"<td>{esc(r['url_type'])}</td>"
                f"<td>{esc(r['checked_at'])}</td></tr>"
            )
        return "".join(parts)

    @staticmethod
    def _build_html(**kw) -> str:
        total = kw["total"]; indexed = kw["indexed"]; not_indexed = kw["not_indexed"]
        other = kw["other"]; pct = kw["pct"]
        trend_js = ""
        if len(kw["history"]) > 1:
            trend_js = (
                "new Chart(document.getElementById('trend'),{type:'line',"
                f"data:{{labels:{kw['hist_dates']},datasets:["
                f"{{label:'Indexed',data:{kw['hist_indexed']},borderColor:'#1a7f37',"
                "backgroundColor:'rgba(26,127,55,.1)',tension:.4,fill:true},"
                f"{{label:'Not Indexed',data:{kw['hist_not']},borderColor:'#cf222e',"
                "backgroundColor:'rgba(207,34,46,.1)',tension:.4,fill:true}]}},"
                "options:{responsive:true,plugins:{legend:{position:'bottom'}},"
                "scales:{y:{beginAtZero:true}}}});"
            )
        other_btn = (f"<button class='btn' onclick=\"filterRows('other',this)\">⚠️ Other ({other})</button>"
                     if other else "")
        return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Indexing Report {datetime.now().strftime('%Y-%m-%d')}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<link rel="stylesheet" href="https://cdn.datatables.net/1.13.8/css/jquery.dataTables.min.css">
<script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
<script src="https://cdn.datatables.net/1.13.8/js/jquery.dataTables.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;color:#1a1a2e}}
header{{background:#1F4E79;color:#fff;padding:20px 40px;display:flex;justify-content:space-between;align-items:center}}
header h1{{font-size:20px;font-weight:700}}
header p{{opacity:.7;font-size:13px}}
.container{{max-width:1300px;margin:24px auto;padding:0 24px}}
.summary-cards{{display:grid;grid-template-columns:repeat(5,1fr);gap:14px;margin-bottom:24px}}
.card{{background:#fff;border-radius:12px;padding:18px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.06);display:flex;flex-direction:column;gap:4px}}
.card .num{{font-size:34px;font-weight:800}}
.card .lbl{{font-size:11px;color:#666;text-transform:uppercase;letter-spacing:.5px}}
.card.blue .num{{color:#0969da}}.card.green .num{{color:#1a7f37}}.card.red .num{{color:#cf222e}}
.card.yellow .num{{color:#9a6700}}.card.purple .num{{color:#6f42c1}}
.section{{background:#fff;border-radius:12px;padding:22px;margin-bottom:22px;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
.section h2{{font-size:15px;font-weight:700;color:#1F4E79;margin-bottom:14px}}
.charts{{display:grid;grid-template-columns:1fr 2fr;gap:22px;margin-bottom:22px}}
.filters{{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap;align-items:center}}
.btn{{padding:6px 14px;border:none;border-radius:20px;cursor:pointer;font-size:12px;background:#e8eaf0;color:#444}}
.btn.active{{background:#1F4E79;color:#fff}}
table{{width:100%;border-collapse:collapse}}
th{{background:#1F4E79;color:#fff;padding:10px 13px;text-align:left;font-size:12px}}
td{{padding:9px 13px;font-size:12px;border-bottom:1px solid #f0f2f5;vertical-align:middle}}
td a{{color:#0969da;text-decoration:none;word-break:break-all}}
td a:hover{{text-decoration:underline}}
tr.indexed td:nth-child(3){{color:#1a7f37;font-weight:600}}
tr.not-indexed td:nth-child(3){{color:#cf222e;font-weight:600}}
tr:hover td{{background:#f8f9fb}}
.badge{{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}}
.badge-high{{background:#d4edda;color:#155724}}
.badge-med{{background:#fff3cd;color:#856404}}
.badge-low{{background:#f8d7da;color:#721c24}}
.pct-bar{{background:#e8eaf0;border-radius:10px;height:8px;margin-top:8px;overflow:hidden}}
.pct-fill{{background:#1a7f37;height:100%;border-radius:10px;width:{pct}%}}
.dataTables_wrapper .dataTables_length select,
.dataTables_wrapper .dataTables_filter input{{border:1px solid #ddd;border-radius:8px;padding:5px 10px;font-size:12px}}
.dataTables_wrapper .dataTables_paginate .paginate_button.current{{background:#1F4E79;color:#fff!important;border-color:#1F4E79}}
.dataTables_wrapper .dataTables_info{{font-size:12px;color:#666}}
</style>
</head>
<body>
<header>
  <div><h1>Google Indexing Status Report</h1><p>Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p></div>
  <div style="text-align:right;color:white"><div style="font-size:28px;font-weight:800">{pct}%</div><div style="opacity:.7;font-size:12px">Index Rate</div></div>
</header>
<div class="container">
  <div class="summary-cards">
    <div class="card blue"><span class="num">{total}</span><span class="lbl">Total Checked</span></div>
    <div class="card green"><span class="num">{indexed}</span><span class="lbl">Indexed</span></div>
    <div class="card red"><span class="num">{not_indexed}</span><span class="lbl">Not Indexed</span></div>
    <div class="card yellow"><span class="num">{other}</span><span class="lbl">Errors/Other</span></div>
    <div class="card purple"><span class="num">{pct}%</span><span class="lbl">Index Rate<div class="pct-bar"><div class="pct-fill"></div></div></span></div>
  </div>
  {kw['compare_html']}
  {kw['trend_section']}
  <div class="charts">
    <div class="section" style="margin:0"><h2>Indexed vs Not Indexed</h2><canvas id="pie"></canvas></div>
    <div class="section" style="margin:0"><h2>By URL Type</h2><canvas id="bar"></canvas></div>
  </div>
  <div class="section">
    <h2>All Results</h2>
    <div class="filters">
      <button class="btn active" onclick="filterRows('all',this)">All ({total})</button>
      <button class="btn" onclick="filterRows('indexed',this)">✅ Indexed ({indexed})</button>
      <button class="btn" onclick="filterRows('not-indexed',this)">❌ Not Indexed ({not_indexed})</button>
      {other_btn}
      <button class="btn" onclick="filterRows('high-pri',this)">🔴 High Priority</button>
    </div>
    <table id="results-table">
      <thead><tr><th>#</th><th>URL</th><th>Status</th><th>Priority</th><th>Depth</th><th>Type</th><th>Checked At</th></tr></thead>
      <tbody id="tb">{kw['rows_html']}</tbody>
    </table>
  </div>
</div>
<script>
new Chart(document.getElementById('pie'),{{
  type:'doughnut',
  data:{{labels:['Indexed','Not Indexed','Other'],datasets:[{{data:[{indexed},{not_indexed},{other}],backgroundColor:['#1a7f37','#cf222e','#9a6700'],borderWidth:0}}]}},
  options:{{plugins:{{legend:{{position:'bottom'}}}},cutout:'60%'}}
}});
new Chart(document.getElementById('bar'),{{
  type:'bar',
  data:{{labels:{kw['t_labels']},datasets:[{{label:'Indexed',data:{kw['t_indexed']},backgroundColor:'#1a7f37'}},{{label:'Not Indexed',data:{kw['t_not']},backgroundColor:'#cf222e'}}]}},
  options:{{responsive:true,plugins:{{legend:{{position:'bottom'}}}},scales:{{x:{{ticks:{{maxRotation:45}}}},y:{{beginAtZero:true}}}}}}
}});
{trend_js}
var _activeFilter='all';
$(document).ready(function(){{
  var table=$('#results-table').DataTable({{
    pageLength:50,
    lengthMenu:[[25,50,100,250,-1],['25','50','100','250','All']],
    order:[[0,'asc']],
    language:{{search:'🔍 Search:',lengthMenu:'Show _MENU_ rows',info:'Showing _START_–_END_ of _TOTAL_ URLs',paginate:{{previous:'‹',next:'›'}}}}
  }});
  $.fn.dataTable.ext.search.push(function(settings,data,dataIndex){{
    var row=$(table.row(dataIndex).node());
    if(_activeFilter==='all') return true;
    if(_activeFilter==='high-pri') return row.find('td:nth-child(4)').text().trim()==='High';
    return row.hasClass(_activeFilter);
  }});
  window._dt=table;
}});
function filterRows(type,btn){{
  document.querySelectorAll('.btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  _activeFilter=type;
  window._dt.draw();
}}
</script>
</body>
</html>"""
