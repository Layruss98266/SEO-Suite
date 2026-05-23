"""
SEO Audit Orchestrator
Runs all Phase 1-4 tools and generates a comprehensive HTML + Excel + CSV report.
"""

import csv
import json
import logging
import sys
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

try:
    from colorama import Fore, Style, init
    init(autoreset=True)
    GREEN = Fore.GREEN; RED = Fore.RED; YELLOW = Fore.YELLOW
    CYAN  = Fore.CYAN;  BOLD = Style.BRIGHT; RESET = Style.RESET_ALL
except ImportError:
    GREEN = RED = YELLOW = CYAN = BOLD = RESET = ""

try:
    from tqdm import tqdm; HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

from core.checker import (
    build_gsc_service,
    fetch_from_domain,
    fetch_sitemap_urls,
    load_config,
    load_from_csv_excel,
)
from core.security import esc, validate_public_url
from tools.phase2 import audit_url as p2_audit
from tools.phase3 import (
    clicks_impressions,
    coverage_errors,
    ctr_analyzer,
    manual_actions,
    position_tracker,
    sitemaps_status,
    top_queries,
)
from tools.phase4 import (
    backlink_check,
    broken_backlinks,
    competitor_comparison,
    domain_authority,
    domain_rank,
    keyword_rank_tracker,
    nofollow_ratio,
    page_authority,
    rank_change,
    referring_domains,
    serp_features,
    spam_score,
    traffic_share,
)

_PROJECT_ROOT = Path(__file__).parent.parent.resolve()
REPORTS_DIR = _PROJECT_ROOT / "data" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
CFG = load_config()
logger = logging.getLogger(__name__)

STATUS_ICON  = {"pass": "✅", "warning": "⚠️", "fail": "❌", "error": "🔴"}
STATUS_COLOR = {"pass": "#1a7f37", "warning": "#9a6700", "fail": "#cf222e", "error": "#6e40c9"}
STATUS_BG    = {"pass": "#d4edda",  "warning": "#fff3cd", "fail": "#f8d7da",  "error": "#e8d5fb"}

_ACRONYMS = {
    "Ssl":"SSL","Http":"HTTP","Og":"OG","Dns":"DNS","Url":"URL","Gsc":"GSC",
    "Lcp":"LCP","Cls":"CLS","Fid":"FID","Fcp":"FCP","Tti":"TTI","Ttfb":"TTFB",
    "Ctr":"CTR","Js":"JS","Css":"CSS","Cdn":"CDN","Api":"API","Id":"ID",
    "H1":"H1","H2":"H2","H3":"H3","Spf":"SPF","Dmarc":"DMARC","Mx":"MX",
    "Ai":"AI","Ui":"UI","Ux":"UX",
}

def _pretty(key: str) -> str:
    parts = key.replace("_", " ").title().split()
    return " ".join(_ACRONYMS.get(p, p) for p in parts)

# ══════════════════════════════════════════════════════════════════════════════
# Use-case registry  (the authoritative definition shared with the dashboard)
# ══════════════════════════════════════════════════════════════════════════════
# Individual tasks per use case — task IDs match the `tool` field returned by checks
TASKS = {
    "crawlability": [
        {"id": "robots",           "label": "robots.txt"},
        {"id": "http_status",      "label": "HTTP status"},
        {"id": "redirect",         "label": "Redirects"},
        {"id": "broken_links",     "label": "Broken links"},
        {"id": "internal_links",   "label": "Internal links"},
        {"id": "sitemap",          "label": "Sitemap validation"},
        {"id": "canonical",        "label": "Canonical URL"},
        {"id": "meta_robots",      "label": "Meta robots / noindex"},
    ],
    "on_page": [
        {"id": "canonical",        "label": "Canonical URL"},
        {"id": "title",            "label": "Title tag"},
        {"id": "meta_description", "label": "Meta description"},
        {"id": "headings",         "label": "Headings (H1–H3)"},
        {"id": "image_alt",        "label": "Image alt text"},
        {"id": "word_count",       "label": "Word count"},
        {"id": "readability",      "label": "Readability"},
        {"id": "schema",           "label": "Schema markup"},
        {"id": "hreflang",         "label": "Hreflang tags"},
        {"id": "ttfb",             "label": "TTFB (response time)"},
        {"id": "og_tags",          "label": "Open Graph / Twitter cards"},
        {"id": "meta_robots",      "label": "Meta robots / X-Robots-Tag"},
        {"id": "favicon",          "label": "Favicon"},
    ],
    "site_health": [
        {"id": "ssl",                "label": "SSL / TLS grade"},
        {"id": "domain_age",         "label": "Domain age & expiry"},
        {"id": "mixed_content",      "label": "Mixed content (HTTP on HTTPS)"},
        {"id": "https_enforcement",  "label": "HTTPS enforcement"},
        {"id": "security_headers",   "label": "Security headers"},
        {"id": "spf",                "label": "SPF record"},
        {"id": "dmarc",              "label": "DMARC policy"},
        {"id": "mx_records",         "label": "MX records"},
    ],
    "performance": [
        {"id": "pagespeed_mobile",  "label": "PageSpeed — Mobile"},
        {"id": "pagespeed_desktop", "label": "PageSpeed — Desktop"},
        {"id": "mobile",            "label": "Mobile vs Desktop gap"},
        {"id": "crawlability",      "label": "GSC URL inspection"},
        {"id": "lcp",               "label": "LCP — Largest Contentful Paint"},
        {"id": "cls",               "label": "CLS — Cumulative Layout Shift"},
        {"id": "fcp",               "label": "FCP — First Contentful Paint"},
        {"id": "inp",               "label": "INP — Interaction to Next Paint"},
    ],
    "search_console": [
        {"id": "clicks_impressions", "label": "Clicks & impressions"},
        {"id": "top_queries",        "label": "Top queries"},
        {"id": "position_tracker",   "label": "Position tracker"},
        {"id": "ctr_analyzer",       "label": "CTR analyzer"},
        {"id": "coverage_errors",    "label": "Coverage errors"},
        {"id": "sitemaps_status",    "label": "Sitemaps status"},
        {"id": "manual_actions",     "label": "Manual actions"},
    ],
    "authority": [
        {"id": "backlinks",          "label": "Backlinks"},
        {"id": "domain_authority",   "label": "Domain authority (DA)"},
        {"id": "page_authority",     "label": "Page authority (PA)"},
        {"id": "referring_domains",  "label": "Referring domains"},
        {"id": "domain_rank",        "label": "Domain rank (DataForSEO)"},
        {"id": "broken_backlinks",   "label": "Broken backlinks"},
        {"id": "nofollow_ratio",     "label": "Nofollow / dofollow ratio"},
        {"id": "spam_score",         "label": "Spam score"},
    ],
    "rankings": [
        {"id": "rank_tracker",   "label": "Keyword rankings"},
        {"id": "serp_features",  "label": "SERP features"},
        {"id": "competitor",     "label": "Competitor analysis"},
        {"id": "rank_change",    "label": "Rank change vs last run"},
        {"id": "traffic_share",  "label": "Estimated traffic share"},
    ],
}

USE_CASES = {
    "crawlability": {
        "label": "Crawl Access",
        "icon": "🕷️",
        "color": "#6366F1",
        "bg": "#EEF2FF",
        "desc": "robots.txt · HTTP status · redirects · broken links · sitemaps",
        "requires": None,
        "group": "Technical",
    },
    "on_page": {
        "label": "On-Page SEO",
        "icon": "📝",
        "color": "#0EA5E9",
        "bg": "#F0F9FF",
        "desc": "Title · meta · headings · readability · OG · schema · favicon",
        "requires": None,
        "group": "Technical",
    },
    "site_health": {
        "label": "Site Health",
        "icon": "🛡️",
        "color": "#14B8A6",
        "bg": "#F0FDFA",
        "desc": "SSL · domain age · DNS · mixed content · archive history",
        "requires": None,
        "group": "Technical",
    },
    "performance": {
        "label": "Performance",
        "icon": "⚡",
        "color": "#F59E0B",
        "bg": "#FFFBEB",
        "desc": "PageSpeed · Core Web Vitals · TTFB · mobile friendliness",
        "requires": "pagespeed_api_key",
        "group": "Technical",
    },
    "search_console": {
        "label": "Search Console",
        "icon": "📊",
        "color": "#10B981",
        "bg": "#ECFDF5",
        "desc": "Clicks · impressions · positions · top queries · coverage",
        "requires": "gsc",
        "group": "Visibility",
    },
    "authority": {
        "label": "Authority",
        "icon": "🔗",
        "color": "#8B5CF6",
        "bg": "#F5F3FF",
        "desc": "Backlink count · domain authority · referring domains",
        "requires": None,
        "group": "Visibility",
    },
    "rankings": {
        "label": "Rankings",
        "icon": "🏆",
        "color": "#EF4444",
        "bg": "#FEF2F2",
        "desc": "Keyword positions · SERP rank · competitor comparison",
        "requires": "serpapi_key",
        "group": "Visibility",
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# SEO Score calculator
# ══════════════════════════════════════════════════════════════════════════════
WEIGHTS = {
    "http_status":      15,
    "robots":           10,
    "meta_robots":      10,
    "title":            10,
    "meta_description":  8,
    "headings":          8,
    "canonical":         7,
    "word_count":        7,
    "image_alt":         6,
    "ssl":               6,
    "schema":            5,
    "ttfb":              5,
    "redirect":          5,
    "broken_links":      5,
    "mixed_content":     5,
    "internal_links":    4,
    "og_tags":           4,
    "readability":       4,
    "dns_health":        3,
    "favicon":           2,
    "domain_age":        2,
    "hreflang":          2,
    "sitemap":           3,
}

def calc_seo_score(results: list[dict]) -> int:
    total_weight  = 0
    earned_weight = 0
    for r in results:
        w = WEIGHTS.get(r["tool"], 2)
        total_weight += w
        if r["status"] == "pass":    earned_weight += w
        elif r["status"] == "warning": earned_weight += w * 0.5
    return round(earned_weight / total_weight * 100) if total_weight else 0


# ══════════════════════════════════════════════════════════════════════════════
# Run audit for a single URL
# ══════════════════════════════════════════════════════════════════════════════
def audit_single_url(url: str, cfg: dict, gsc_service=None,
                     use_cases: list | None = None,
                     tasks: list | None = None,
                     phases: dict | None = None) -> dict:
    """
    Run the selected use-cases (and optionally a subset of tasks) against a URL.
    `tasks` is a flat list of task IDs that overrides use-case-default behavior.
    Passing neither runs everything.
    """
    # Refuse private/internal hosts up front rather than crashing somewhere
    # downstream in phase1's `requests.get`. Returning a structured result
    # keeps the audit pipeline shape consistent.
    try:
        url = validate_public_url(url)
    except ValueError as exc:
        # Generic client-facing message — the raw exception can leak internal
        # hostnames / resolved IPs. Full reason still goes to server logs.
        import logging as _log
        _log.getLogger(__name__).info("URL refused by SSRF guard: %s (%s)", url, exc)
        err = {"tool": "url_validation", "status": "error",
               "message": "URL refused (private, local, or non-public host)",
               "value": None, "details": {}}
        return {
            "url": url, "score": 0,
            "results": [err],
            "issues": [err], "warnings": [], "passed": [],
            "use_cases": [],
            "counts": {"pass": 0, "warning": 0, "fail": 0, "error": 1},
        }
    if use_cases is None and phases is not None:
        uc = []
        if phases.get("p1"): uc += ["crawlability", "on_page"]
        if phases.get("p2"): uc.append("performance")
        if phases.get("p3"): uc.append("search_console")
        if phases.get("p4"): uc += ["authority", "rankings"]
        use_cases = uc
    elif use_cases is None:
        use_cases = list(USE_CASES.keys())

    active = set(use_cases)
    task_filter = set(tasks) if tasks else None  # None = no filtering
    def _keep(tool_id: str) -> bool:
        return task_filter is None or tool_id in task_filter

    all_results: list[dict] = []

    if "crawlability" in active:
        from tools.phase1 import (
            broken_link_check,
            canonical_check,
            http_status_check,
            internal_links_check,
            meta_robots_check,
            redirect_check,
            robots_check,
            sitemap_validate,
        )
        crawl_fns = [
            (robots_check,         "robots"),
            (http_status_check,    "http_status"),
            (redirect_check,       "redirect"),
            (broken_link_check,    "broken_links"),
            (internal_links_check, "internal_links"),
            (sitemap_validate,     "sitemap"),
            (canonical_check,      "canonical"),
            (meta_robots_check,    "meta_robots"),
        ]
        crawl_fns = [(fn, tid) for fn, tid in crawl_fns if _keep(tid)]
        if crawl_fns:
            with ThreadPoolExecutor(max_workers=min(len(crawl_fns), 8)) as ex:
                futs = [ex.submit(fn, url) for fn, _ in crawl_fns]
                all_results += [f.result() for f in futs]

    if "on_page" in active:
        from tools.phase1 import (
            canonical_check,
            favicon_check,
            heading_check,
            hreflang_check,
            image_alt_check,
            meta_description_check,
            meta_robots_check,
            og_check,
            readability_check,
            schema_check,
            title_check,
            ttfb_check,
            word_count_check,
        )
        onpage_fns = [
            (canonical_check,        "canonical"),
            (title_check,            "title"),
            (meta_description_check, "meta_description"),
            (heading_check,          "headings"),
            (image_alt_check,        "image_alt"),
            (word_count_check,       "word_count"),
            (readability_check,      "readability"),
            (schema_check,           "schema"),
            (hreflang_check,         "hreflang"),
            (ttfb_check,             "ttfb"),
            (og_check,               "og_tags"),
            (meta_robots_check,      "meta_robots"),
            (favicon_check,          "favicon"),
        ]
        onpage_fns = [(fn, tid) for fn, tid in onpage_fns if _keep(tid)]
        if onpage_fns:
            with ThreadPoolExecutor(max_workers=min(len(onpage_fns), 8)) as ex:
                futs = [ex.submit(fn, url) for fn, _ in onpage_fns]
                all_results += [f.result() for f in futs]

    if "site_health" in active:
        from tools.phase1 import (
            dmarc_check,
            domain_age_check,
            https_enforcement_check,
            mixed_content_check,
            mx_records_check,
            security_headers_check,
            spf_check,
            ssl_check,
        )
        sh_fns = [
            (ssl_check,                "ssl"),
            (domain_age_check,         "domain_age"),
            (mixed_content_check,      "mixed_content"),
            (https_enforcement_check,  "https_enforcement"),
            (security_headers_check,   "security_headers"),
            (spf_check,                "spf"),
            (dmarc_check,              "dmarc"),
            (mx_records_check,         "mx_records"),
        ]
        sh_fns = [(fn, tid) for fn, tid in sh_fns if _keep(tid)]
        if sh_fns:
            with ThreadPoolExecutor(max_workers=min(len(sh_fns), 8)) as ex:
                futs = [ex.submit(fn, url) for fn, _ in sh_fns]
                all_results += [f.result() for f in futs]

    if "performance" in active:
        from tools.phase2 import extract_cwv
        api_key  = cfg.get("pagespeed_api_key", "")
        _parsed  = urlparse(url)
        site_url = f"{_parsed.scheme}://{_parsed.netloc}/" if _parsed.netloc else ""
        if api_key:
            perf = p2_audit(url, api_key=api_key, gsc_service=gsc_service, site_url=site_url)
            # Extract Core Web Vitals as separate result items from mobile pagespeed
            mob_ps = next((r for r in perf if r.get("tool") == "pagespeed_mobile"), None)
            if mob_ps:
                cwv = [c for c in extract_cwv(mob_ps) if _keep(c["tool"])]
                perf = perf + cwv
            if task_filter is not None:
                perf = [r for r in perf if r.get("tool") in task_filter]
            all_results += perf

    if "search_console" in active and gsc_service:
        _parsed  = urlparse(url)
        site_url = f"{_parsed.scheme}://{_parsed.netloc}/" if _parsed.netloc else ""
        if site_url:
            # url-scoped: need both url + service + site_url
            url_fns: list = []
            if _keep("clicks_impressions"): url_fns.append(("clicks_impressions", clicks_impressions))
            if _keep("top_queries"):        url_fns.append(("top_queries",        top_queries))
            if _keep("position_tracker"):   url_fns.append(("position_tracker",   position_tracker))
            if _keep("manual_actions"):     url_fns.append(("manual_actions",     manual_actions))
            # site-scoped: only service + site_url
            site_fns: list = []
            if _keep("ctr_analyzer"):       site_fns.append(("ctr_analyzer",      ctr_analyzer))
            if _keep("coverage_errors"):    site_fns.append(("coverage_errors",   coverage_errors))
            if _keep("sitemaps_status"):    site_fns.append(("sitemaps_status",   sitemaps_status))
            workers = len(url_fns) + len(site_fns)
            if workers:
                with ThreadPoolExecutor(max_workers=workers) as ex:
                    futs  = [ex.submit(fn, url, gsc_service, site_url) for _, fn in url_fns]
                    futs += [ex.submit(fn, gsc_service, site_url)      for _, fn in site_fns]
                    all_results += [f.result() for f in futs]

    if "authority" in active:
        dfs_login    = cfg.get("dataforseo_login", "")
        dfs_password = cfg.get("dataforseo_password", "")
        moz_id       = cfg.get("moz_access_id", "")
        moz_key      = cfg.get("moz_secret_key", "")
        au_fns = []
        if _keep("backlinks"):
            au_fns.append(lambda: backlink_check(url, dataforseo_login=dfs_login,
                                                  dataforseo_password=dfs_password))
        if _keep("domain_authority"):
            au_fns.append(lambda: domain_authority(url, moz_id, moz_key))
        if _keep("page_authority"):
            au_fns.append(lambda: page_authority(url, moz_id, moz_key))
        if _keep("referring_domains"):
            au_fns.append(lambda: referring_domains(url, dfs_login, dfs_password))
        if _keep("domain_rank"):
            au_fns.append(lambda: domain_rank(url, dfs_login, dfs_password))
        if _keep("broken_backlinks"):
            au_fns.append(lambda: broken_backlinks(url, dfs_login, dfs_password))
        if _keep("nofollow_ratio"):
            au_fns.append(lambda: nofollow_ratio(url, dfs_login, dfs_password))
        if _keep("spam_score"):
            au_fns.append(lambda: spam_score(url, moz_id, moz_key))
        if au_fns:
            with ThreadPoolExecutor(max_workers=min(len(au_fns), 4)) as ex:
                futs = [ex.submit(fn) for fn in au_fns]
                all_results += [f.result() for f in futs]

    if "rankings" in active:
        keywords    = cfg.get("track_keywords", [])
        serp_key    = cfg.get("serpapi_key", "")
        dfs_login   = cfg.get("dataforseo_login", "")
        dfs_pass    = cfg.get("dataforseo_password", "")
        rk_fns = []
        if _keep("rank_tracker") and keywords:
            rk_fns.append(lambda: keyword_rank_tracker(url, keywords,
                serpapi_key=serp_key, dataforseo_login=dfs_login,
                dataforseo_password=dfs_pass))
        if _keep("serp_features") and keywords:
            rk_fns.append(lambda: serp_features(url, keywords,
                serpapi_key=serp_key, dataforseo_login=dfs_login,
                dataforseo_password=dfs_pass))
        if _keep("rank_change") and keywords:
            rk_fns.append(lambda: rank_change(url, keywords,
                serpapi_key=serp_key, dataforseo_login=dfs_login,
                dataforseo_password=dfs_pass))
        if _keep("traffic_share") and keywords:
            rk_fns.append(lambda: traffic_share(url, keywords,
                serpapi_key=serp_key, dataforseo_login=dfs_login,
                dataforseo_password=dfs_pass))
        if _keep("competitor") and keywords:
            rk_fns.append(lambda: competitor_comparison(url, keywords[0],
                serpapi_key=serp_key, dataforseo_login=dfs_login,
                dataforseo_password=dfs_pass))
        if rk_fns:
            with ThreadPoolExecutor(max_workers=min(len(rk_fns), 3)) as ex:
                futs = [ex.submit(fn) for fn in rk_fns]
                all_results += [f.result() for f in futs]
        elif not keywords:
            # Return a warning if no keywords configured
            if _keep("rank_tracker"):
                from tools.phase4 import result as _p4result
                all_results.append(_p4result(url, "rank_tracker", "warning", [],
                    "No keywords configured — add track_keywords in Settings"))

    score    = calc_seo_score(all_results)
    issues   = [r for r in all_results if r["status"] in ("fail", "error")]
    warnings = [r for r in all_results if r["status"] == "warning"]
    passed   = [r for r in all_results if r["status"] == "pass"]

    return {
        "url":       url,
        "score":     score,
        "results":   all_results,
        "issues":    issues,
        "warnings":  warnings,
        "passed":    passed,
        "use_cases": sorted(active),
        "counts":    {"pass": len(passed), "warning": len(warnings), "fail": len(issues)},
    }


# ══════════════════════════════════════════════════════════════════════════════
# Value formatter — renders r["value"] + r["details"] into a readable string
# ══════════════════════════════════════════════════════════════════════════════
def format_value(r: dict) -> str:
    tool    = r.get("tool", "")
    value   = r.get("value")
    details = r.get("details") or {}

    if value is None:
        return "<span style='color:#999'>—</span>"

    if tool == "http_status":
        return f"<b>HTTP {value}</b>"

    if tool == "robots":
        if isinstance(value, dict):
            ok_all  = value.get("allowed", False)
            ok_g    = value.get("googlebot", False)
            delay   = details.get("crawl_delay")
            parts   = []
            parts.append("All bots: " + ("✓ allowed" if ok_all else "✗ blocked"))
            parts.append("Googlebot: " + ("✓ allowed" if ok_g else "✗ blocked"))
            if delay: parts.append(f"Crawl-delay: {delay}s")
            return " &nbsp;·&nbsp; ".join(parts)
        return str(value)

    if tool == "redirect":
        hops  = details.get("hops", 0)
        final = details.get("final_url", "")
        codes = details.get("codes", [])
        if hops == 0: return "No redirect"
        code_str = " → ".join(str(c) for c in codes) if codes else ""
        return f"{hops} hop(s)&nbsp; {esc(code_str)}&nbsp; → <code style='font-size:11px'>{esc(final[:70])}</code>"

    if tool == "canonical":
        if not value: return "<span style='color:#999'>None</span>"
        self_ref = details.get("self_referencing")
        if self_ref: return "✓ Self-referencing"
        return f"→ <code style='font-size:11px'>{esc(str(value)[:80])}</code>"

    if tool == "title":
        if not value: return "<span style='color:#cf222e'>Missing</span>"
        length = details.get("length", len(str(value)))
        opt    = details.get("optimal", False)
        bar_color = "#1a7f37" if opt else "#9a6700"
        bar_pct   = min(100, round(length / 60 * 100))
        return (f'<div style="font-style:italic;color:#333">&ldquo;{esc(str(value)[:70])}{"…" if len(str(value))>70 else ""}&rdquo;</div>'
                f'<div style="margin-top:4px;display:flex;align-items:center;gap:6px">'
                f'<div style="flex:1;height:4px;background:#e2e8f0;border-radius:9px;overflow:hidden">'
                f'<div style="width:{bar_pct}%;height:100%;background:{bar_color}"></div></div>'
                f'<span style="font-size:11px;color:#666">{length} chars</span>'
                f'<span style="font-size:10px;color:#999">(30–60 optimal)</span></div>')

    if tool == "meta_description":
        if not value: return "<span style='color:#cf222e'>Missing</span>"
        length  = details.get("length", len(str(value)))
        opt     = details.get("optimal", False)
        bar_color = "#1a7f37" if opt else "#9a6700"
        bar_pct   = min(100, round(length / 160 * 100))
        return (f'<div style="font-style:italic;color:#333">&ldquo;{esc(str(value)[:100])}{"…" if len(str(value))>100 else ""}&rdquo;</div>'
                f'<div style="margin-top:4px;display:flex;align-items:center;gap:6px">'
                f'<div style="flex:1;height:4px;background:#e2e8f0;border-radius:9px;overflow:hidden">'
                f'<div style="width:{bar_pct}%;height:100%;background:{bar_color}"></div></div>'
                f'<span style="font-size:11px;color:#666">{length} chars</span>'
                f'<span style="font-size:10px;color:#999">(70–160 optimal)</span></div>')

    if tool == "headings":
        if isinstance(value, dict):
            h1s = value.get("h1", [])
            h2s = value.get("h2", [])
            h3s = value.get("h3", [])
            h1_text = f': <i>"{esc(h1s[0][:50])}"</i>' if h1s else " — <span style='color:#cf222e'>none</span>"
            return (f"<b>H1</b>{h1_text}<br>"
                    f"<span style='color:#666'>H2: {len(h2s)}&nbsp;&nbsp;H3: {len(h3s)}</span>")
        return str(value)

    if tool == "image_alt":
        if isinstance(value, dict):
            total = details.get("total", 0)
            miss  = details.get("missing_count", 0)
            pct   = details.get("pct_missing", 0)
            bar_color = "#1a7f37" if miss == 0 else "#cf222e" if pct > 30 else "#9a6700"
            ok_pct    = 100 - pct
            return (f"<div style='display:flex;align-items:center;gap:6px'>"
                    f"<div style='flex:1;height:5px;background:#e2e8f0;border-radius:9px;overflow:hidden'>"
                    f"<div style='width:{ok_pct}%;height:100%;background:{bar_color}'></div></div>"
                    f"<span style='font-size:11px'>{total - miss}/{total} have alt text</span></div>"
                    + (f"<div style='font-size:11px;color:#cf222e;margin-top:2px'>{miss} missing ({pct}%)</div>" if miss else ""))
        return str(value)

    if tool == "word_count":
        try:
            count = int(value) if value is not None else 0
        except (ValueError, TypeError):
            count = 0
        opt_min  = details.get("optimal_min", 600)
        opt_max  = details.get("optimal_max", 2500)
        bar_pct  = min(100, round(count / opt_max * 100))
        bar_color = "#1a7f37" if opt_min <= count <= opt_max else "#9a6700" if count < opt_min else "#9a6700"
        return (f"<b>{count:,} words</b>"
                f"<div style='margin-top:4px;display:flex;align-items:center;gap:6px'>"
                f"<div style='flex:1;height:4px;background:#e2e8f0;border-radius:9px;overflow:hidden'>"
                f"<div style='width:{bar_pct}%;height:100%;background:{bar_color}'></div></div>"
                f"<span style='font-size:10px;color:#999'>optimal {opt_min:,}–{opt_max:,}</span></div>")

    if tool == "broken_links":
        total  = details.get("total_checked", 0)
        broken = details.get("broken_count", 0)
        if not broken:
            return f"<span style='color:#1a7f37'>All {total} links OK</span>"
        return f"<b style='color:#cf222e'>{broken}</b> broken of {total} checked"

    if tool == "internal_links":
        ic = details.get("internal_count", 0)
        ec = details.get("external_count", 0)
        nf = details.get("nofollow_count", 0)
        parts = [f"<b>{ic}</b> internal", f"<b>{ec}</b> external"]
        if nf: parts.append(f"<b>{nf}</b> nofollow")
        return " &nbsp;·&nbsp; ".join(parts)

    if tool == "sitemap":
        if isinstance(value, dict):
            count = value.get("url_count", 0)
            return f"<b>{count:,}</b> URLs in sitemap"
        return str(value)

    if tool == "schema":
        if isinstance(value, list) and value:
            tags = "".join(
                f"<span style='display:inline-block;margin:2px;padding:1px 7px;background:#e8f4fd;"
                f"border-radius:10px;font-size:11px;color:#0969da'>{esc(s.get('format',''))}: {esc(s.get('type',''))}</span>"
                for s in value[:6])
            return tags
        return "<span style='color:#cf222e'>None found</span>"

    if tool == "hreflang":
        if isinstance(value, list) and value:
            langs = [e["lang"] for e in value]
            tags  = "".join(
                f"<span style='display:inline-block;margin:2px;padding:1px 7px;background:#f0fdf4;"
                f"border-radius:10px;font-size:11px;color:#166534'>{esc(lang)}</span>"
                for lang in langs[:8])
            return f"{len(langs)} tag(s): {tags}"
        return "<span style='color:#999'>None</span>"

    if tool == "ttfb":
        ms    = int(value) if value else 0
        color = "#1a7f37" if ms < 500 else "#9a6700" if ms < 800 else "#cf222e"
        bar_pct = min(100, round(ms / 1500 * 100))
        return (f"<b style='color:{color}'>{ms}ms</b>"
                f"<div style='margin-top:4px;display:flex;align-items:center;gap:6px'>"
                f"<div style='flex:1;height:4px;background:#e2e8f0;border-radius:9px;overflow:hidden'>"
                f"<div style='width:{bar_pct}%;height:100%;background:{color}'></div></div>"
                f"<span style='font-size:10px;color:#999'>target &lt;500ms</span></div>")

    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        return esc(value[:120])
    return "—"


# ── Fix hints ─────────────────────────────────────────────────────────────────
_FIX_HINTS: dict[str, list[tuple[str, str]]] = {
    "title": [
        ("Missing",     "Add a &lt;title&gt; tag between 30–60 chars with your primary keyword"),
        ("too short",   "Expand the title to 30–60 characters — include your target keyword"),
        ("too long",    "Trim title to under 60 chars so it doesn't get truncated in search results"),
    ],
    "meta_description": [
        ("Missing",     "Add a &lt;meta name='description'&gt; of 70–160 chars to improve click-through rate"),
        ("short",       "Expand meta description to 70–160 characters"),
        ("long",        "Trim meta description to under 160 characters"),
    ],
    "headings": [
        ("Missing H1",  "Add exactly one H1 tag describing the page's main topic"),
        ("Multiple H1", "Remove extra H1 tags — one H1 per page is the standard"),
        ("No H2",       "Add H2 subheadings to structure content for readers and search engines"),
    ],
    "image_alt": [
        ("missing alt", "Add descriptive alt text to every image for accessibility and image search"),
    ],
    "word_count": [
        ("Thin",        "Add more content — aim for at least 600 words to compete for rankings"),
        ("Below",       "Expand content to at least 600 words"),
    ],
    "canonical": [
        ("No canonical","Add &lt;link rel='canonical' href='https://...'&gt; in the &lt;head&gt;"),
        ("empty",       "Set the href attribute on your canonical tag to the preferred URL"),
        ("Canonical →", "Verify the canonical is intentional — it points search equity to another URL"),
    ],
    "broken_links": [
        ("broken",      "Fix or remove each broken link — they harm crawling and user experience"),
    ],
    "robots": [
        ("Blocked",     "Edit robots.txt to allow Googlebot — currently this page cannot be crawled"),
    ],
    "redirect": [
        ("chain",       "Collapse the redirect chain — redirect directly from source to final URL"),
        ("loop",        "Remove the redirect loop — check your server/CMS URL settings"),
    ],
    "ttfb": [
        ("Slow",        "Improve server response: add page caching, use a CDN, optimise DB queries"),
        ("Very slow",   "Critical: investigate server performance — TTFB over 800ms hurts Core Web Vitals"),
    ],
    "schema": [
        ("No structured","Add JSON-LD markup (Article, Product, FAQPage, etc.) to enable rich results"),
        ("error",       "Fix the JSON-LD syntax errors — use Google's Rich Results Test to validate"),
    ],
    "sitemap": [
        ("malformed",   "Regenerate or fix your sitemap XML — use an online validator"),
        ("duplicate",   "Remove duplicate URLs from your sitemap"),
        ("HTTP",        "Change HTTP URLs to HTTPS in your sitemap"),
        ("50,000",      "Split your sitemap into multiple files — each must be under 50,000 URLs"),
    ],
    "hreflang": [
        ("x-default",  "Add hreflang x-default pointing to the default language version"),
        ("Duplicate",  "Remove duplicate lang codes from your hreflang tags"),
    ],
    "http_status": [
        ("404",         "Page not found — fix the URL or set up a redirect to the correct page"),
        ("5",           "Server error — check your hosting logs and server configuration"),
        ("Redirect",    "This URL redirects — update internal links to point directly to the final URL"),
    ],
}

def get_fix_hint(r: dict) -> str:
    if r.get("status") == "pass":
        return ""
    hints = _FIX_HINTS.get(r.get("tool", ""), [])
    msg   = r.get("message", "")
    for key, hint in hints:
        if key.lower() in msg.lower():
            return hint
    return ""


def _sub_items(r: dict) -> str:
    """Render expandable detail rows for checks that have lists."""
    tool  = r.get("tool", "")
    value = r.get("value")
    html  = ""

    if tool == "broken_links" and isinstance(value, list) and value:
        def _bl_row(x):
            u = x.get("url","")
            s = x.get("status","")
            return (f"<div style='padding:3px 0 3px 16px;font-size:11px;color:#666'>"
                    f"<span style='color:#cf222e'>✗</span> "
                    f"<a href='{esc(u)}' target='_blank' style='color:#0969da'>{esc(u[:80])}</a>"
                    f" <span style='color:#999;margin-left:4px'>({esc(s)})</span></div>")
        items = "".join(_bl_row(x) for x in value[:10])
        if len(value) > 10:
            items += f"<div style='font-size:11px;color:#999;padding-left:16px'>+ {len(value)-10} more…</div>"
        html = f"<div style='border-top:1px solid #f0f2f5;margin-top:6px;padding-top:4px'>{items}</div>"

    elif tool == "image_alt" and isinstance(value, dict) and value.get("missing"):
        items = "".join(
            f"<div style='padding:2px 0 2px 16px;font-size:11px;color:#666'>"
            f"<span style='color:#9a6700'>⚠</span> {esc(src[:90])}</div>"
            for src in value["missing"][:8]
        )
        html = f"<div style='border-top:1px solid #f0f2f5;margin-top:6px;padding-top:4px'>{items}</div>"

    elif tool == "redirect" and isinstance(value, list) and len(value) > 1:
        chain = " &rarr; ".join(f"<code style='font-size:10px'>{esc(u[:60])}</code>" for u in value)
        html  = f"<div style='margin-top:6px;font-size:11px;color:#666;border-top:1px solid #f0f2f5;padding-top:4px'>{chain}</div>"

    elif tool == "headings" and isinstance(value, dict):
        h2s = value.get("h2", [])
        if h2s:
            items = "".join(
                f"<div style='padding:1px 0 1px 16px;font-size:11px;color:#666'>H2: {esc(h[:80])}</div>"
                for h in h2s[:5]
            )
            html = f"<div style='margin-top:6px;border-top:1px solid #f0f2f5;padding-top:4px'>{items}</div>"

    elif tool == "hreflang" and isinstance(value, list) and value:
        items = "".join(
            f"<div style='padding:1px 0 1px 16px;font-size:11px;color:#666'>"
            f"<b style='color:#333'>{esc(e['lang'])}</b> → {esc(e['url'][:70])}</div>"
            for e in value[:8]
        )
        html = f"<div style='margin-top:6px;border-top:1px solid #f0f2f5;padding-top:4px'>{items}</div>"

    elif tool == "schema" and isinstance(value, list):
        errors = (r.get("details") or {}).get("errors", [])
        if errors:
            items = "".join(
                f"<div style='padding:2px 0 2px 16px;font-size:11px;color:#cf222e'>✗ {esc(e[:100])}</div>"
                for e in errors[:5]
            )
            html = f"<div style='margin-top:6px;border-top:1px solid #f0f2f5;padding-top:4px'>{items}</div>"

    elif tool == "internal_links" and isinstance(value, dict):
        ext = value.get("external", [])
        if ext:
            items = "".join(
                f"<div style='padding:1px 0 1px 16px;font-size:11px;color:#666'>"
                f"<span style='color:#999'>↗</span> {esc(u[:70])}</div>"
                for u in ext[:5]
            )
            html = f"<div style='margin-top:6px;border-top:1px solid #f0f2f5;padding-top:4px'>{items}</div>"

    return html


# ══════════════════════════════════════════════════════════════════════════════
# HTML Report
# ══════════════════════════════════════════════════════════════════════════════
def score_color(score: int) -> str:
    if score >= 80: return "score-excellent"
    if score >= 50: return "score-good"
    return "score-poor"

def generate_html_report(audits: list[dict], p3_site_data: list[dict], timestamp: str) -> str:
    total_urls  = len(audits)
    avg_score   = round(sum(a["score"] for a in audits) / total_urls) if total_urls else 0
    total_fails  = sum(len(a["issues"]) for a in audits)
    total_warns  = sum(len(a["warnings"]) for a in audits)

    # Score distribution
    excellent = sum(1 for a in audits if a["score"] >= 80)
    good      = sum(1 for a in audits if 50 <= a["score"] < 80)
    poor      = sum(1 for a in audits if a["score"] < 50)

    # Issues frequency
    issue_freq: dict[str, int] = {}
    for a in audits:
        for r in a["issues"] + a["warnings"]:
            issue_freq[r["tool"]] = issue_freq.get(r["tool"], 0) + 1
    top_issues = sorted(issue_freq.items(), key=lambda x: x[1], reverse=True)[:10]

    # URL cards
    url_cards = ""
    for a in audits:
        sc    = a["score"]
        color = score_color(sc)
        rows_html = ""
        for r in a["results"]:
            icon       = STATUS_ICON.get(r["status"], "❓")
            bg         = STATUS_BG.get(r["status"], "#fff")
            txt_color  = STATUS_COLOR.get(r["status"], "#333")
            tool_name  = _pretty(r["tool"])
            fmtval     = format_value(r)
            hint       = get_fix_hint(r)
            sub        = _sub_items(r)
            hint_html  = f"<div class='fix-hint'>💡 {hint}</div>" if hint else ""
            rows_html += f"""
            <tr style="background:{bg}">
              <td style="vertical-align:top">
                <div class="check-tool">{icon} {esc(tool_name)}</div>
              </td>
              <td style="vertical-align:top">
                <div class="check-val">{fmtval}</div>
              </td>
              <td style="vertical-align:top">
                <span class="check-msg" style="color:{txt_color};font-weight:500">{esc(r.get('message',''))}</span>
                {hint_html}
                {sub}
              </td>
            </tr>"""

        fail_count = a['counts']['fail']
        url_cards += f"""
        <div class="url-card" data-score="{sc}" data-issues="{fail_count}" data-url="{esc(a['url'])}">
          <div class="url-header" onclick="toggle(this)">
            <div class="url-left">
              <div class="url-text"><a href="{esc(a['url'])}" target="_blank" onclick="event.stopPropagation()">{esc(a['url'])}</a></div>
              <div class="url-meta">
                <span class="url-meta-pill pill-pass">✓ {a['counts']['pass']} passed</span>
                {'<span class="url-meta-pill pill-warn">⚠ ' + str(a['counts']['warning']) + ' warn</span>' if a['counts']['warning'] else ''}
                {'<span class="url-meta-pill pill-fail">✕ ' + str(a['counts']['fail']) + ' issues</span>' if a['counts']['fail'] else ''}
              </div>
            </div>
            <div class="url-right">
              <div class="score-ring {color}">{sc}</div>
              <span class="expand-icon">▾</span>
            </div>
          </div>
          <div class="url-body">
            <table class="check-table" style="width:100%;border-collapse:collapse">
              <thead>
                <tr>
                  <th style="width:160px">Check</th>
                  <th style="width:240px">Value</th>
                  <th>Status &amp; Recommendation</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
          </div>
        </div>"""

    # Phase 3 site-wide data
    p3_html = ""
    for r in p3_site_data:
        if r["status"] == "error": continue
        if r["tool"] == "ctr_analyzer" and r["value"]:
            rows = "".join(f"<tr><td>{esc(x['url'][:60])}</td><td>{esc(x['impressions'])}</td><td>{esc(x['ctr'])}%</td><td>{esc(x['position'])}</td></tr>" for x in r["value"][:10])
            p3_html += f"""<div class="section"><h2>Low CTR Pages (top 10)</h2>
            <table><thead><tr><th>URL</th><th>Impressions</th><th>CTR</th><th>Position</th></tr></thead><tbody>{rows}</tbody></table></div>"""
        if r["tool"] == "sitemaps_status" and r["value"]:
            rows = "".join(f"<tr><td>{esc(x['path'])}</td><td>{esc(x['submitted'])}</td><td>{esc(x['indexed'])}</td><td>{esc(x['index_rate'])}%</td></tr>" for x in r["value"])
            p3_html += f"""<div class="section"><h2>Sitemaps Status</h2>
            <table><thead><tr><th>Sitemap</th><th>Submitted</th><th>Indexed</th><th>Index Rate</th></tr></thead><tbody>{rows}</tbody></table></div>"""

    _max_issue = top_issues[0][1] if top_issues else 1
    top_issues_html = "".join(
        f"<tr><td>{esc(_pretty(t))}</td><td>{c}</td>"
        f"<td><div style='width:120px;height:7px;background:#F1F5F9;border-radius:4px;overflow:hidden'>"
        f"<div style='width:{round(c/_max_issue*100)}%;height:100%;background:#EF4444;border-radius:4px'></div></div></td></tr>"
        for t, c in top_issues
    )

    total_pass = sum(a['counts']['pass'] for a in audits)
    total_warn_checks = sum(a['counts']['warning'] for a in audits)
    total_fail_checks = sum(a['counts']['fail'] for a in audits)
    score_color_cls = "excellent" if avg_score >= 80 else "good" if avg_score >= 50 else "poor"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>SEO Audit Report — {timestamp[:4]}-{timestamp[4:6]}-{timestamp[6:8]}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --primary:#6366F1;--success:#10B981;--warn:#F59E0B;--danger:#EF4444;
  --bg:#F5F7FB;--surface:#fff;--surface2:#F8FAFC;--border:#E5E9F2;
  --text:#0F172A;--text2:#475569;--text3:#94A3B8;
  --r:12px;--shadow:0 1px 3px rgba(15,23,42,.06),0 4px 12px rgba(15,23,42,.04);
}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--text);line-height:1.5}}
a{{color:var(--primary);text-decoration:none}}a:hover{{text-decoration:underline}}

/* ── Header ── */
.rpt-header{{
  background:linear-gradient(135deg,#6366F1 0%,#8B5CF6 100%);
  color:#fff;padding:28px 40px;
  display:flex;justify-content:space-between;align-items:center;gap:24px;flex-wrap:wrap;
}}
.rpt-header-left h1{{font-size:22px;font-weight:800;margin-bottom:4px}}
.rpt-header-left p{{font-size:13px;opacity:.8}}
.rpt-score-ring{{
  text-align:center;background:rgba(255,255,255,.15);border-radius:16px;
  padding:18px 28px;backdrop-filter:blur(8px);
}}
.rpt-score-n{{font-size:44px;font-weight:900;line-height:1}}
.rpt-score-n.excellent{{color:#86EFAC}}
.rpt-score-n.good{{color:#FDE68A}}
.rpt-score-n.poor{{color:#FCA5A5}}
.rpt-score-l{{font-size:12px;opacity:.75;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}}

/* ── Layout ── */
.container{{max-width:1300px;margin:28px auto;padding:0 28px}}
.stat-grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:14px;margin-bottom:24px}}
.stat-card{{
  background:var(--surface);border-radius:var(--r);padding:18px 16px;text-align:center;
  box-shadow:var(--shadow);border:1px solid var(--border);
}}
.stat-card .n{{font-size:30px;font-weight:800;line-height:1;margin-bottom:5px}}
.stat-card .l{{font-size:10.5px;color:var(--text3);text-transform:uppercase;letter-spacing:.5px;font-weight:600}}
.c-blue{{color:#3B82F6}}.c-purple{{color:#8B5CF6}}.c-green{{color:var(--success)}}
.c-yellow{{color:var(--warn)}}.c-red{{color:var(--danger)}}

.section{{background:var(--surface);border-radius:var(--r);padding:20px 22px;margin-bottom:20px;box-shadow:var(--shadow);border:1px solid var(--border)}}
.section h2{{font-size:14px;font-weight:700;color:var(--text);margin-bottom:14px;display:flex;align-items:center;gap:8px}}
.section h2 .badge{{font-size:11px;font-weight:600;padding:2px 8px;border-radius:20px;background:var(--surface2);color:var(--text2);border:1px solid var(--border)}}
.charts-row{{display:grid;grid-template-columns:1fr 1.4fr 1fr;gap:18px;margin-bottom:20px}}

/* ── Table ── */
table{{width:100%;border-collapse:collapse;font-size:12.5px}}
thead th{{
  background:var(--surface2);color:var(--text2);padding:9px 13px;
  text-align:left;font-size:11px;font-weight:700;text-transform:uppercase;
  letter-spacing:.4px;border-bottom:2px solid var(--border);
}}
td{{padding:9px 13px;border-bottom:1px solid var(--border);color:var(--text2)}}
tr:last-child td{{border-bottom:none}}
tbody tr:hover td{{background:var(--surface2)}}

/* ── URL cards ── */
.url-cards-wrap{{display:flex;flex-direction:column;gap:8px}}
.url-card{{background:var(--surface);border-radius:var(--r);box-shadow:var(--shadow);border:1px solid var(--border);overflow:hidden;transition:box-shadow .15s}}
.url-card:hover{{box-shadow:0 4px 16px rgba(99,102,241,.1)}}
.url-header{{
  display:flex;justify-content:space-between;align-items:center;
  padding:14px 16px;cursor:pointer;gap:12px;
}}
.url-header:hover{{background:var(--surface2)}}
.url-left{{flex:1;min-width:0}}
.url-text{{font-size:13px;font-weight:600;color:var(--primary);word-break:break-all;margin-bottom:3px}}
.url-meta{{font-size:11px;color:var(--text3);display:flex;gap:10px;flex-wrap:wrap}}
.url-meta-pill{{
  display:inline-flex;align-items:center;gap:3px;
  padding:2px 7px;border-radius:999px;font-size:10.5px;font-weight:600;
}}
.pill-pass{{background:#DCFCE7;color:#166534}}
.pill-warn{{background:#FEF3C7;color:#92400E}}
.pill-fail{{background:#FEE2E2;color:#991B1B}}
.url-right{{display:flex;align-items:center;gap:10px;flex-shrink:0}}
.score-ring{{
  width:46px;height:46px;border-radius:50%;display:flex;align-items:center;justify-content:center;
  font-size:16px;font-weight:800;color:#fff;flex-shrink:0;
}}
.score-excellent{{background:#1a7f37}}
.score-good{{background:#9a6700}}
.score-poor{{background:#cf222e}}
.expand-icon{{color:var(--text3);font-size:11px;transition:transform .2s}}
.url-card.open .expand-icon{{transform:rotate(180deg)}}
.url-body{{border-top:1px solid var(--border);display:none;overflow:hidden}}
.url-body.vis{{display:block}}
.check-table th{{background:var(--surface2)}}
.check-tool{{font-size:11.5px;font-weight:700;color:var(--text);white-space:nowrap;display:flex;align-items:center;gap:5px}}
.check-val{{font-size:12px;color:var(--text2);max-width:240px;word-break:break-word}}
.check-msg{{font-size:12px}}
.fix-hint{{
  margin-top:5px;padding:6px 10px;
  background:#FFFBEB;border-left:3px solid var(--warn);
  border-radius:0 6px 6px 0;font-size:11px;color:#78350F;line-height:1.4;
}}

/* ── Filter bar ── */
.filter-bar{{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap;align-items:center}}
.fbtn{{
  padding:6px 14px;border:1.5px solid var(--border);border-radius:999px;
  cursor:pointer;font-size:12px;font-weight:500;background:var(--surface);color:var(--text2);
  transition:all .15s;white-space:nowrap;
}}
.fbtn:hover{{border-color:var(--primary);color:var(--primary)}}
.fbtn.active{{background:var(--primary);border-color:var(--primary);color:#fff}}
.fbtn .fc{{
  display:inline-flex;align-items:center;justify-content:center;
  margin-left:5px;min-width:17px;height:17px;padding:0 4px;
  background:rgba(0,0,0,.12);border-radius:999px;font-size:10px;font-weight:700;
}}
.fbtn.active .fc{{background:rgba(255,255,255,.25)}}
.fsrch{{
  padding:7px 14px;border:1.5px solid var(--border);border-radius:999px;
  font-size:12px;background:var(--surface);color:var(--text);outline:none;min-width:200px;
  transition:border-color .15s;
}}
.fsrch:focus{{border-color:var(--primary)}}
.factions{{margin-left:auto;display:flex;gap:6px}}
.fact{{
  padding:6px 12px;border:1.5px solid var(--border);border-radius:999px;
  cursor:pointer;font-size:12px;font-weight:500;background:var(--surface);color:var(--text2);
}}
.fact:hover{{background:var(--surface2)}}

/* ── Sort controls ── */
.sort-bar{{display:flex;align-items:center;gap:8px;margin-bottom:12px;font-size:12px;color:var(--text3)}}
.sort-btn{{
  padding:4px 10px;border:1.5px solid var(--border);border-radius:8px;
  cursor:pointer;font-size:11px;font-weight:600;background:var(--surface);color:var(--text2);
}}
.sort-btn.active{{border-color:var(--primary);color:var(--primary)}}

/* ── Print ── */
@media print{{
  body{{background:#fff}}
  .rpt-header{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}
  .url-body{{display:block!important}}
  .filter-bar,.sort-bar,.factions,.fact,.fsrch{{display:none}}
  .url-card{{box-shadow:none;border:1px solid #ddd;break-inside:avoid}}
  .charts-row,.stat-grid{{break-inside:avoid}}
  canvas{{max-height:200px}}
}}
@media(max-width:900px){{
  .stat-grid{{grid-template-columns:repeat(3,1fr)}}
  .charts-row{{grid-template-columns:1fr}}
  .rpt-header{{padding:20px 20px}}
  .container{{padding:0 16px}}
}}
</style>
</head>
<body>

<div class="rpt-header">
  <div class="rpt-header-left">
    <h1>🔍 SEO Audit Report</h1>
    <p>Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')} · {total_urls} URL{'s' if total_urls!=1 else ''} audited</p>
  </div>
  <div class="rpt-score-ring">
    <div class="rpt-score-n {score_color_cls}">{avg_score}</div>
    <div class="rpt-score-l">Avg SEO Score / 100</div>
  </div>
</div>

<div class="container">
  <!-- Stat cards -->
  <div class="stat-grid">
    <div class="stat-card"><div class="n c-blue">{total_urls}</div><div class="l">URLs Audited</div></div>
    <div class="stat-card"><div class="n c-purple">{avg_score}</div><div class="l">Avg Score</div></div>
    <div class="stat-card"><div class="n c-green">{excellent}</div><div class="l">Excellent 80+</div></div>
    <div class="stat-card"><div class="n c-yellow">{total_warns}</div><div class="l">Warnings</div></div>
    <div class="stat-card"><div class="n c-red">{total_fails}</div><div class="l">Issues</div></div>
  </div>

  <!-- Charts -->
  <div class="charts-row">
    <div class="section" style="margin:0">
      <h2>Score Distribution</h2>
      <canvas id="dist" height="160"></canvas>
    </div>
    <div class="section" style="margin:0">
      <h2>Top Issues <span class="badge">by frequency</span></h2>
      <table>
        <thead><tr><th>Check</th><th style="width:50px">Count</th><th>Frequency</th></tr></thead>
        <tbody>{top_issues_html}</tbody>
      </table>
    </div>
    <div class="section" style="margin:0">
      <h2>Pass / Warn / Fail</h2>
      <canvas id="pie" height="160"></canvas>
    </div>
  </div>

  {p3_html}

  <!-- URL results -->
  <div class="section">
    <h2>URL-by-URL Results <span class="badge" id="visible-count">{total_urls} shown</span></h2>
    <div class="filter-bar">
      <button class="fbtn active" onclick="filter('all',this)">All<span class="fc">{total_urls}</span></button>
      <button class="fbtn" onclick="filter('excellent',this)">✅ 80+<span class="fc">{excellent}</span></button>
      <button class="fbtn" onclick="filter('good',this)">⚠ 50–79<span class="fc">{good}</span></button>
      <button class="fbtn" onclick="filter('poor',this)">❌ &lt;50<span class="fc">{poor}</span></button>
      <input class="fsrch" id="srch" type="text" placeholder="🔍  Filter by URL…" oninput="search()">
      <div class="factions">
        <button class="fact" onclick="expandAll(true)">Expand all</button>
        <button class="fact" onclick="expandAll(false)">Collapse all</button>
        <button class="fact" onclick="window.print()">🖨 Print</button>
      </div>
    </div>
    <div class="sort-bar">
      Sort:
      <button class="sort-btn active" id="sort-score-desc" onclick="sortCards('score','desc',this)">Score ↓</button>
      <button class="sort-btn" id="sort-score-asc" onclick="sortCards('score','asc',this)">Score ↑</button>
      <button class="sort-btn" id="sort-issues" onclick="sortCards('issues','desc',this)">Most Issues</button>
    </div>
    <div class="url-cards-wrap" id="cards-wrap">{url_cards}</div>
  </div>
</div>

<script>
new Chart(document.getElementById('dist'),{{
  type:'bar',
  data:{{
    labels:['Excellent (80+)','Good (50–79)','Poor (<50)'],
    datasets:[{{data:[{excellent},{good},{poor}],
      backgroundColor:['#10B981','#F59E0B','#EF4444'],
      borderRadius:6,borderSkipped:false}}]
  }},
  options:{{responsive:true,plugins:{{legend:{{display:false}}}},
    scales:{{y:{{beginAtZero:true,grid:{{color:'#F1F5F9'}}}},x:{{grid:{{display:false}}}}}}
  }}
}});
new Chart(document.getElementById('pie'),{{
  type:'doughnut',
  data:{{
    labels:['Pass','Warning','Fail'],
    datasets:[{{
      data:[{total_pass},{total_warn_checks},{total_fail_checks}],
      backgroundColor:['#10B981','#F59E0B','#EF4444'],
      borderWidth:2,borderColor:'#fff',hoverOffset:4
    }}]
  }},
  options:{{cutout:'62%',plugins:{{legend:{{position:'bottom',labels:{{padding:14,font:{{size:12}}}}}}}}}}
}});

// ── Toggle card open/closed ──
function toggle(el){{
  const card=el.closest('.url-card');
  const body=card.querySelector('.url-body');
  const open=card.classList.toggle('open');
  body.classList.toggle('vis',open);
}}

// ── Expand / collapse all ──
function expandAll(open){{
  document.querySelectorAll('.url-card:not([style*="display: none"])').forEach(c=>{{
    c.classList.toggle('open',open);
    c.querySelector('.url-body').classList.toggle('vis',open);
  }});
}}

// ── Filter by score bucket ──
let _fType='all';
function filter(type,btn){{
  _fType=type;
  document.querySelectorAll('.filter-bar .fbtn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  applyFilters();
}}

// ── Text search ──
function search(){{ applyFilters(); }}

function applyFilters(){{
  const q=(document.getElementById('srch').value||'').toLowerCase();
  let shown=0;
  document.querySelectorAll('.url-card').forEach(c=>{{
    const s=parseInt(c.dataset.score||0);
    const u=(c.dataset.url||'').toLowerCase();
    const scoreOk=_fType==='all'||(_fType==='excellent'&&s>=80)||(_fType==='good'&&s>=50&&s<80)||(_fType==='poor'&&s<50);
    const textOk=!q||u.includes(q);
    const show=scoreOk&&textOk;
    c.style.display=show?'':'none';
    if(show) shown++;
  }});
  const vc=document.getElementById('visible-count');
  if(vc) vc.textContent=shown+' shown';
}}

// ── Sort cards ──
function sortCards(field,dir,btn){{
  document.querySelectorAll('.sort-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  const wrap=document.getElementById('cards-wrap');
  const cards=[...wrap.querySelectorAll('.url-card')];
  cards.sort((a,b)=>{{
    let va=parseInt(a.dataset[field]||0), vb=parseInt(b.dataset[field]||0);
    return dir==='asc'?va-vb:vb-va;
  }});
  cards.forEach(c=>wrap.appendChild(c));
}}
</script>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════════
# Excel Report
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(audits: list[dict], path: Path):
    if not HAS_XLSX: return
    wb  = openpyxl.Workbook()
    ws  = wb.active; ws.title = "Summary"
    HDR = PatternFill("solid", fgColor="1F4E79")
    GF  = PatternFill("solid", fgColor="C6EFCE")
    RF  = PatternFill("solid", fgColor="FFC7CE")
    YF  = PatternFill("solid", fgColor="FFEB9C")

    ws.append(["#","URL","SEO Score","Pass","Warnings","Issues","Top Issue"])
    for cell in ws[1]: cell.fill=HDR; cell.font=Font(bold=True,color="FFFFFF")

    for i, a in enumerate(audits, 1):
        top_issue = a["issues"][0]["message"] if a["issues"] else (a["warnings"][0]["message"] if a["warnings"] else "")
        ws.append([i, a["url"], a["score"], a["counts"]["pass"], a["counts"]["warning"], a["counts"]["fail"], top_issue])
        sc = ws.cell(ws.max_row, 3)
        sc.fill = GF if a["score"]>=80 else YF if a["score"]>=50 else RF

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 50

    # Detailed sheet
    ws2 = wb.create_sheet("Details")
    ws2.append(["URL","Tool","Status","Message","Value"])
    for cell in ws2[1]: cell.fill=HDR; cell.font=Font(bold=True,color="FFFFFF")
    for a in audits:
        for r in a["results"]:
            ws2.append([a["url"], r["tool"], r["status"], r["message"], str(r.get("value",""))[:200]])
            c = ws2.cell(ws2.max_row, 3)
            c.fill = GF if r["status"]=="pass" else RF if r["status"]=="fail" else YF

    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
# Main CLI
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print(f"\n{BOLD}{CYAN}{'='*56}")
    print("            SEO AUDIT TOOL")
    print("  Phase 1 · Phase 2 · Phase 3 GSC · Phase 4 APIs")
    print(f"{'='*56}{RESET}\n")

    # Input
    print(f"{BOLD}Input type:{RESET}")
    print("  1. Sitemap URL")
    print("  2. Domain (auto-detect sitemap)")
    print("  3. CSV / Excel file")
    choice = input("Choice (1-3, default 1): ").strip() or "1"

    if choice == "1":
        sitemap = input("Sitemap URL: ").strip()
        if not sitemap:
            print(f"{RED}No URL entered.{RESET}"); return
        urls = fetch_sitemap_urls(sitemap)
    elif choice == "2":
        domain = input("Domain: ").strip()
        urls   = fetch_from_domain(domain)
    else:
        path = input("File path: ").strip()
        urls = load_from_csv_excel(path)

    if not urls:
        print(f"{RED}No URLs found.{RESET}"); sys.exit(1)

    print(f"\n{GREEN}Found {len(urls)} URLs.{RESET}")
    raw = input("How many to audit? (number or 'all'): ").strip()
    if raw.lower() != "all":
        try:
            urls = urls[:int(raw)]
        except ValueError:
            logger.warning("Invalid audit limit %r — defaulting to 10", raw)
            urls = urls[:10]

    pattern = input("Filter by pattern (e.g. /category/) or Enter to skip: ").strip()
    if pattern: urls = [u for u in urls if pattern in u]

    # GSC
    gsc_service = None
    if CFG.get("gsc", {}).get("enabled"):
        gsc_service = build_gsc_service()
        if gsc_service:
            print(f"{GREEN}GSC API connected ✓{RESET}")

    # Phase 3 site-wide
    p3_site_data = []
    if gsc_service and urls:
        from tools.phase3 import audit_site as p3_site
        _p = urlparse(urls[0])
        site_url     = f"{_p.scheme}://{_p.netloc}/" if _p.netloc else ""
        p3_site_data = p3_site(gsc_service, site_url, urls[:5])

    # Audit each URL
    print(f"\n{BOLD}Auditing {len(urls)} URLs…{RESET}\n")
    audits   = []
    iterator = tqdm(urls, unit="url") if HAS_TQDM else urls

    for url in iterator:
        if not HAS_TQDM:
            print(f"  Auditing: {url}")
        audit = audit_single_url(url, CFG, gsc_service)
        audits.append(audit)

        score = audit["score"]
        color = GREEN if score>=80 else YELLOW if score>=50 else RED
        if not HAS_TQDM:
            print(f"  {color}Score: {score}/100{RESET} | {len(audit['issues'])} issues | {len(audit['warnings'])} warnings\n")

    # Reports
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_path  = REPORTS_DIR / f"seo_audit_{timestamp}.html"
    excel_path = REPORTS_DIR / f"seo_audit_{timestamp}.xlsx"
    csv_path   = REPORTS_DIR / f"seo_audit_{timestamp}.csv"

    print(f"\n{BOLD}Generating reports…{RESET}")

    html_content = generate_html_report(audits, p3_site_data, timestamp)
    html_path.write_text(html_content, encoding="utf-8")
    print(f"  HTML  → {BOLD}{html_path}{RESET}")

    # JSON sidecar — lightweight summary for the dashboard home panel
    _avg = round(sum(a["score"] for a in audits) / len(audits)) if audits else 0
    _sidecar = {
        "avg_score":     _avg,
        "total_issues":  sum(len(a["issues"])   for a in audits),
        "total_warnings":sum(len(a["warnings"])  for a in audits),
        "urls":          len(audits),
    }
    (REPORTS_DIR / f"seo_audit_{timestamp}.json").write_text(
        json.dumps(_sidecar), encoding="utf-8"
    )

    generate_excel_report(audits, excel_path)
    print(f"  Excel → {BOLD}{excel_path}{RESET}")

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["URL","SEO Score","Pass","Warnings","Issues"])
        for a in audits:
            w.writerow([a["url"], a["score"], a["counts"]["pass"], a["counts"]["warning"], a["counts"]["fail"]])
    print(f"  CSV   → {BOLD}{csv_path}{RESET}")

    # Summary
    avg_score = round(sum(a["score"] for a in audits) / len(audits)) if audits else 0
    print(f"\n{BOLD}{CYAN}{'='*56}\n  SUMMARY\n{'='*56}{RESET}")
    print(f"  URLs audited   : {len(audits)}")
    print(f"  Avg SEO score  : {avg_score}/100")
    print(f"  {GREEN}Excellent 80+  : {sum(1 for a in audits if a['score']>=80)}{RESET}")
    print(f"  {YELLOW}Good 50-79     : {sum(1 for a in audits if 50<=a['score']<80)}{RESET}")
    print(f"  {RED}Poor <50       : {sum(1 for a in audits if a['score']<50)}{RESET}")
    print(f"\n  Open → {BOLD}{html_path}{RESET}\n")


if __name__ == "__main__":
    main()
